"""
api_app.py — CRM REST API (Phase 2)
=====================================
FastAPI REST backend untuk Nuxt 3 frontend.

Fitur:
  - JWT Bearer token authentication
  - CORS untuk localhost:3000 (Nuxt dev) dan domain produksi
  - Semua endpoint return JSON
  - Semua business logic sama dengan main.py (Jinja2 SSR)

Jalankan:
    uvicorn api_app:app --host 0.0.0.0 --port 8001 --reload

Endpoint utama:
  POST /api/v1/auth/login          → login, terima JWT token
  GET  /api/v1/auth/me             → info user saat ini
  GET  /api/v1/dashboard           → statistik dashboard
  CRUD /api/v1/pipeline            → leads pipeline
  GET  /api/v1/today               → aktivitas hari ini
  GET  /api/v1/schedule            → jadwal follow-up
  CRUD /api/v1/followup            → follow-up log
  GET  /api/v1/winloss             → win/loss records
  CRUD /api/v1/contacts            → contacts
  GET  /api/v1/insights            → pipeline insights
  GET  /api/v1/revenue/...         → semua modul revenue
  GET  /api/v1/roles               → master roles
  GET  /api/v1/users               → master users
  GET  /api/v1/sales               → master sales
"""

from fastapi import FastAPI, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List, Any
from jose import JWTError, jwt
from datetime import datetime, date, timedelta
from decimal import Decimal
import os
import random
import string

from database_pg import (
    get_conn, init_db, next_lead_id, next_fu_id,
    hash_pw, ALL_MENUS, get_user_menus
)

# ── Konstanta ─────────────────────────────────────────────────────────────────
SECRET_KEY  = os.getenv("CRM_JWT_SECRET", "crm-dcss-pkp-jwt-2026-super-secret-key")
ALGORITHM   = "HS256"
TOKEN_EXPIRE_HOURS = 24

MONTHS_EN = ['January','February','March','April','May','June',
             'July','August','September','October','November','December']

STALE_HOT  = 7
STALE_WARM = 14
STALE_ANY  = 30

# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(d: dict) -> dict:
    """Konversi datetime.date / Decimal ke tipe JSON-serializable."""
    out = {}
    for k, v in d.items():
        if isinstance(v, (date, datetime)):
            out[k] = v.isoformat()
        elif isinstance(v, Decimal):
            out[k] = float(v)
        else:
            out[k] = v
    return out


def compute_stale_flag(prioritas: str, days) -> str:
    if days is None:
        return "OK"
    days = float(days)
    if (prioritas or '').lower() == 'hot'  and days > STALE_HOT:
        return "URGENT"
    if (prioritas or '').lower() == 'warm' and days > STALE_WARM:
        return "WARNING"
    if days > STALE_ANY:
        return "STALE"
    return "OK"


def auto_status_risk(revenue_target, actual_revenue, type_: str,
                     tahun: int, target_month: str = None):
    revenue_target = float(revenue_target or 0)
    actual_revenue = float(actual_revenue or 0)
    if not revenue_target or revenue_target <= 0:
        return "Critical", "LOW"
    ach = actual_revenue / revenue_target
    today = date.today()
    if today.year == tahun:
        months_elapsed = today.month
    elif today.year > tahun:
        months_elapsed = 12
    else:
        months_elapsed = 0
    pace = months_elapsed / 12

    if pace == 0:
        status = "Critical" if ach == 0 else "On Track"
    elif ach >= max(pace * 0.9, 0.95):
        status = "On Track"
    elif ach >= pace * 0.5:
        status = "At Risk"
    else:
        status = "Critical"
    if ach >= 1.0:
        status = "On Track"

    type_lower = (type_ or '').lower()
    gap = pace - ach
    if 'termin' in type_lower:
        risk = "HIGH" if actual_revenue == 0 else ("HIGH" if gap > 0.3 else "MEDIUM")
    elif 'tahunan' in type_lower:
        risk = "HIGH" if actual_revenue == 0 else ("MEDIUM" if gap > 0.2 else "LOW")
    elif 'bulanan' in type_lower:
        risk = "LOW" if ach >= 0.7 else ("MEDIUM" if ach >= 0.4 else "HIGH")
    elif 'one time' in type_lower:
        risk = "HIGH" if actual_revenue == 0 else "LOW"
    else:
        risk = "HIGH" if ach < 0.3 else ("MEDIUM" if ach < 0.7 else "LOW")
    return status, risk


def sync_project_status(project_id: str):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM revenue_projects WHERE project_id=%s", (project_id,))
    p = c.fetchone()
    if not p:
        conn.close(); return
    p = dict(p)
    yr = int(p.get('tahun') or date.today().year)

    # Total paid dari semua invoice proyek ini
    c.execute("SELECT COALESCE(SUM(paid_amount),0) AS total FROM invoices WHERE project_id=%s", (project_id,))
    actual = float(c.fetchone()['total'] or 0)

    # Update revenue_projects
    st, rk = auto_status_risk(float(p['revenue_target'] or 0), actual,
                               p['type'], yr, p.get('target_month'))
    c.execute("""UPDATE revenue_projects SET actual_revenue=%s, status=%s, risk_level=%s,
        updated_at=NOW() WHERE project_id=%s""", (actual, st, rk, project_id))

    # Sync revenue_monthly.actual per bulan dari paid invoices
    c.execute("""
        SELECT EXTRACT(MONTH FROM invoice_date)::int AS month_num,
               COALESCE(SUM(paid_amount), 0) AS paid
        FROM invoices
        WHERE project_id=%s AND tahun=%s AND invoice_date IS NOT NULL
        GROUP BY 1
    """, (project_id, yr))
    paid_by_month = {r['month_num']: float(r['paid'] or 0) for r in c.fetchall()}

    # Update setiap baris revenue_monthly yang ada untuk proyek ini
    c.execute("SELECT id, month_num FROM revenue_monthly WHERE project_id=%s", (project_id,))
    for row in c.fetchall():
        mn = row['month_num']
        paid = paid_by_month.get(mn, 0)
        c.execute("UPDATE revenue_monthly SET actual=%s WHERE id=%s", (paid, row['id']))

    conn.commit(); conn.close()


def next_rev_id() -> str:
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT project_id FROM revenue_projects ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    if row:
        try:
            num = int(row['project_id'].replace('REV-', '')) + 1
        except:
            num = 9000
    else:
        num = 1
    return f"REV-{num:04d}"


def get_sales_list():
    """Ambil daftar nama sales dari tabel users (bukan tabel sales terpisah)."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT u.nama FROM users u LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.nama")
    names = [r['nama'] for r in c.fetchall()]
    conn.close()
    return names


# ── JWT helpers ───────────────────────────────────────────────────────────────

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/v1/auth/login")


def create_token(data: dict) -> str:
    payload = data.copy()
    payload["exp"] = datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token tidak valid atau sudah kadaluarsa.",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: int = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.*, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id = r.id
                 WHERE u.id=%s AND u.is_active=1""", (int(user_id),))
    row = c.fetchone()
    conn.close()
    if not row:
        raise credentials_exception
    return _norm(dict(row))


def require_admin(user: dict = Depends(get_current_user)):
    if user.get("role_id") != 1:
        raise HTTPException(status_code=403, detail="Hanya admin yang dapat melakukan tindakan ini")
    return user


def require_menu(menu_key: str):
    """Dependency factory — pastikan user punya akses ke menu tertentu."""
    def _check(user: dict = Depends(get_current_user)):
        allowed = get_user_menus(user["role_id"])
        if menu_key not in allowed:
            raise HTTPException(status_code=403, detail=f"Akses ditolak: menu '{menu_key}'")
        return user
    return _check


# ── App ───────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="CRM Leads Tracker — REST API",
    version="2.0.0",
    description="REST API untuk CRM DCSS — dikonsumsi oleh Nuxt 3 frontend",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    openapi_url="/api/openapi.json",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3000",
        os.getenv("CRM_FRONTEND_URL", "http://localhost:3000"),
    ],
    allow_origin_regex=r"http://localhost:\d+",  # semua port localhost (dev)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

init_db()

# Serve uploaded files at /storage/*
_static_dir = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(_static_dir, exist_ok=True)
app.mount("/storage", StaticFiles(directory=_static_dir, follow_symlink=True), name="storage")

# Serve Laravel public uploads (foto dari mobile apex) at /laravel-uploads/*
_laravel_uploads = os.path.join(os.path.dirname(__file__), "laravel-api", "public", "uploads")
if os.path.isdir(_laravel_uploads):
    app.mount("/laravel-uploads", StaticFiles(directory=_laravel_uploads, follow_symlink=True), name="laravel_uploads")

# ═══════════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════════

class LoginRequest(BaseModel):
    email: str
    password: str


@app.post("/api/v1/auth/login", tags=["Auth"])
def login(form: LoginRequest):
    """Login → terima JWT token."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.*, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id = r.id
                 WHERE u.email=%s AND u.password=%s AND u.is_active=1""",
              (form.email, hash_pw(form.password)))
    row = c.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=401, detail="Email atau password salah.")
    user = _norm(dict(row))
    token = create_token({"sub": str(user["id"]), "email": user["email"],
                          "role_id": user["role_id"]})
    allowed_menus = list(get_user_menus(user["role_id"]))
    return {
        "access_token": token,
        "token_type": "bearer",
        "expires_in": TOKEN_EXPIRE_HOURS * 3600,
        "user": {
            "id": user["id"], "nama": user["nama"], "email": user["email"],
            "role_id": user["role_id"], "role_nama": user["role_nama"],
            "allowed_menus": allowed_menus,
        }
    }


class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    email: str
    otp: str
    new_password: str

def _ensure_otp_table():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_otps (
            id         SERIAL PRIMARY KEY,
            email      VARCHAR(255) NOT NULL,
            otp        VARCHAR(10)  NOT NULL,
            expires_at TIMESTAMP    NOT NULL,
            used       BOOLEAN      DEFAULT FALSE,
            created_at TIMESTAMP    DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

_ensure_otp_table()

@app.post("/api/v1/auth/forgot-password", tags=["Auth"])
def forgot_password(req: ForgotPasswordRequest):
    """Generate OTP reset password — OTP ditampilkan di response untuk diteruskan admin ke user."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, email, nama FROM users WHERE email=%s AND is_active=1", (req.email,))
    user = c.fetchone()
    if not user:
        # Respons sama agar tidak bisa dipakai untuk enumerasi email
        conn.close()
        return {"message": "Jika email terdaftar, OTP telah dikirim."}

    otp = ''.join(random.choices(string.digits, k=6))
    expires = datetime.utcnow() + timedelta(minutes=15)

    # Hapus OTP lama untuk email ini
    c.execute("DELETE FROM password_reset_otps WHERE email=%s", (req.email,))
    c.execute(
        "INSERT INTO password_reset_otps (email, otp, expires_at) VALUES (%s,%s,%s)",
        (req.email, otp, expires)
    )
    conn.commit()
    conn.close()

    print(f"\n[RESET PASSWORD] Email: {req.email} | OTP: {otp} | Expires: {expires} WIB\n")
    return {
        "message": "Jika email terdaftar, OTP telah dikirim.",
        "debug_otp": otp,  # Hanya untuk development — hapus di production
    }


@app.post("/api/v1/auth/reset-password", tags=["Auth"])
def reset_password(req: ResetPasswordRequest):
    """Reset password menggunakan OTP yang valid."""
    if len(req.new_password) < 6:
        raise HTTPException(status_code=422, detail="Password minimal 6 karakter.")

    conn = get_conn()
    c = conn.cursor()
    c.execute(
        """SELECT id FROM password_reset_otps
           WHERE email=%s AND otp=%s AND used=FALSE AND expires_at > NOW()""",
        (req.email, req.otp)
    )
    row = c.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=400, detail="OTP tidak valid atau sudah kadaluarsa.")

    c.execute("UPDATE users SET password=%s WHERE email=%s", (hash_pw(req.new_password), req.email))
    c.execute("UPDATE password_reset_otps SET used=TRUE WHERE id=%s", (row[0],))
    conn.commit()
    conn.close()
    return {"message": "Password berhasil direset. Silakan login."}


@app.get("/api/v1/auth/me", tags=["Auth"])
def me(user: dict = Depends(get_current_user)):
    """Info user yang sedang login + daftar menu yang diizinkan."""
    allowed = list(get_user_menus(user["role_id"]))
    nav_menus = [m for m in ALL_MENUS if m["key"] in allowed]
    return {
        "id": user["id"], "nama": user["nama"], "email": user["email"],
        "role_id": user["role_id"], "role_nama": user["role_nama"],
        "allowed_menus": allowed,
        "nav_menus": nav_menus,
    }


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


@app.put("/api/v1/auth/change-password", tags=["Auth"])
def change_password(req: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    """Ganti password user yang sedang login."""
    if len(req.new_password) < 6:
        raise HTTPException(status_code=422, detail="Password baru minimal 6 karakter.")

    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT password FROM users WHERE id=%s", (user["id"],))
    row = c.fetchone()
    if not row or row["password"] != hash_pw(req.current_password):
        conn.close()
        raise HTTPException(status_code=400, detail="Password lama tidak sesuai.")

    c.execute("UPDATE users SET password=%s WHERE id=%s", (hash_pw(req.new_password), user["id"]))
    conn.commit()
    conn.close()
    return {"message": "Password berhasil diubah."}


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/dashboard", tags=["Dashboard"])
def dashboard(user: dict = Depends(require_menu("dashboard"))):
    conn = get_conn()
    c = conn.cursor()
    today = date.today()
    cur_year  = today.year
    is_sales  = user.get("role_id") == 3
    sf  = " AND sales_owner=%s" if is_sales else ""    # append to existing WHERE
    sw  = " WHERE sales_owner=%s" if is_sales else ""  # standalone WHERE
    sp  = [user["nama"]] if is_sales else []

    # ── Stats aggregate ──────────────────────────────────────────────────────
    c.execute(f"""
        SELECT
            COUNT(*)                                                          as total,
            COUNT(*) FILTER (WHERE stage='Won')                              as won,
            COUNT(*) FILTER (WHERE stage='Lost')                             as lost,
            COUNT(*) FILTER (WHERE stage='On Hold')                         as on_hold,
            COUNT(*) FILTER (WHERE stage NOT IN ('Won','Lost','On Hold'))    as aktif,
            COUNT(*) FILTER (WHERE (sales_owner IS NULL OR sales_owner='')
                              AND stage NOT IN ('Won','Lost'))                as unassigned,
            COUNT(*) FILTER (WHERE
                COALESCE((CURRENT_DATE-last_fu_date)::int,(CURRENT_DATE-tgl_masuk)::int,999) > 30
                AND stage NOT IN ('Won','Lost'))                              as stale,
            COUNT(*) FILTER (WHERE next_fu_date < CURRENT_DATE
                              AND stage NOT IN ('Won','Lost'))                as overdue_fu_count,
            COALESCE(SUM(propose_value) FILTER (WHERE stage NOT IN ('Won','Lost')), 0) as total_pipeline,
            COALESCE(SUM(propose_value) FILTER (WHERE stage NOT IN ('Won','Lost','On Hold')), 0) as active_pipeline,
            COALESCE(SUM(weighted_value) FILTER (WHERE stage NOT IN ('Won','Lost')), 0) as weighted_pipeline,
            COALESCE(SUM(COALESCE(NULLIF(deal_value,0), propose_value)) FILTER (WHERE stage='Won'), 0) as total_won,
            COALESCE(SUM(propose_value) FILTER (WHERE stage='On Hold'), 0)   as onhold_value
        FROM leads{sw}
    """, sp)
    agg = _norm(dict(c.fetchone()))
    stats = {
        "total":             agg["total"],
        "won":               agg["won"],
        "lost":              agg["lost"],
        "on_hold":           agg["on_hold"],
        "aktif":             agg["aktif"],
        "unassigned":        agg["unassigned"],
        "stale":             agg["stale"],
        "overdue_fu":        agg["overdue_fu_count"],
        "total_pipeline":    float(agg["total_pipeline"]),
        "active_pipeline":   float(agg["active_pipeline"]),
        "weighted_pipeline": float(agg["weighted_pipeline"]),
        "total_won":         float(agg["total_won"]),
        "onhold_value":      float(agg["onhold_value"]),
    }

    # ── Health Score ─────────────────────────────────────────────────────────
    if agg["total"] == 0:
        health_score = 0
    else:
        total = agg["total"]
        won_rate   = agg["won"] / total * 100
        stale_rate = agg["stale"] / total * 100
        unasgn_rate= agg["unassigned"] / total * 100
        health_score = max(0, min(100, int(
            won_rate * 0.4 +
            (100 - stale_rate) * 0.35 +
            (100 - unasgn_rate) * 0.25
        )))

    # ── Revenue bulan ini ────────────────────────────────────────────────────
    rev_filter = " AND p.pic=%s" if is_sales else ""
    c.execute(f"""
        SELECT COALESCE(SUM(m.actual), 0) as actual,
               COALESCE(SUM(m.target), 0) as target
        FROM revenue_monthly m
        JOIN revenue_projects p ON m.project_id = p.project_id
        WHERE p.is_active = 1 AND p.tahun = %s AND m.month_num = %s{rev_filter}
    """, [cur_year, today.month] + sp)
    rev_row = c.fetchone()
    rev_actual = float(rev_row["actual"] or 0)
    rev_target = float(rev_row["target"] or 0)
    rev_ach = round(rev_actual / rev_target * 100) if rev_target > 0 else 0

    # ── By Stage ─────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT stage,
               COUNT(*) as jumlah,
               COALESCE(SUM(propose_value),0) as total_nilai
        FROM leads{sw} GROUP BY stage ORDER BY jumlah DESC
    """, sp)
    by_stage = [_norm(dict(r)) for r in c.fetchall()]

    # ── By Segmen ────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(segmen,'—') as segmen,
               COUNT(*) as jumlah,
               COALESCE(SUM(propose_value),0) as total_nilai
        FROM leads{sw} GROUP BY segmen ORDER BY jumlah DESC
    """, sp)
    by_segmen = [_norm(dict(r)) for r in c.fetchall()]

    # ── By Source ────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(source,'—') as source, COUNT(*) as jumlah
        FROM leads{sw} GROUP BY source ORDER BY jumlah DESC
    """, sp)
    by_source = [dict(r) for r in c.fetchall()]

    # ── By Priority ──────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(prioritas,'—') as prioritas,
               COUNT(*) as jumlah,
               COALESCE(SUM(propose_value),0) as total_nilai
        FROM leads WHERE stage NOT IN ('Won','Lost'){sf}
        GROUP BY prioritas ORDER BY jumlah DESC
    """, sp)
    by_priority = [_norm(dict(r)) for r in c.fetchall()]

    # ── By Organisasi ────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(organisasi,'—') as organisasi,
               COUNT(*) as jumlah,
               COALESCE(SUM(propose_value),0) as total_nilai,
               COUNT(*) FILTER (WHERE stage NOT IN ('Won','Lost')) as aktif,
               COUNT(*) FILTER (WHERE stage='Won') as won
        FROM leads WHERE organisasi IS NOT NULL AND organisasi != ''{sf}
        GROUP BY organisasi ORDER BY jumlah DESC LIMIT 12
    """, sp)
    by_organisasi = [_norm(dict(r)) for r in c.fetchall()]

    # ── Upcoming FU (7 hari ke depan) ───────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, next_fu_date, prioritas, sales_owner
        FROM leads
        WHERE next_fu_date BETWEEN CURRENT_DATE AND CURRENT_DATE + 7
          AND stage NOT IN ('Won','Lost'){sf}
        ORDER BY next_fu_date ASC LIMIT 10
    """, sp)
    upcoming_fu = [_norm(dict(r)) for r in c.fetchall()]

    # ── Ready to Close ────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, organisasi, prioritas, propose_value, probability, sales_owner
        FROM leads
        WHERE stage IN ('Proposal Sent','Negotiation'){sf}
        ORDER BY propose_value DESC NULLS LAST LIMIT 10
    """, sp)
    ready_to_close = [_norm(dict(r)) for r in c.fetchall()]

    # ── Hot Stale ─────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, organisasi, propose_value, sales_owner,
               COALESCE((CURRENT_DATE-last_fu_date)::int,
                        (CURRENT_DATE-tgl_masuk)::int, 9999) as days_since_fu
        FROM leads
        WHERE prioritas='Hot' AND stage NOT IN ('Won','Lost'){sf}
          AND COALESCE((CURRENT_DATE-last_fu_date)::int,
                       (CURRENT_DATE-tgl_masuk)::int, 9999) > 14
        ORDER BY days_since_fu DESC LIMIT 10
    """, sp)
    hot_stale = [_norm(dict(r)) for r in c.fetchall()]

    # ── Closing Soon (60 hari ke depan) ──────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, organisasi, propose_value,
               (exp_close_date - CURRENT_DATE)::int as days_until_close
        FROM leads
        WHERE exp_close_date BETWEEN CURRENT_DATE AND CURRENT_DATE + 60
          AND stage NOT IN ('Won','Lost'){sf}
        ORDER BY exp_close_date ASC LIMIT 10
    """, sp)
    closing_soon = [_norm(dict(r)) for r in c.fetchall()]

    # ── On Hold at Risk (> 30 hari idle) ─────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, organisasi, propose_value, sales_owner,
               COALESCE((CURRENT_DATE-last_fu_date)::int,
                        (CURRENT_DATE-tgl_masuk)::int, 9999) as days_idle
        FROM leads
        WHERE stage='On Hold'{sf}
          AND COALESCE((CURRENT_DATE-last_fu_date)::int,
                       (CURRENT_DATE-tgl_masuk)::int, 9999) > 30
        ORDER BY days_idle DESC LIMIT 10
    """, sp)
    onhold_at_risk = [_norm(dict(r)) for r in c.fetchall()]

    # ── Overdue FU (list preview — full data via /dashboard/overdue-fu) ──────
    c.execute(f"""
        SELECT lead_id, nama_company, prioritas, next_fu_date,
               (CURRENT_DATE - next_fu_date)::int as days_overdue
        FROM leads
        WHERE next_fu_date < CURRENT_DATE AND stage NOT IN ('Won','Lost'){sf}
        ORDER BY next_fu_date ASC LIMIT 10
    """, sp)
    overdue_fu = [_norm(dict(r)) for r in c.fetchall()]

    # ── Recent Activity ───────────────────────────────────────────────────────
    c.execute(f"""
        SELECT f.lead_id, l.nama_company, f.tgl_fu, f.metode_fu, f.hasil_fu, f.catatan_fu
        FROM follow_up_log f
        LEFT JOIN leads l ON f.lead_id = l.lead_id
        WHERE 1=1{" AND l.sales_owner=%s" if is_sales else ""}
        ORDER BY f.tgl_fu DESC, f.created_at DESC LIMIT 15
    """, sp)
    recent_activity = [_norm(dict(r)) for r in c.fetchall()]

    conn.close()
    return {
        "stats":          stats,
        "health_score":   health_score,
        "rev_actual":     rev_actual,
        "rev_target":     rev_target,
        "rev_ach":        rev_ach,
        "by_stage":       by_stage,
        "by_segmen":      by_segmen,
        "by_source":      by_source,
        "by_priority":    by_priority,
        "by_organisasi":  by_organisasi,
        "upcoming_fu":    upcoming_fu,
        "ready_to_close": ready_to_close,
        "hot_stale":      hot_stale,
        "closing_soon":   closing_soon,
        "onhold_at_risk": onhold_at_risk,
        "overdue_fu":     overdue_fu,
        "recent_activity":recent_activity,
    }


@app.get("/api/v1/dashboard/overdue-fu", tags=["Dashboard"])
def dashboard_overdue_fu(
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=50),
    user: dict = Depends(require_menu("dashboard"))
):
    import math as _m
    conn = get_conn(); c = conn.cursor()
    is_sales = user.get("role_id") == 3
    sf = " AND sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []
    c.execute(
        f"SELECT COUNT(*) cnt FROM leads WHERE next_fu_date < CURRENT_DATE AND stage NOT IN ('Won','Lost'){sf}",
        sp
    )
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(f"""
        SELECT lead_id, nama_company, prioritas, sales_owner, next_fu_date,
               (CURRENT_DATE - next_fu_date)::int as days_overdue
        FROM leads
        WHERE next_fu_date < CURRENT_DATE AND stage NOT IN ('Won','Lost'){sf}
        ORDER BY next_fu_date ASC LIMIT %s OFFSET %s
    """, sp + [per_page, offset])
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"data": rows, "total": total, "page": page, "per_page": per_page,
            "total_pages": _m.ceil(total / per_page) if total else 1}


@app.get("/api/v1/dashboard/recent-activity", tags=["Dashboard"])
def dashboard_recent_activity(
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=50),
    user: dict = Depends(require_menu("dashboard"))
):
    import math as _m
    conn = get_conn(); c = conn.cursor()
    is_sales = user.get("role_id") == 3
    sf = " AND l.sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []
    c.execute(
        f"SELECT COUNT(*) cnt FROM follow_up_log f LEFT JOIN leads l ON f.lead_id = l.lead_id WHERE 1=1{sf}",
        sp
    )
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(f"""
        SELECT f.lead_id, l.nama_company, f.tgl_fu, f.metode_fu, f.hasil_fu, f.catatan_fu
        FROM follow_up_log f
        LEFT JOIN leads l ON f.lead_id = l.lead_id
        WHERE 1=1{sf}
        ORDER BY f.tgl_fu DESC, f.created_at DESC LIMIT %s OFFSET %s
    """, sp + [per_page, offset])
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"data": rows, "total": total, "page": page, "per_page": per_page,
            "total_pages": _m.ceil(total / per_page) if total else 1}


# ═══════════════════════════════════════════════════════════════════════════════
# PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

class LeadCreate(BaseModel):
    nama_company: str
    product: str = ""
    contact_person: str = ""
    segmen: str = ""
    sub_segmen: str = ""
    source: str = ""
    stage: str = "New"
    prioritas: str = "Warm"
    tgl_masuk: str = ""
    propose_value: float = 0
    deal_value: float = 0
    probability: float = 0
    exp_close_date: str = ""
    sales_owner: str = ""
    organisasi: str = ""
    next_fu_date: str = ""
    remarks: str = ""
    loss_reason: str = ""


class LeadUpdate(LeadCreate):
    pass


@app.get("/api/v1/pipeline", tags=["Pipeline"])
def pipeline_list(
    stage: str = "", segmen: str = "", search: str = "", sales: str = "",
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=100),
    user: dict = Depends(require_menu("pipeline"))
):
    conn = get_conn()
    c = conn.cursor()
    is_sales = user.get("role_id") == 3
    base = "FROM leads WHERE 1=1"
    params: list = []
    if is_sales:
        base += " AND sales_owner=%s"; params.append(user["nama"])
    else:
        if sales: base += " AND sales_owner=%s"; params.append(sales)
    if stage:  base += " AND stage=%s";       params.append(stage)
    if segmen: base += " AND segmen=%s";      params.append(segmen)
    if search: base += " AND (nama_company LIKE %s OR product LIKE %s OR contact_person LIKE %s)"; params += [f"%{search}%"]*3

    c.execute(f"SELECT COUNT(*) cnt {base}", params)
    total = int(c.fetchone()["cnt"])

    offset = (page - 1) * per_page
    c.execute(f"""SELECT *,
        COALESCE((CURRENT_DATE-last_fu_date)::integer,(CURRENT_DATE-tgl_masuk)::integer,999) as days_without_fu
        {base} ORDER BY id DESC LIMIT %s OFFSET %s""", params + [per_page, offset])
    leads = [_norm(dict(r)) for r in c.fetchall()]
    for l in leads:
        l["stale_flag"] = compute_stale_flag(l["prioritas"], l["days_without_fu"]) \
                          if l["stage"] not in ("Won","Lost") else "OK"
    conn.close()
    import math
    return {
        "leads": leads, "total": total,
        "page": page, "per_page": per_page,
        "total_pages": math.ceil(total / per_page) if total else 1,
        "sales_list": get_sales_list(),
    }


@app.post("/api/v1/pipeline", status_code=201, tags=["Pipeline"])
def pipeline_create(payload: LeadCreate, user: dict = Depends(require_menu("pipeline"))):
    conn = get_conn()
    c = conn.cursor()
    pv = float(payload.propose_value or 0)
    dv = float(payload.deal_value or 0)
    prob = float(payload.probability or 0)
    wv = dv * (prob/100) if dv else pv * (prob/100)
    c.execute("""INSERT INTO leads
        (lead_id,nama_company,product,contact_person,segmen,sub_segmen,source,stage,prioritas,
         tgl_masuk,propose_value,deal_value,probability,exp_close_date,weighted_value,
         sales_owner,next_fu_date,remarks,loss_reason)
        VALUES ('TEMP',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (payload.nama_company, payload.product, payload.contact_person,
         payload.segmen, payload.sub_segmen, payload.source, payload.stage, payload.prioritas,
         payload.tgl_masuk or str(date.today()), pv, dv, prob,
         payload.exp_close_date or None, wv,
         payload.sales_owner, payload.next_fu_date or None, payload.remarks,
         payload.loss_reason or None))
    new_id = c.fetchone()['id']
    lid = f"LD-{new_id:04d}"
    c.execute("UPDATE leads SET lead_id=%s WHERE id=%s", (lid, new_id))
    conn.commit(); conn.close()
    return {"message": "Lead berhasil dibuat.", "lead_id": lid}


@app.get("/api/v1/pipeline/forecast", tags=["Pipeline"])
def pipeline_forecast(tahun: int = 0, user: dict = Depends(require_menu("pipeline"))):
    from datetime import date as _date
    cur_year = tahun or _date.today().year
    is_sales = user.get("role_id") == 3
    nama     = user["nama"] if is_sales else None

    conn = get_conn()
    c = conn.cursor()

    def _params(*extra): return (cur_year, nama) + extra if is_sales else (cur_year,) + extra
    owner_sql = "AND sales_owner = %s" if is_sales else ""

    c.execute(f"""
        SELECT
            EXTRACT(MONTH FROM exp_close_date)::integer as bulan,
            TO_CHAR(exp_close_date, 'Mon') as bulan_label,
            COUNT(*) as jumlah_lead,
            COALESCE(SUM(propose_value), 0) as total_propose,
            COALESCE(SUM(weighted_value), 0) as total_weighted,
            COALESCE(SUM(CASE WHEN stage='Won' THEN deal_value ELSE 0 END), 0) as actual_won
        FROM leads
        WHERE EXTRACT(YEAR FROM exp_close_date) = %s
          AND exp_close_date IS NOT NULL
          AND stage NOT IN ('Lost')
          {owner_sql}
        GROUP BY 1, 2 ORDER BY 1
    """, _params())
    monthly = [dict(r) for r in c.fetchall()]

    c.execute(f"""
        SELECT
            COALESCE(SUM(weighted_value), 0) as total_weighted,
            COALESCE(SUM(CASE WHEN stage='Won' THEN deal_value ELSE 0 END), 0) as total_won,
            COUNT(*) as total_leads,
            COUNT(CASE WHEN stage='Won' THEN 1 END) as total_won_count
        FROM leads
        WHERE EXTRACT(YEAR FROM exp_close_date) = %s
          AND exp_close_date IS NOT NULL
          {owner_sql}
    """, _params())
    summary = dict(c.fetchone() or {})

    c.execute(f"""
        SELECT
            sales_owner,
            COUNT(*) as jumlah_lead,
            COALESCE(SUM(weighted_value), 0) as total_weighted,
            COALESCE(SUM(CASE WHEN stage='Won' THEN deal_value ELSE 0 END), 0) as actual_won,
            COALESCE(AVG(probability), 0) as avg_probability
        FROM leads
        WHERE EXTRACT(YEAR FROM exp_close_date) = %s
          AND exp_close_date IS NOT NULL
          AND stage NOT IN ('Lost')
          AND sales_owner IS NOT NULL AND sales_owner <> ''
          {owner_sql}
        GROUP BY sales_owner ORDER BY total_weighted DESC
    """, _params())
    by_sales = [dict(r) for r in c.fetchall()]

    c.execute(f"""
        SELECT
            COALESCE(NULLIF(loss_reason, ''), 'Tidak dicatat') as reason,
            COUNT(*) as jumlah,
            COALESCE(SUM(propose_value), 0) as nilai_hilang
        FROM leads
        WHERE stage = 'Lost'
          AND EXTRACT(YEAR FROM COALESCE(updated_at, tgl_masuk)) = %s
          {owner_sql}
        GROUP BY 1 ORDER BY 2 DESC
    """, _params())
    loss_analysis = [dict(r) for r in c.fetchall()]

    conn.close()
    return {
        "tahun": cur_year,
        "summary": {k: float(v) if isinstance(v, (int, float)) else int(v) if v is not None else 0 for k, v in summary.items()},
        "monthly_forecast": monthly,
        "by_sales": by_sales,
        "loss_analysis": loss_analysis,
    }


@app.get("/api/v1/pipeline/{lead_id}", tags=["Pipeline"])
def pipeline_detail(lead_id: str, user: dict = Depends(require_menu("pipeline"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM leads WHERE lead_id=%s", (lead_id,))
    row = c.fetchone()
    if not row:
        conn.close(); raise HTTPException(404, "Lead tidak ditemukan.")
    lead = _norm(dict(row))
    if user.get("role_id") == 3 and lead.get("sales_owner") != user["nama"]:
        conn.close(); raise HTTPException(403, "Akses ditolak.")
    c.execute("SELECT * FROM follow_up_log WHERE lead_id=%s ORDER BY tgl_fu DESC", (lead_id,))
    logs = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT * FROM contacts WHERE lead_id=%s", (lead_id,))
    contacts = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"lead": lead, "fu_logs": logs, "contacts": contacts}


_LEAD_TRACKED_FIELDS = {
    "stage": "Stage", "prioritas": "Prioritas", "sales_owner": "Sales Owner",
    "propose_value": "Propose Value", "deal_value": "Deal Value",
    "probability": "Probability", "exp_close_date": "Exp. Close Date",
    "next_fu_date": "Next FU Date", "nama_company": "Nama Perusahaan",
    "product": "Produk", "loss_reason": "Loss Reason",
}

def _log_lead_changes(c, lead_id: str, old: dict, new: dict, changed_by: str):
    for field, label in _LEAD_TRACKED_FIELDS.items():
        ov = str(old.get(field) or "")
        nv = str(new.get(field) or "")
        if ov != nv:
            c.execute("""INSERT INTO lead_history (lead_id, field_name, old_value, new_value, changed_by)
                VALUES (%s,%s,%s,%s,%s)""", (lead_id, label, ov or None, nv or None, changed_by))


@app.put("/api/v1/pipeline/{lead_id}", tags=["Pipeline"])
def pipeline_update(lead_id: str, payload: LeadUpdate,
                    user: dict = Depends(require_menu("pipeline"))):
    conn = get_conn(); c = conn.cursor()
    if user.get("role_id") == 3:
        c.execute("SELECT sales_owner FROM leads WHERE lead_id=%s", (lead_id,))
        row = c.fetchone()
        if not row or row["sales_owner"] != user["nama"]:
            conn.close(); raise HTTPException(403, "Akses ditolak.")

    # Ambil data lama untuk audit trail
    c.execute("SELECT * FROM leads WHERE lead_id=%s", (lead_id,))
    old = dict(c.fetchone() or {})

    pv = float(payload.propose_value or 0)
    dv = float(payload.deal_value or 0)
    prob = float(payload.probability or 0)
    wv = dv * (prob/100) if dv else pv * (prob/100)
    c.execute("""UPDATE leads SET
        nama_company=%s,product=%s,contact_person=%s,segmen=%s,sub_segmen=%s,source=%s,
        stage=%s,prioritas=%s,tgl_masuk=%s,propose_value=%s,deal_value=%s,probability=%s,
        exp_close_date=%s,weighted_value=%s,sales_owner=%s,organisasi=%s,
        next_fu_date=%s,remarks=%s,loss_reason=%s,updated_at=NOW() WHERE lead_id=%s""",
        (payload.nama_company, payload.product, payload.contact_person,
         payload.segmen, payload.sub_segmen, payload.source,
         payload.stage, payload.prioritas, payload.tgl_masuk or None,
         pv, dv, prob, payload.exp_close_date or None, wv,
         payload.sales_owner, payload.organisasi or None,
         payload.next_fu_date or None, payload.remarks,
         payload.loss_reason or None, lead_id))

    new = {
        "stage": payload.stage, "prioritas": payload.prioritas,
        "sales_owner": payload.sales_owner, "propose_value": pv,
        "deal_value": dv, "probability": prob,
        "exp_close_date": str(payload.exp_close_date or ""),
        "next_fu_date": str(payload.next_fu_date or ""),
        "nama_company": payload.nama_company, "product": payload.product,
        "loss_reason": payload.loss_reason or "",
    }
    _log_lead_changes(c, lead_id, old, new, user["email"])
    conn.commit(); conn.close()
    return {"message": "Lead berhasil diperbarui.", "lead_id": lead_id}


@app.get("/api/v1/pipeline/{lead_id}/history", tags=["Pipeline"])
def lead_history(lead_id: str, user: dict = Depends(require_menu("pipeline"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("""SELECT id, field_name, old_value, new_value, changed_by, changed_at
        FROM lead_history WHERE lead_id=%s ORDER BY changed_at DESC LIMIT 100""", (lead_id,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"history": rows}


@app.delete("/api/v1/pipeline/{lead_id}", tags=["Pipeline"])
def pipeline_delete(lead_id: str, user: dict = Depends(require_menu("pipeline"))):
    conn = get_conn()
    c = conn.cursor()
    if user.get("role_id") == 3:
        c.execute("SELECT sales_owner FROM leads WHERE lead_id=%s", (lead_id,))
        row = c.fetchone()
        if not row or row["sales_owner"] != user["nama"]:
            conn.close(); raise HTTPException(403, "Akses ditolak.")
    # Hapus records terkait terlebih dahulu (FK constraint)
    c.execute("DELETE FROM lead_history WHERE lead_id=%s", (lead_id,))
    c.execute("DELETE FROM follow_up_log WHERE lead_id=%s", (lead_id,))
    c.execute("DELETE FROM contacts WHERE lead_id=%s", (lead_id,))
    c.execute("DELETE FROM leads WHERE lead_id=%s", (lead_id,))
    conn.commit(); conn.close()
    return {"message": "Lead berhasil dihapus.", "lead_id": lead_id}


# ── FU Templates ─────────────────────────────────────────────────────────────
class FuTemplateCreate(BaseModel):
    nama: str
    catatan: str
    hasil_fu: str = ""
    metode_fu: str = ""

@app.get("/api/v1/fu-templates", tags=["Activity"])
def fu_templates_list(user: dict = Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM fu_templates ORDER BY created_at DESC")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows

@app.post("/api/v1/fu-templates", status_code=201, tags=["Activity"])
def fu_templates_create(payload: FuTemplateCreate, user: dict = Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute("""INSERT INTO fu_templates (nama, catatan, hasil_fu, metode_fu, created_by)
        VALUES (%s,%s,%s,%s,%s)""",
        (payload.nama, payload.catatan, payload.hasil_fu, payload.metode_fu, user["email"]))
    conn.commit(); conn.close()
    return {"message": "Template disimpan."}

@app.delete("/api/v1/fu-templates/{tid}", tags=["Activity"])
def fu_templates_delete(tid: int, user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    c.execute("DELETE FROM fu_templates WHERE id=%s", (tid,))
    conn.commit(); conn.close()
    return {"message": "Template dihapus."}


# ═══════════════════════════════════════════════════════════════════════════════
# TODAY / ACTIVITY
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/today", tags=["Activity"])
def today_dashboard(user: dict = Depends(require_menu("today"))):
    conn = get_conn()
    c = conn.cursor()
    today_str = str(date.today())
    is_sales = user.get("role_id") == 3
    sf = " AND sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []

    c.execute(f"""SELECT l.*, (CURRENT_DATE - l.next_fu_date)::integer as days_overdue
        FROM leads l WHERE l.next_fu_date < %s AND l.stage NOT IN ('Won','Lost')
        AND l.next_fu_date IS NOT NULL{sf} ORDER BY l.next_fu_date ASC""", [today_str] + sp)
    overdue = [_norm(dict(r)) for r in c.fetchall()]

    c.execute(f"""SELECT l.* FROM leads l WHERE l.next_fu_date = %s AND l.stage NOT IN ('Won','Lost')
        {sf} ORDER BY l.prioritas DESC""", [today_str] + sp)
    due_today = [_norm(dict(r)) for r in c.fetchall()]

    c.execute(f"""SELECT l.*, (l.next_fu_date - CURRENT_DATE)::integer as days_until
        FROM leads l WHERE l.next_fu_date > %s AND l.next_fu_date <= (CURRENT_DATE + INTERVAL '7 days')
        AND l.stage NOT IN ('Won','Lost'){sf} ORDER BY l.next_fu_date ASC""", [today_str] + sp)
    upcoming = [_norm(dict(r)) for r in c.fetchall()]

    c.execute(f"""SELECT * FROM leads WHERE next_fu_date IS NULL
        AND stage NOT IN ('Won','Lost'){sf} ORDER BY prioritas DESC, propose_value DESC""", sp)
    unscheduled = [_norm(dict(r)) for r in c.fetchall()]

    c.execute(f"SELECT COUNT(*) as cnt FROM leads WHERE stage NOT IN ('Won','Lost'){sf}", sp)
    total_active = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM follow_up_log WHERE tgl_fu = %s", (today_str,))
    fu_done_today = c.fetchone()["cnt"]

    c.execute(f"""SELECT *,
        COALESCE((CURRENT_DATE-last_fu_date)::integer,(CURRENT_DATE-tgl_masuk)::integer,999) as days_without_fu
        FROM leads WHERE stage NOT IN ('Won','Lost') AND next_fu_date IS NULL{sf}
        ORDER BY CASE prioritas WHEN 'Hot' THEN 1 WHEN 'Warm' THEN 2 ELSE 3 END, days_without_fu DESC""", sp)
    stale_raw = [_norm(dict(r)) for r in c.fetchall()]
    for l in stale_raw:
        l["stale_flag"] = compute_stale_flag(l["prioritas"], l["days_without_fu"])

    # Normalize days field name for frontend
    for l in stale_raw:
        l["days_since_fu"] = l.get("days_without_fu", 0)

    conn.close()
    stale_urgent  = [l for l in stale_raw if l["stale_flag"] == "URGENT"]
    stale_warning = [l for l in stale_raw if l["stale_flag"] == "WARNING"]
    stale_stale   = [l for l in stale_raw if l["stale_flag"] == "STALE"]
    return {
        "today": today_str,
        "overdue": overdue, "due_today": due_today,
        "upcoming": upcoming, "unscheduled": unscheduled,
        "total_active": total_active, "fu_done_today": fu_done_today,
        "stale": stale_urgent + stale_warning + stale_stale,
        "stale_urgent":  stale_urgent,
        "stale_warning": stale_warning,
        "stale_stale":   stale_stale,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FU SCHEDULE
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/schedule", tags=["Activity"])
def fu_schedule(user: dict = Depends(require_menu("schedule"))):
    conn = get_conn()
    c = conn.cursor()
    is_sales = user.get("role_id") == 3
    sf = " AND l.sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []
    c.execute(f"""SELECT l.*,
        CASE WHEN l.next_fu_date < CURRENT_DATE THEN 'Overdue'
             WHEN l.next_fu_date = CURRENT_DATE THEN 'Today'
             ELSE 'Upcoming' END as fu_status
        FROM leads l WHERE l.stage NOT IN ('Won','Lost') AND l.next_fu_date IS NOT NULL{sf}
        ORDER BY l.next_fu_date ASC""", sp)
    schedule = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    from_date = schedule[0]["next_fu_date"] if schedule else None
    to_date   = schedule[-1]["next_fu_date"] if schedule else None
    return {
        "schedule": schedule,
        "total": len(schedule),
        "from": from_date,
        "to": to_date,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FOLLOW-UP LOG
# ═══════════════════════════════════════════════════════════════════════════════

class FollowUpCreate(BaseModel):
    lead_id: str
    tgl_fu: str
    metode_fu: str = ""
    kontak: str = ""
    hasil_fu: str = ""
    catatan_fu: str = ""
    stage_saat_fu: str = ""
    next_action: str = ""
    tgl_fu_berikut: str = ""
    sales_owner: str = ""
    status: str = "Done"


@app.get("/api/v1/followup", tags=["Activity"])
def followup_list(
    lead_id: str = "", search: str = "",
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=200),
    user: dict = Depends(require_menu("followup"))
):
    import math
    conn = get_conn()
    c = conn.cursor()
    is_sales = user.get("role_id") == 3
    base = """FROM follow_up_log f
        JOIN leads l ON l.lead_id = f.lead_id
        WHERE 1=1"""
    params: list = []
    if is_sales: base += " AND l.sales_owner=%s"; params.append(user["nama"])
    if lead_id:  base += " AND f.lead_id=%s"; params.append(lead_id)
    if search:   base += " AND (f.nama_company LIKE %s OR f.catatan_fu LIKE %s OR f.kontak LIKE %s)"; params += [f"%{search}%"]*3
    c.execute(f"SELECT COUNT(*) cnt {base}", params)
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(f"SELECT f.* {base} ORDER BY f.tgl_fu DESC LIMIT %s OFFSET %s", params + [per_page, offset])
    logs = [_norm(dict(r)) for r in c.fetchall()]
    if is_sales:
        c.execute("SELECT lead_id, nama_company FROM leads WHERE sales_owner=%s ORDER BY nama_company", (user["nama"],))
    else:
        c.execute("SELECT lead_id, nama_company FROM leads ORDER BY nama_company")
    leads_dropdown = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"logs": logs, "leads": leads_dropdown, "total": total, "page": page, "per_page": per_page, "total_pages": math.ceil(total/per_page) if total else 1}


@app.post("/api/v1/followup", status_code=201, tags=["Activity"])
def followup_create(payload: FollowUpCreate, user: dict = Depends(require_menu("followup"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT nama_company FROM leads WHERE lead_id=%s", (payload.lead_id,))
    row = c.fetchone()
    nama = row["nama_company"] if row else ""
    c.execute("""INSERT INTO follow_up_log
        (fu_id,lead_id,tgl_fu,nama_company,sales_owner,metode_fu,kontak,hasil_fu,
         catatan_fu,stage_saat_fu,next_action,tgl_fu_berikut,status)
        VALUES ('TEMP',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (payload.lead_id, payload.tgl_fu, nama, payload.sales_owner,
         payload.metode_fu, payload.kontak, payload.hasil_fu, payload.catatan_fu,
         payload.stage_saat_fu, payload.next_action,
         payload.tgl_fu_berikut or None, payload.status))
    new_id = c.fetchone()['id']
    fid = f"FU-{new_id:03d}"
    c.execute("UPDATE follow_up_log SET fu_id=%s WHERE id=%s", (fid, new_id))
    if payload.tgl_fu_berikut:
        c.execute("""UPDATE leads SET last_fu_date=%s, last_fu_notes=%s, fu_count=fu_count+1,
            next_fu_date=%s, updated_at=NOW() WHERE lead_id=%s""",
            (payload.tgl_fu, payload.catatan_fu, payload.tgl_fu_berikut, payload.lead_id))
    else:
        c.execute("""UPDATE leads SET last_fu_date=%s, last_fu_notes=%s, fu_count=fu_count+1,
            updated_at=NOW() WHERE lead_id=%s""",
            (payload.tgl_fu, payload.catatan_fu, payload.lead_id))
    conn.commit(); conn.close()
    return {"message": "Follow-up berhasil dicatat.", "fu_id": fid}


# ═══════════════════════════════════════════════════════════════════════════════
# WIN-LOSS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/winloss", tags=["Pipeline"])
def winloss(tahun: int = 0, user: dict = Depends(require_menu("winloss"))):
    conn = get_conn()
    c = conn.cursor()
    cur_year = tahun or date.today().year
    is_sales = user.get("role_id") == 3
    sf = " AND sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []

    # ── Summary ───────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT
            COUNT(*) FILTER (WHERE stage='Won')  AS won,
            COUNT(*) FILTER (WHERE stage='Lost') AS lost,
            COALESCE(SUM(COALESCE(NULLIF(deal_value,0), propose_value)) FILTER (WHERE stage='Won'),  0) AS won_value,
            COALESCE(SUM(COALESCE(NULLIF(deal_value,0), propose_value)) FILTER (WHERE stage='Lost'), 0) AS lost_value
        FROM leads
        WHERE stage IN ('Won','Lost')
          AND DATE_PART('year', COALESCE(updated_at, tgl_masuk)) = %s{sf}
    """, [cur_year] + sp)
    s = _norm(dict(c.fetchone()))
    won  = int(s["won"]  or 0)
    lost = int(s["lost"] or 0)
    total_closed = won + lost
    summary = {
        "won":        won,
        "lost":       lost,
        "won_value":  float(s["won_value"]  or 0),
        "lost_value": float(s["lost_value"] or 0),
        "avg_deal":   float(s["won_value"] or 0) / won if won else 0,
        "win_rate":   round(won / total_closed * 100) if total_closed else 0,
    }

    # ── Won leads ─────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, product, segmen, sales_owner,
               deal_value, propose_value, exp_close_date, updated_at,
               COALESCE(NULLIF(deal_value,0), propose_value) AS effective_value
        FROM leads WHERE stage='Won'
          AND DATE_PART('year', COALESCE(updated_at, tgl_masuk)) = %s{sf}
        ORDER BY updated_at DESC
    """, [cur_year] + sp)
    won_leads = [_norm(dict(r)) for r in c.fetchall()]

    # ── Lost leads ────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, product, segmen, sales_owner,
               propose_value, loss_reason, last_fu_notes, updated_at
        FROM leads WHERE stage='Lost'
          AND DATE_PART('year', COALESCE(updated_at, tgl_masuk)) = %s{sf}
        ORDER BY updated_at DESC
    """, [cur_year] + sp)
    lost_leads = [_norm(dict(r)) for r in c.fetchall()]

    # ── By Product ────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(NULLIF(product,''),'—') AS product,
               COUNT(*) FILTER (WHERE stage='Won')  AS won,
               COUNT(*) FILTER (WHERE stage='Lost') AS lost
        FROM leads WHERE stage IN ('Won','Lost')
          AND DATE_PART('year', COALESCE(updated_at, tgl_masuk)) = %s{sf}
        GROUP BY 1 ORDER BY 2+3 DESC
    """, [cur_year] + sp)
    by_product = [dict(r) for r in c.fetchall()]

    # ── By Segmen ─────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(NULLIF(segmen,''),'—') AS segmen,
               COUNT(*) FILTER (WHERE stage='Won')  AS won,
               COUNT(*) FILTER (WHERE stage='Lost') AS lost
        FROM leads WHERE stage IN ('Won','Lost')
          AND DATE_PART('year', COALESCE(updated_at, tgl_masuk)) = %s{sf}
        GROUP BY 1 ORDER BY 2+3 DESC
    """, [cur_year] + sp)
    by_segmen = [dict(r) for r in c.fetchall()]

    # ── By Sales (hanya untuk non-sales) ─────────────────────────────────────
    by_sales = []
    if not is_sales:
        c.execute("""
            SELECT COALESCE(NULLIF(sales_owner,''),'Unassigned') AS sales_owner,
                   COUNT(*) FILTER (WHERE stage='Won')  AS won,
                   COUNT(*) FILTER (WHERE stage='Lost') AS lost
            FROM leads WHERE stage IN ('Won','Lost')
              AND DATE_PART('year', COALESCE(updated_at, tgl_masuk)) = %s
            GROUP BY 1 ORDER BY 2 DESC
        """, (cur_year,))
        by_sales = [dict(r) for r in c.fetchall()]

    conn.close()
    return {
        "tahun":     cur_year,
        "summary":   summary,
        "won_leads": won_leads,
        "lost_leads":lost_leads,
        "by_product":by_product,
        "by_segmen": by_segmen,
        "by_sales":  by_sales,
    }


@app.get("/api/v1/winloss/reason/{lead_id}", tags=["Pipeline"])
def winloss_reason_get(lead_id: str, user: dict = Depends(require_menu("winloss"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM win_loss WHERE lead_id=%s ORDER BY id DESC LIMIT 1", (lead_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return None
    return _norm(dict(row))


class WinLossReason(BaseModel):
    lead_id: str
    alasan: str
    kompetitor: str = ""
    lesson_learned: str = ""


@app.post("/api/v1/winloss/reason", tags=["Pipeline"])
def winloss_reason_save(payload: WinLossReason, user: dict = Depends(require_menu("winloss"))):
    conn = get_conn()
    c = conn.cursor()
    # Ambil info lead
    c.execute("SELECT nama_company, segmen, stage, deal_value, propose_value, sales_owner, tgl_masuk FROM leads WHERE lead_id=%s", (payload.lead_id,))
    lead = c.fetchone()
    if not lead:
        conn.close()
        raise HTTPException(404, "Lead tidak ditemukan.")
    hasil = lead["stage"]
    c.execute("SELECT id FROM win_loss WHERE lead_id=%s", (payload.lead_id,))
    existing = c.fetchone()
    if existing:
        c.execute("UPDATE win_loss SET alasan=%s, kompetitor=%s, lesson_learned=%s WHERE lead_id=%s",
                  (payload.alasan, payload.kompetitor or None, payload.lesson_learned or None, payload.lead_id))
    else:
        c.execute("""INSERT INTO win_loss
            (lead_id, nama_company, segmen, hasil, deal_value, sales_owner, alasan, kompetitor, lesson_learned, tgl_masuk, tgl_close)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,CURRENT_DATE)""",
            (payload.lead_id, lead["nama_company"], lead["segmen"], hasil,
             float(lead["deal_value"] or lead["propose_value"] or 0),
             lead["sales_owner"], payload.alasan,
             payload.kompetitor or None, payload.lesson_learned or None,
             lead["tgl_masuk"]))
    conn.commit(); conn.close()
    return {"message": "Alasan berhasil disimpan."}


# ═══════════════════════════════════════════════════════════════════════════════
# CONTACTS
# ═══════════════════════════════════════════════════════════════════════════════

class ContactCreate(BaseModel):
    lead_id: Optional[str] = None
    nama_company: Optional[str] = ""
    nama_contact: str
    jabatan: Optional[str] = ""; dept: Optional[str] = ""; role: Optional[str] = ""
    no_hp: Optional[str] = None; email: Optional[str] = None; telepon: Optional[str] = None
    linkedin: Optional[str] = None; preferensi_kontak: Optional[str] = None; catatan: Optional[str] = None


@app.get("/api/v1/contacts", tags=["Contacts"])
def contacts_list(
    q: str = "", search: str = "",
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=200),
    user: dict = Depends(require_menu("contacts"))
):
    import math
    conn = get_conn()
    c = conn.cursor()
    keyword = q or search
    base = "FROM contacts"
    params: list = []
    if keyword:
        base += " WHERE nama_contact ILIKE %s OR nama_company ILIKE %s OR no_hp ILIKE %s OR email ILIKE %s"
        params = [f"%{keyword}%"]*4
    c.execute(f"SELECT COUNT(*) cnt {base}", params)
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(f"SELECT * {base} ORDER BY nama_company LIMIT %s OFFSET %s", params + [per_page, offset])
    ct = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT lead_id, nama_company FROM leads WHERE stage NOT IN ('Lost') ORDER BY nama_company")
    leads = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"contacts": ct, "total": total, "page": page, "per_page": per_page, "total_pages": math.ceil(total/per_page) if total else 1, "leads": leads}


@app.post("/api/v1/contacts", status_code=201, tags=["Contacts"])
def contact_create(payload: ContactCreate, user: dict = Depends(require_menu("contacts"))):
    conn = get_conn()
    c = conn.cursor()
    # Ambil nama_company dari lead jika lead_id diisi
    nama_co = payload.nama_company or ""
    if payload.lead_id:
        c.execute("SELECT nama_company FROM leads WHERE lead_id=%s", (payload.lead_id,))
        row = c.fetchone()
        if row:
            nama_co = row["nama_company"]
    c.execute("""INSERT INTO contacts
        (lead_id,nama_company,nama_contact,jabatan,dept,role,no_hp,email,telepon,linkedin,preferensi_kontak,catatan)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (payload.lead_id or None, nama_co, payload.nama_contact, payload.jabatan, payload.dept,
         payload.role, payload.no_hp or None, payload.email or None, payload.telepon or None,
         payload.linkedin or None, payload.preferensi_kontak or None, payload.catatan or None))
    conn.commit(); conn.close()
    return {"message": "Kontak berhasil ditambahkan."}


# ═══════════════════════════════════════════════════════════════════════════════
# PIPELINE INSIGHTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/insights", tags=["Pipeline"])
def insights(user: dict = Depends(require_menu("insights"))):
    conn = get_conn()
    c = conn.cursor()
    today_str = date.today().isoformat()
    is_sales = user.get("role_id") == 3
    sf = " AND sales_owner=%s" if is_sales else ""
    sw = (" WHERE sales_owner=%s" if is_sales else "")
    sp = [user["nama"]] if is_sales else []

    # ── Stats ─────────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (WHERE stage='Won')  AS won,
            COUNT(*) FILTER (WHERE stage='Lost') AS lost,
            COUNT(*) FILTER (WHERE stage='On Hold') AS on_hold,
            COUNT(*) FILTER (WHERE stage NOT IN ('Won','Lost','On Hold')) AS aktif,
            COUNT(*) FILTER (WHERE (sales_owner IS NULL OR sales_owner='') AND stage NOT IN ('Won','Lost')) AS unassigned,
            COALESCE(SUM(propose_value) FILTER (WHERE stage NOT IN ('Won','Lost')),0) AS total_pipeline,
            COALESCE(SUM(propose_value) FILTER (WHERE stage NOT IN ('Won','Lost','On Hold')),0) AS active_pipeline,
            COALESCE(SUM(weighted_value) FILTER (WHERE stage NOT IN ('Won','Lost')),0) AS weighted_pipeline,
            COALESCE(SUM(COALESCE(NULLIF(deal_value,0), propose_value)) FILTER (WHERE stage='Won'),0) AS total_won
        FROM leads{sw}
    """, sp)
    stats = _norm(dict(c.fetchone()))

    # ── By Stage ─────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT stage,
               COUNT(*) AS jumlah,
               COALESCE(SUM(
                   CASE WHEN stage='Won'
                        THEN COALESCE(NULLIF(deal_value,0), propose_value)
                        ELSE propose_value END
               ),0) AS total_nilai,
               COALESCE(SUM(weighted_value),0) AS weighted
        FROM leads{sw} GROUP BY stage
        ORDER BY CASE stage
            WHEN 'New' THEN 1 WHEN 'In Progress' THEN 2 WHEN 'Demo Scheduled' THEN 3
            WHEN 'Proposal Sent' THEN 4 WHEN 'Negotiation' THEN 5 WHEN 'Won' THEN 6
            WHEN 'On Hold' THEN 7 WHEN 'Lost' THEN 8 ELSE 9 END
    """, sp)
    by_stage = [_norm(dict(r)) for r in c.fetchall()]

    # ── Monthly Trend (12 bulan terakhir) ────────────────────────────────────
    c.execute(f"""
        SELECT TO_CHAR(tgl_masuk,'YYYY-MM') AS bulan, COUNT(*) AS jumlah,
               COALESCE(SUM(propose_value),0) AS total_nilai
        FROM leads WHERE tgl_masuk >= CURRENT_DATE - INTERVAL '12 months'{sf}
        GROUP BY 1 ORDER BY 1
    """, sp)
    monthly_trend = [_norm(dict(r)) for r in c.fetchall()]

    # ── Velocity: avg days to close ─────────────────────────────────────────
    c.execute(f"""
        SELECT stage,
               ROUND(AVG(
                   COALESCE(
                       NULLIF(days_in_stage, 0),
                       (updated_at::date - tgl_masuk)
                   )
               )) AS avg_days,
               COUNT(*) AS cnt
        FROM leads
        WHERE stage IN ('Won','Lost')
          AND tgl_masuk IS NOT NULL{sf}
        GROUP BY stage ORDER BY avg_days DESC
    """, sp)
    velocity = [_norm(dict(r)) for r in c.fetchall()]

    # ── By Sales (hanya untuk non-sales) ─────────────────────────────────────
    by_sales = []
    if not is_sales:
        c.execute("""
            SELECT COALESCE(NULLIF(sales_owner,''),'Unassigned') AS sales_owner,
                   COUNT(*) AS jumlah,
                   COUNT(*) FILTER (WHERE stage NOT IN ('Won','Lost','On Hold')) AS aktif,
                   COALESCE(SUM(propose_value),0) AS total_nilai
            FROM leads GROUP BY 1 ORDER BY jumlah DESC
        """)
        by_sales = [_norm(dict(r)) for r in c.fetchall()]

    # ── Weighted Forecast per Stage ───────────────────────────────────────────
    c.execute(f"""
        SELECT stage,
               COUNT(*) AS jumlah,
               COALESCE(SUM(propose_value),0) AS propose_value,
               COALESCE(SUM(weighted_value),0) AS weighted_value,
               ROUND(AVG(probability)) AS avg_probability
        FROM leads WHERE stage NOT IN ('Won','Lost'){sf}
        GROUP BY stage ORDER BY weighted_value DESC
    """, sp)
    weighted_forecast = [_norm(dict(r)) for r in c.fetchall()]

    # ── Stale Leads ──────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, stage, prioritas, sales_owner,
               propose_value, last_fu_date, next_fu_date,
               CURRENT_DATE - COALESCE(last_fu_date::date, tgl_masuk) AS days_since_fu
        FROM leads WHERE stage NOT IN ('Won','Lost'){sf}
          AND CURRENT_DATE - COALESCE(last_fu_date::date, tgl_masuk) > 30
        ORDER BY days_since_fu DESC LIMIT 20
    """, sp)
    stale_leads = [_norm(dict(r)) for r in c.fetchall()]

    # ── Hot Stale ─────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, stage, prioritas, sales_owner,
               propose_value, last_fu_date,
               CURRENT_DATE - COALESCE(last_fu_date::date, tgl_masuk) AS days_since_fu
        FROM leads WHERE stage NOT IN ('Won','Lost')
          AND prioritas='Hot'{sf}
          AND CURRENT_DATE - COALESCE(last_fu_date::date, tgl_masuk) > 14
        ORDER BY days_since_fu DESC LIMIT 10
    """, sp)
    hot_stale = [_norm(dict(r)) for r in c.fetchall()]

    # ── On Hold at Risk ───────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, stage, prioritas, sales_owner,
               propose_value, last_fu_date,
               CURRENT_DATE - COALESCE(last_fu_date::date, tgl_masuk) AS days_stagnant
        FROM leads WHERE stage='On Hold'{sf}
        ORDER BY days_stagnant DESC LIMIT 10
    """, sp)
    onhold_risk = [_norm(dict(r)) for r in c.fetchall()]

    # ── Source Conversion ─────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(NULLIF(source,''),'—') AS source,
               COUNT(*) AS total,
               COUNT(*) FILTER (WHERE stage='Won')  AS won,
               COUNT(*) FILTER (WHERE stage='Lost') AS lost,
               COALESCE(SUM(propose_value),0) AS pipeline_value
        FROM leads{sw} GROUP BY 1 ORDER BY total DESC
    """, sp)
    source_conversion = [_norm(dict(r)) for r in c.fetchall()]

    # ── Closing Soon ─────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, stage, prioritas, sales_owner,
               propose_value, exp_close_date,
               (exp_close_date - CURRENT_DATE) AS days_until_close
        FROM leads
        WHERE stage NOT IN ('Won','Lost')
          AND exp_close_date IS NOT NULL{sf}
        ORDER BY exp_close_date ASC LIMIT 15
    """, sp)
    closing_soon = [_norm(dict(r)) for r in c.fetchall()]

    # ── By Product ────────────────────────────────────────────────────────────
    c.execute(f"""
        SELECT COALESCE(NULLIF(product,''),'—') AS product,
               COUNT(*) AS jumlah,
               COALESCE(SUM(propose_value),0) AS total_nilai
        FROM leads WHERE stage NOT IN ('Won','Lost'){sf}
        GROUP BY 1 ORDER BY total_nilai DESC
    """, sp)
    by_product = [_norm(dict(r)) for r in c.fetchall()]

    # ── Top 5 High Value Aktif ────────────────────────────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, stage, sales_owner,
               propose_value, probability, exp_close_date
        FROM leads WHERE stage NOT IN ('Won','Lost','On Hold'){sf}
        ORDER BY propose_value DESC NULLS LAST LIMIT 5
    """, sp)
    high_value = [_norm(dict(r)) for r in c.fetchall()]

    # ── Ready to Close (Proposal Sent / Negotiation) ──────────────────────────
    c.execute(f"""
        SELECT lead_id, nama_company, stage, sales_owner,
               propose_value, probability, exp_close_date
        FROM leads WHERE stage IN ('Proposal Sent','Negotiation'){sf}
        ORDER BY probability DESC NULLS LAST, propose_value DESC LIMIT 10
    """, sp)
    ready_to_close = [_norm(dict(r)) for r in c.fetchall()]

    # avg_days_close: rata-rata dari velocity Won
    won_velocity = next((v for v in velocity if v["stage"] == "Won"), None)
    avg_days_close = int(won_velocity["avg_days"]) if won_velocity and won_velocity["avg_days"] else None

    conn.close()
    return {
        "stats":             stats,
        "avg_days_close":    avg_days_close,
        "by_stage":          by_stage,
        "monthly_trend":     monthly_trend,
        "velocity":          velocity,
        "by_sales":          by_sales,
        "weighted_forecast": weighted_forecast,
        "stale_leads":       stale_leads,
        "hot_stale":         hot_stale,
        "onhold_risk":       onhold_risk,
        "source_conversion": source_conversion,
        "high_value":        high_value,
        "ready_to_close":    ready_to_close,
        "closing_soon":      closing_soon,
        "by_product":        by_product,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# REVENUE — SUMMARY & INSIGHTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/revenue/summary", tags=["Revenue"])
def revenue_summary_api(
    tahun: int = Query(0),
    user: dict = Depends(require_menu("rev_dashboard"))
):
    cur_year = tahun if tahun else date.today().year
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    s = _norm(dict(c.fetchone())); tt = s['t'] or 0; ta = s['a'] or 0
    ach_pct = round(float(ta or 0)/float(tt or 1)*100,1) if tt else 0

    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE kategori='Recurring' AND is_active=1 AND tahun=%s", (cur_year,))
    r = _norm(dict(c.fetchone())); rec_target = float(r['t'] or 0); rec_actual = float(r['a'] or 0)
    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE kategori='Project' AND is_active=1 AND tahun=%s", (cur_year,))
    r = _norm(dict(c.fetchone())); prj_target = float(r['t'] or 0); prj_actual = float(r['a'] or 0)
    c.execute("SELECT status, COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY status", (cur_year,))
    by_status = {r['status']: r['cnt'] for r in c.fetchall()}
    c.execute("SELECT COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    total_projects = c.fetchone()['cnt']

    c.execute("""SELECT m.month_num, m.month_name, SUM(m.target) total_target, SUM(m.actual) total_actual
        FROM revenue_monthly m JOIN revenue_projects p ON m.project_id=p.project_id
        WHERE p.is_active=1 AND p.tahun=%s GROUP BY m.month_num, m.month_name ORDER BY m.month_num""", (cur_year,))
    monthly_trend_rows = {r['month_num']: _norm(dict(r)) for r in c.fetchall()}

    # Billed per bulan dari tabel invoices (berdasarkan EXTRACT month dari invoice_date)
    c.execute("""SELECT EXTRACT(MONTH FROM invoice_date)::int AS month_num,
                        COALESCE(SUM(invoice_amount), 0) AS total_billed
                 FROM invoices WHERE tahun=%s
                 GROUP BY month_num ORDER BY month_num""", (cur_year,))
    for r in c.fetchall():
        mn = r['month_num']
        if mn in monthly_trend_rows:
            monthly_trend_rows[mn]['total_billed'] = float(r['total_billed'] or 0)
        # bulan yang ada invoice tapi tidak ada di revenue_monthly diabaikan

    monthly_trend = sorted(monthly_trend_rows.values(), key=lambda x: x['month_num'])

    c.execute("""SELECT * FROM revenue_projects WHERE status IN ('Critical','At Risk') AND is_active=1 AND tahun=%s
        ORDER BY risk_level DESC, revenue_target DESC LIMIT 10""", (cur_year,))
    critical = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("SELECT pic, COUNT(*) cnt, SUM(revenue_target) target, SUM(actual_revenue) actual FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY pic", (cur_year,))
    by_pic = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]

    c.execute("SELECT organisasi, SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY organisasi ORDER BY SUM(revenue_target) DESC", (cur_year,))
    org_breakdown = [
        {"organisasi": r['organisasi'], "target": float(r['t'] or 0), "actual": float(r['a'] or 0),
         "ach": round(float(r['a'] or 0) / max(float(r['t'] or 0), 1) * 100, 1)}
        for r in c.fetchall()
    ]

    c.execute("""
        SELECT CEIL(m.month_num / 3.0)::int AS q,
               SUM(m.target) AS total_target, SUM(m.actual) AS total_actual
        FROM revenue_monthly m
        JOIN revenue_projects p ON m.project_id = p.project_id
        WHERE p.is_active=1 AND p.tahun=%s
        GROUP BY q ORDER BY q
    """, (cur_year,))
    quarter_trend = [
        {"quarter": f"Q{r['q']}", "target": float(r['total_target'] or 0), "actual": float(r['total_actual'] or 0),
         "ach": round(float(r['total_actual'] or 0) / max(float(r['total_target'] or 0), 1) * 100, 1)}
        for r in c.fetchall()
    ]

    c.execute("SELECT COALESCE(SUM(invoice_amount),0) t, COALESCE(SUM(paid_amount),0) p FROM invoices WHERE tahun=%s", (cur_year,))
    iv = _norm(dict(c.fetchone()))
    total_billed = float(iv['t'] or 0)

    conn.close()

    return {
        "cur_year": cur_year, "years": years,
        "total_target": float(tt or 0), "total_actual": float(ta or 0), "ach_pct": ach_pct,
        "total_billed": total_billed,
        "rec_target": rec_target, "rec_actual": rec_actual,
        "prj_target": prj_target, "prj_actual": prj_actual,
        "by_status": by_status, "total_projects": total_projects,
        "monthly_trend": monthly_trend, "critical": critical, "by_pic": by_pic,
        "org_breakdown": org_breakdown, "quarter_trend": quarter_trend,
    }


@app.get("/api/v1/revenue/insights", tags=["Revenue"])
def revenue_insights_api(tahun: int = Query(0), user: dict = Depends(require_menu("rev_dashboard"))):
    cur_year = tahun if tahun else date.today().year
    cur_month = date.today().month if tahun == 0 or tahun == date.today().year else 12
    conn = get_conn()
    c = conn.cursor()

    # ── Years ─────────────────────────────────────────────────────────────────
    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]

    # ── Total Target & Actual ─────────────────────────────────────────────────
    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    s = _norm(dict(c.fetchone()))
    total_target = float(s['t'] or 0)
    total_actual = float(s['a'] or 0)
    ach_pct = round(total_actual / max(total_target, 1) * 100, 1)

    # ── Per Kategori ──────────────────────────────────────────────────────────
    c.execute("SELECT kategori, SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY kategori", (cur_year,))
    by_kat = {r['kategori']: _norm(dict(r)) for r in c.fetchall()}
    project_target   = float(by_kat.get('Project',   {}).get('t', 0) or 0)
    project_actual   = float(by_kat.get('Project',   {}).get('a', 0) or 0)
    recurring_target = float(by_kat.get('Recurring', {}).get('t', 0) or 0)
    recurring_actual = float(by_kat.get('Recurring', {}).get('a', 0) or 0)
    project_ach   = round(project_actual   / max(project_target,   1) * 100, 1)
    recurring_ach = round(recurring_actual / max(recurring_target, 1) * 100, 1)

    # ── Per Type Kontrak ──────────────────────────────────────────────────────
    c.execute("""SELECT COALESCE(NULLIF(type,''),'—') AS type,
        COUNT(*) AS cnt, SUM(revenue_target) AS target, SUM(actual_revenue) AS actual
        FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY 1 ORDER BY actual DESC""", (cur_year,))
    by_type = [_norm(dict(r)) for r in c.fetchall()]

    # ── Per Status ────────────────────────────────────────────────────────────
    c.execute("""
        SELECT status, COUNT(*) AS cnt,
               COALESCE(SUM(revenue_target),0) AS target,
               COALESCE(SUM(actual_revenue),0) AS actual
        FROM revenue_projects WHERE is_active=1 AND tahun=%s
        GROUP BY status
        ORDER BY CASE status WHEN 'On Track' THEN 1 WHEN 'At Risk' THEN 2 WHEN 'Critical' THEN 3 ELSE 4 END
    """, (cur_year,))
    by_status = [_norm(dict(r)) for r in c.fetchall()]
    critical_count = sum(r['cnt'] for r in by_status if r['status'] in ('Critical', 'At Risk'))
    c.execute("SELECT COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    total_projects = c.fetchone()['cnt']

    # ── Per Risk Level ────────────────────────────────────────────────────────
    c.execute("""SELECT COALESCE(NULLIF(risk_level,''),'—') AS risk_level,
        COUNT(*) AS cnt, SUM(revenue_target) AS target, SUM(actual_revenue) AS actual
        FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY 1 ORDER BY target DESC""", (cur_year,))
    by_risk = [_norm(dict(r)) for r in c.fetchall()]

    # ── Per Organisasi (FSP-ECO / FSP-CORE) ──────────────────────────────────
    c.execute("SELECT organisasi, SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY organisasi", (cur_year,))
    by_org_rows = c.fetchall()
    by_org = {r['organisasi']: _norm(dict(r)) for r in by_org_rows}
    def _org(key):
        o = by_org.get(key, {})
        t = float(o.get('t', 0) or 0); a = float(o.get('a', 0) or 0)
        return t, a, round(a / max(t, 1) * 100, 1)
    fsp_eco_target,  fsp_eco_actual,  fsp_eco_ach  = _org('FSP-ECO')
    fsp_core_target, fsp_core_actual, fsp_core_ach = _org('FSP-CORE')
    org_breakdown = [
        {"organisasi": r['organisasi'], "target": float(r['t'] or 0), "actual": float(r['a'] or 0),
         "ach": round(float(r['a'] or 0) / max(float(r['t'] or 0), 1) * 100, 1)}
        for r in sorted(by_org_rows, key=lambda x: -(float(x['t'] or 0)))
    ]

    # ── Monthly ───────────────────────────────────────────────────────────────
    c.execute("""SELECT m.month_num, m.month_name, SUM(m.target) AS target, SUM(m.actual) AS actual
        FROM revenue_monthly m JOIN revenue_projects p ON m.project_id=p.project_id
        WHERE p.is_active=1 AND p.tahun=%s GROUP BY m.month_num, m.month_name ORDER BY m.month_num""", (cur_year,))
    monthly = [dict(_norm(dict(r)), is_past=r['month_num'] <= cur_month) for r in c.fetchall()]

    # ── YTD gap & run rate & projection ──────────────────────────────────────
    past_months = [m for m in monthly if m['month_num'] <= cur_month]
    ytd_target  = sum(float(m['target'] or 0) for m in past_months)
    ytd_actual  = sum(float(m['actual'] or 0) for m in past_months)
    gap_ytd     = ytd_target - ytd_actual
    past_months_count = len(past_months)
    ach_months_count  = sum(1 for m in past_months if float(m['actual'] or 0) >= float(m['target'] or 0) * 0.8)
    miss_months_count = past_months_count - ach_months_count
    run_rate    = ytd_actual / max(past_months_count, 1)
    remain_months = 12 - cur_month
    projected_eoy = ytd_actual + run_rate * remain_months
    projected_ach = round(projected_eoy / max(total_target, 1) * 100, 1)
    remain_target = total_target - ytd_actual

    # ── At Risk Projects ──────────────────────────────────────────────────────
    c.execute("""SELECT project_id, product, client, organisasi, pic,
        revenue_target, actual_revenue, status, risk_level, kategori
        FROM revenue_projects WHERE status IN ('Critical','At Risk') AND is_active=1 AND tahun=%s
        ORDER BY revenue_target DESC""", (cur_year,))
    at_risk_projects = [_norm(dict(r)) for r in c.fetchall()]

    # ── Zero Realisasi ────────────────────────────────────────────────────────
    c.execute("""SELECT project_id, product, client, organisasi, pic,
        revenue_target, actual_revenue, type, risk_level, kategori
        FROM revenue_projects WHERE actual_revenue=0 AND revenue_target>0 AND is_active=1 AND tahun=%s
        ORDER BY revenue_target DESC""", (cur_year,))
    zero_projects = [_norm(dict(r)) for r in c.fetchall()]

    # ── Top Contributors ──────────────────────────────────────────────────────
    c.execute("""SELECT project_id, product, client, organisasi, pic,
        revenue_target, actual_revenue,
        ROUND(actual_revenue::numeric / NULLIF(revenue_target, 0) * 100, 1) AS ach_pct
        FROM revenue_projects WHERE is_active=1 AND tahun=%s AND actual_revenue > 0
        ORDER BY actual_revenue DESC LIMIT 5""", (cur_year,))
    top_contributors = [_norm(dict(r)) for r in c.fetchall()]

    # ── Invoice Outstanding ───────────────────────────────────────────────────
    c.execute("""SELECT COUNT(*) AS cnt, COALESCE(SUM(invoice_amount - paid_amount), 0) AS outstanding
        FROM invoices WHERE paid_amount < invoice_amount AND tahun=%s""", (cur_year,))
    inv = _norm(dict(c.fetchone()))
    outstanding_count  = int(inv.get('cnt') or 0)
    outstanding_amount = float(inv.get('outstanding') or 0)

    # ── Recurring Behind YTD Target ───────────────────────────────────────────
    c.execute("""
        SELECT rp.project_id, rp.product, rp.client, rp.organisasi, rp.pic,
               rp.revenue_target,
               ROUND(rp.revenue_target::numeric * %s / 12, 0) AS target_ytd,
               COALESCE(SUM(i.invoice_amount), 0) AS billed,
               rp.actual_revenue AS collected,
               ROUND(rp.revenue_target::numeric * %s / 12, 0) - rp.actual_revenue AS gap_collected,
               ROUND(rp.revenue_target::numeric * %s / 12, 0) - COALESCE(SUM(i.invoice_amount), 0) AS gap_billed,
               ROUND(rp.actual_revenue / NULLIF(rp.revenue_target, 0) * 100, 1) AS ach_pct
        FROM revenue_projects rp
        LEFT JOIN invoices i ON i.project_id = rp.project_id AND i.tahun = %s
        WHERE rp.kategori = 'Recurring' AND rp.is_active = 1 AND rp.tahun = %s
          AND rp.actual_revenue < ROUND(rp.revenue_target::numeric * %s / 12, 0)
        GROUP BY rp.project_id, rp.product, rp.client, rp.organisasi, rp.pic,
                 rp.revenue_target, rp.actual_revenue
        ORDER BY gap_collected DESC
    """, (cur_month, cur_month, cur_month, cur_year, cur_year, cur_month))
    recurring_behind = [_norm(dict(r)) for r in c.fetchall()]

    conn.close()
    return {
        "cur_year": cur_year, "cur_month": cur_month, "years": years,
        # KPI
        "total_target": total_target, "total_actual": total_actual, "ach_pct": ach_pct,
        "gap_ytd": gap_ytd, "remain_target": remain_target,
        "run_rate": run_rate, "projected_eoy": projected_eoy, "projected_ach": projected_ach,
        # Monthly analysis
        "monthly": monthly,
        "past_months_count": past_months_count,
        "ach_months_count":  ach_months_count,
        "miss_months_count": miss_months_count,
        # Per kategori
        "project_target": project_target, "project_actual": project_actual, "project_ach": project_ach,
        "recurring_target": recurring_target, "recurring_actual": recurring_actual, "recurring_ach": recurring_ach,
        # Per type & risk
        "by_type": by_type, "by_risk": by_risk, "by_status": by_status,
        # Per organisasi
        "fsp_eco_target":  fsp_eco_target,  "fsp_eco_actual":  fsp_eco_actual,  "fsp_eco_ach":  fsp_eco_ach,
        "fsp_core_target": fsp_core_target, "fsp_core_actual": fsp_core_actual, "fsp_core_ach": fsp_core_ach,
        "org_breakdown": org_breakdown,
        # Projects
        "critical_count": critical_count, "total_projects": total_projects,
        "at_risk_projects": at_risk_projects,
        "zero_projects": zero_projects, "zero_count": len(zero_projects),
        "top_contributors": top_contributors,
        "recurring_behind": recurring_behind,
        # Invoice
        "outstanding_count": outstanding_count, "outstanding_amount": outstanding_amount,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# REVENUE TRACKER (Projects)
# ═══════════════════════════════════════════════════════════════════════════════

class TerminItem(BaseModel):
    month: int
    pct: float = 0

class BulananTarget(BaseModel):
    month: int
    target: float = 0

class ProjectCreate(BaseModel):
    lob: str = "DCSS"
    owner: str = ""
    product: str = ""
    client: str = ""
    organisasi: str = ""
    kategori: str = "Project"
    type: str = "One Time"
    tgl_penagihan_pertama: str = ""
    revenue_target: float = 0
    notes: str = ""
    from_lead_id: str = ""
    termins: list[TerminItem] = []
    bulanan_targets: list[BulananTarget] = []


@app.get("/api/v1/revenue/projects", tags=["Revenue"])
def revenue_tracker_api(
    owner: str = "", kategori: str = "", status: str = "",
    search: str = "", tahun: int = Query(0), month: int = Query(0),
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=200),
    user: dict = Depends(require_menu("rev_tracker"))
):
    import math as _math
    cur_year  = tahun if tahun else date.today().year
    cur_month = month if month else 0
    conn = get_conn()
    c = conn.cursor()

    if cur_month:
        month_name_en = MONTHS_EN[cur_month - 1]
        inv_sub  = f"COALESCE((SELECT SUM(paid_amount) FROM invoices i WHERE i.project_id=p.project_id AND i.tahun={cur_year} AND i.period='{month_name_en}'),0)"
        inv_cnt  = f"COALESCE((SELECT COUNT(*) FROM invoices i WHERE i.project_id=p.project_id AND i.tahun={cur_year} AND i.period='{month_name_en}'),0)"
    else:
        inv_sub  = "COALESCE((SELECT SUM(paid_amount) FROM invoices i WHERE i.project_id=p.project_id),0)"
        inv_cnt  = "COALESCE((SELECT COUNT(*) FROM invoices i WHERE i.project_id=p.project_id),0)"

    is_sales = user.get("role_id") == 3
    q = f"""SELECT p.*,
        {inv_sub} as invoice_actual,
        {inv_cnt} as invoice_count,
        CASE WHEN p.revenue_target > 0 THEN CAST({inv_sub} AS FLOAT) / p.revenue_target ELSE 0 END as achievement_pct
        FROM revenue_projects p WHERE p.is_active=1 AND p.tahun=%s"""
    params = [cur_year]
    if is_sales:
        q += " AND p.pic=%s"; params.append(user["nama"])
    elif owner:  q += " AND p.pic=%s"; params.append(owner)
    if kategori: q += " AND p.kategori=%s"; params.append(kategori)
    if status:   q += " AND p.status=%s"; params.append(status)
    if search:   q += " AND (p.product LIKE %s OR p.client LIKE %s)"; params += [f"%{search}%"]*2
    effective_owner = user["nama"] if is_sales else owner
    c.execute(f"SELECT COUNT(*) cnt FROM revenue_projects p WHERE p.is_active=1 AND p.tahun=%s" +
              (" AND p.pic=%s" if effective_owner else "") + (" AND p.kategori=%s" if kategori else "") +
              (" AND p.status=%s" if status else "") + (" AND (p.product LIKE %s OR p.client LIKE %s)" if search else ""),
              [cur_year] + ([effective_owner] if effective_owner else []) + ([kategori] if kategori else []) +
              ([status] if status else []) + ([f"%{search}%"]*2 if search else []))
    total_projects = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(q + f" ORDER BY p.project_id LIMIT {per_page} OFFSET {offset}", params)
    projects = [_norm(dict(r)) for r in c.fetchall()]

    # ── Enrich per-project invoice summary ───────────────────────────────────
    project_ids = [p['project_id'] for p in projects]
    inv_map: dict = {}
    if project_ids:
        placeholders = ','.join(['%s'] * len(project_ids))
        c.execute(f"""
            SELECT project_id,
                   COUNT(*) AS total_inv,
                   COALESCE(SUM(invoice_amount),0) AS total_amount,
                   COALESCE(SUM(paid_amount),0) AS total_paid,
                   COUNT(*) FILTER (WHERE paid_amount >= invoice_amount AND invoice_amount > 0) AS lunas_count,
                   COUNT(*) FILTER (WHERE paid_amount < invoice_amount) AS belum_count
            FROM invoices WHERE project_id IN ({placeholders}) AND tahun=%s
            GROUP BY project_id
        """, project_ids + [cur_year])
        for r in c.fetchall():
            r = _norm(dict(r))
            outstanding = float(r['total_amount'] or 0) - float(r['total_paid'] or 0)
            inv_map[r['project_id']] = {
                'total_inv':   int(r['total_inv'] or 0),
                'total_amount': float(r['total_amount'] or 0),
                'total_paid':   float(r['total_paid'] or 0),
                'lunas_count':  int(r['lunas_count'] or 0),
                'belum_count':  int(r['belum_count'] or 0),
                'outstanding':  outstanding,
            }

    real_cur_month = date.today().month if cur_year == date.today().year else 12
    for p in projects:
        p['inv'] = inv_map.get(p['project_id'])
        # YTD achievement untuk proyek Bulanan/Recurring
        p['is_ytd'] = p.get('kategori') == 'Recurring' and not cur_month
        p['cur_month'] = real_cur_month
        if p['is_ytd'] and p.get('revenue_target'):
            ytd_target_share = float(p['revenue_target']) * real_cur_month / 12
            p['ytd_ach_pct'] = round(float(p.get('invoice_actual') or 0) / max(ytd_target_share, 1) * 100, 1)
        else:
            p['ytd_ach_pct'] = round(float(p.get('achievement_pct') or 0) * 100, 1)
        # target_period_label & invoice_period_status (dari target_invoice_date jika ada)
        tid = p.get('target_invoice_date')
        if tid:
            import datetime as _dt
            try:
                td = _dt.date.fromisoformat(str(tid)[:10])
                p['target_period_label'] = td.strftime('%b %Y')
                if p.get('inv') and p['inv']['total_inv'] > 0:
                    p['actual_period_label'] = p['target_period_label']
                    p['invoice_period_status'] = 'Sudah Terbit' if td <= date.today() else 'Belum Jatuh Tempo'
                else:
                    p['actual_period_label'] = None
                    p['invoice_period_status'] = 'Terlambat' if td < date.today() else 'Belum Jatuh Tempo'
            except Exception:
                p['target_period_label'] = None
                p['actual_period_label'] = None
                p['invoice_period_status'] = None
        else:
            p['target_period_label'] = None
            p['actual_period_label'] = None
            p['invoice_period_status'] = None

    c.execute("SELECT DISTINCT pic FROM revenue_projects WHERE pic IS NOT NULL ORDER BY pic")
    owners = [r['pic'] for r in c.fetchall()]
    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years  = [r['tahun'] for r in c.fetchall() if r['tahun']]

    # Summary totals
    if cur_month:
        c.execute("""SELECT COALESCE(SUM(i.paid_amount),0) a FROM invoices i
            JOIN revenue_projects p ON i.project_id=p.project_id
            WHERE p.is_active=1 AND i.tahun=%s AND i.period=%s""", (cur_year, MONTHS_EN[cur_month-1]))
        ta = float(c.fetchone()['a'] or 0)
    else:
        c.execute("SELECT COALESCE(SUM(actual_revenue),0) a FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
        ta = float(c.fetchone()['a'] or 0)
    c.execute("SELECT COALESCE(SUM(revenue_target),0) t FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    tt = float(c.fetchone()['t'] or 0)

    # ── Invoice Summary ───────────────────────────────────────────────────────
    c.execute("""
        SELECT
            COUNT(*)                                          AS total_inv,
            COALESCE(SUM(invoice_amount), 0)                  AS total_amount,
            COALESCE(SUM(paid_amount), 0)                     AS total_paid,
            COUNT(*) FILTER (WHERE paid_amount >= invoice_amount AND invoice_amount > 0) AS lunas_count,
            COUNT(*) FILTER (WHERE paid_amount < invoice_amount)  AS belum_count
        FROM invoices WHERE tahun=%s
    """, (cur_year,))
    iv = _norm(dict(c.fetchone()))
    iv_total_amount = float(iv['total_amount'] or 0)
    iv_total_paid   = float(iv['total_paid']   or 0)
    iv_outstanding  = iv_total_amount - iv_total_paid
    iv_collection   = round(iv_total_paid / max(iv_total_amount, 1) * 100, 1) if iv_total_amount else 0
    inv_summary = {
        "total_inv":          int(iv['total_inv']   or 0),
        "total_amount":       iv_total_amount,
        "total_paid":         iv_total_paid,
        "outstanding":        iv_outstanding,
        "lunas_count":        int(iv['lunas_count'] or 0),
        "belum_count":        int(iv['belum_count'] or 0),
        "collection_rate":    iv_collection,
    }

    c.execute("""
        SELECT i.id, i.project_id, i.invoice_no, i.invoice_date, i.period,
               i.invoice_amount, i.paid_amount,
               (i.invoice_amount - i.paid_amount) AS outstanding,
               p.client
        FROM invoices i
        JOIN revenue_projects p ON i.project_id = p.project_id
        WHERE i.tahun=%s AND i.paid_amount < i.invoice_amount AND p.is_active=1
        ORDER BY i.invoice_date ASC
        LIMIT 20
    """, (cur_year,))
    unpaid_invoices = [_norm(dict(r)) for r in c.fetchall()]

    conn.close()

    return {
        "projects": projects, "owners": owners, "years": years,
        "cur_year": cur_year, "cur_month": cur_month,
        "total_target": tt, "total_actual": ta,
        "ach_pct": round(ta / max(tt, 1) * 100, 1) if tt else 0,
        "total_projects": total_projects,
        "months": [{"num": i+1, "name": m} for i, m in enumerate(MONTHS_EN)],
        "inv_summary": inv_summary, "unpaid_invoices": unpaid_invoices,
        "total": total_projects, "page": page, "per_page": per_page,
        "total_pages": _math.ceil(total_projects/per_page) if total_projects else 1,
    }


@app.post("/api/v1/revenue/projects", status_code=201, tags=["Revenue"])
def project_create(payload: ProjectCreate, user: dict = Depends(require_menu("rev_tracker"))):
    import datetime as _dt
    try:
        billing_date = _dt.date.fromisoformat(payload.tgl_penagihan_pertama[:10]) if payload.tgl_penagihan_pertama else None
    except Exception:
        billing_date = None
    yr          = billing_date.year  if billing_date else _dt.date.today().year
    start_month = billing_date.month if billing_date else _dt.date.today().month

    pid = next_rev_id()
    target = float(payload.revenue_target or 0)
    st, rk = auto_status_risk(target, 0.0, payload.type, yr)
    conn = get_conn(); c = conn.cursor()
    notes_txt = payload.notes
    if payload.from_lead_id:
        notes_txt = f"Dari Pipeline CRM: {payload.from_lead_id}. {notes_txt}"

    c.execute("""INSERT INTO revenue_projects
        (project_id,lob,pic,product,client,organisasi,kategori,type,tahun,
         revenue_target,actual_revenue,status,risk_level,notes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (project_id) DO NOTHING""",
        (pid, payload.lob, payload.owner, payload.product, payload.client, payload.organisasi,
         payload.kategori, payload.type, yr, target, 0, st, rk, notes_txt))

    type_lower = payload.type.lower()
    if 'bulanan' in type_lower or 'tahunan' in type_lower:
        bulanan_map = {bt.month: bt.target for bt in payload.bulanan_targets} if payload.bulanan_targets else {}
        for m in range(start_month, 13):
            c.execute("""INSERT INTO revenue_monthly (project_id,month_num,month_name,target,actual,termin_no)
                VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (project_id,month_num) DO NOTHING""",
                (pid, m, MONTHS_EN[m-1], float(bulanan_map.get(m, 0)), 0, m - start_month + 1))
    elif 'termin' in type_lower and payload.termins:
        for i, t in enumerate(payload.termins):
            m = max(1, min(12, int(t.month)))
            c.execute("""INSERT INTO revenue_monthly (project_id,month_num,month_name,target,actual,termin_no)
                VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (project_id,month_num) DO UPDATE SET target=EXCLUDED.target,termin_no=EXCLUDED.termin_no""",
                (pid, m, MONTHS_EN[m-1], round(target * float(t.pct) / 100, 2), 0, i + 1))
    else:
        c.execute("""INSERT INTO revenue_monthly (project_id,month_num,month_name,target,actual,termin_no)
            VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (project_id,month_num) DO NOTHING""",
            (pid, start_month, MONTHS_EN[start_month-1], target, 0, 1))

    conn.commit(); conn.close()
    return {"message": "Proyek revenue berhasil dibuat.", "project_id": pid}


class ProjectUpdate(BaseModel):
    client: str = ""
    product: str = ""
    organisasi: str = ""
    lob: str = ""
    kategori: str = "Project"
    type: str = "One Time"
    target_invoice_date: str = ""
    tahun: int = 0
    revenue_target: float = 0
    status: str = ""
    risk_level: str = ""
    notes: str = ""


@app.put("/api/v1/revenue/projects/{pid}", tags=["Revenue"])
def project_update(pid: str, payload: ProjectUpdate, user: dict = Depends(require_menu("rev_tracker"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT project_id, actual_revenue, type, tahun FROM revenue_projects WHERE project_id=%s", (pid,))
    row = c.fetchone()
    if not row:
        conn.close(); raise HTTPException(404, "Proyek tidak ditemukan")
    yr = payload.tahun or row['tahun'] or date.today().year
    actual = float(row['actual_revenue'] or 0)
    target = float(payload.revenue_target or 0)
    st, rk = auto_status_risk(target, actual, payload.type or row['type'], yr)
    if payload.status: st = payload.status
    if payload.risk_level: rk = payload.risk_level
    tid = payload.target_invoice_date or None
    c.execute("""UPDATE revenue_projects SET
        client=%s, product=%s, organisasi=%s, lob=%s, kategori=%s, type=%s,
        target_invoice_date=%s, tahun=%s, revenue_target=%s,
        status=%s, risk_level=%s, notes=%s
        WHERE project_id=%s""",
        (payload.client, payload.product, payload.organisasi, payload.lob,
         payload.kategori, payload.type, tid, yr, target, st, rk, payload.notes, pid))
    conn.commit(); conn.close()
    return {"message": "Proyek berhasil diupdate.", "project_id": pid}


@app.patch("/api/v1/revenue/projects/{pid}/status", tags=["Revenue"])
def project_patch_status(pid: str, payload: dict, user: dict = Depends(require_menu("rev_tracker"))):
    ps = payload.get("project_status")
    allowed = ["Active", "On Hold", "Completed", "Failed"]
    if ps not in allowed:
        raise HTTPException(422, "Status tidak valid.")
    conn = get_conn(); c = conn.cursor()
    c.execute("UPDATE revenue_projects SET project_status=%s WHERE project_id=%s", (ps, pid))
    conn.commit(); conn.close()
    return {"ok": True}


@app.delete("/api/v1/revenue/projects/{pid}", tags=["Revenue"])
def project_delete(pid: str, user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    c.execute("UPDATE revenue_projects SET is_active=0 WHERE project_id=%s", (pid,))
    conn.commit(); conn.close()
    return {"message": "Proyek dipindahkan ke recycle bin."}


@app.get("/api/v1/revenue/trashed", tags=["Revenue"])
def revenue_trashed(user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM revenue_projects WHERE is_active=0 ORDER BY project_id DESC")
    trashed = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"trashed": trashed, "count": len(trashed)}


@app.post("/api/v1/revenue/trashed/{pid}/restore", tags=["Revenue"])
def project_restore(pid: str, user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    c.execute("UPDATE revenue_projects SET is_active=1 WHERE project_id=%s", (pid,))
    conn.commit(); conn.close()
    return {"message": "Proyek berhasil dipulihkan."}


@app.delete("/api/v1/revenue/trashed/{pid}/permanent", tags=["Revenue"])
def project_delete_permanent(pid: str, user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    # Pastikan project benar-benar di recycle bin (is_active=0 / deleted_at IS NOT NULL)
    c.execute("SELECT project_id FROM revenue_projects WHERE project_id=%s AND (is_active=0 OR deleted_at IS NOT NULL)", (pid,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404, "Proyek tidak ditemukan di recycle bin.")
    # Hapus invoice terkait lebih dulu (FK)
    c.execute("DELETE FROM revenue_monthly WHERE project_id=%s", (pid,))
    c.execute("DELETE FROM invoices WHERE project_id=%s", (pid,))
    c.execute("DELETE FROM revenue_projects WHERE project_id=%s", (pid,))
    conn.commit(); conn.close()
    return {"message": "Proyek berhasil dihapus permanen."}


@app.get("/api/v1/revenue/won-leads", tags=["Revenue"])
def revenue_won_leads(user: dict = Depends(require_menu("rev_tracker"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        SELECT l.lead_id, l.nama_company, l.product, l.sales_owner,
               COALESCE(NULLIF(l.deal_value,0), l.propose_value) AS deal_value,
               l.remarks, l.segmen, l.tgl_masuk,
               CASE WHEN rp.project_id IS NOT NULL THEN TRUE ELSE FALSE END AS is_imported,
               CASE WHEN l.segmen IN ('Recurring','Maintenance') THEN 'Recurring' ELSE 'Project' END AS suggested_kategori,
               CASE WHEN l.segmen IN ('Recurring','Maintenance') THEN 'Annual' ELSE 'One Time' END AS suggested_type,
               EXTRACT(YEAR FROM COALESCE(l.exp_close_date, CURRENT_DATE))::int AS suggested_tahun
        FROM leads l
        LEFT JOIN revenue_projects rp ON rp.notes LIKE '%' || l.lead_id || '%' AND rp.is_active=1
        WHERE l.stage = 'Won' AND COALESCE(l.won_import_excluded, FALSE) = FALSE
        ORDER BY l.updated_at DESC
    """)
    leads = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    pending = sum(1 for l in leads if not l['is_imported'])
    return {"leads": leads, "pending": pending}


@app.delete("/api/v1/revenue/won-leads/{lead_id}", tags=["Revenue"])
def exclude_won_lead(lead_id: str, user: dict = Depends(require_menu("rev_tracker"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("UPDATE leads SET won_import_excluded=TRUE WHERE lead_id=%s AND stage='Won'", (lead_id,))
    conn.commit(); conn.close()
    return {"ok": True}


class ImportWonItem(BaseModel):
    lead_id: str
    client: str = ""
    product: str = ""
    owner: str = ""
    lob: str = "DCSS"
    organisasi: str = ""
    kategori: str = "Project"
    type: str = "One Time"
    tgl_penagihan_pertama: str = ""
    revenue_target: float = 0
    notes: str = ""
    termins: list[TerminItem] = []
    bulanan_targets: list[BulananTarget] = []


class ImportWonPayload(BaseModel):
    items: list[ImportWonItem]


@app.get("/api/v1/revenue/export-invoice-template", tags=["Revenue"])
def revenue_export_invoice_template(
    tahun: int = Query(0),
    user: dict = Depends(require_menu("rev_tracker"))
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    cur_year = tahun if tahun else date.today().year
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        SELECT project_id, lob, product, client, organisasi, tahun, revenue_target
        FROM revenue_projects
        WHERE is_active=1 AND tahun=%s
        ORDER BY project_id
    """, (cur_year,))
    projects = [dict(r) for r in c.fetchall()]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Invoice"

    headers = ["project_id","lob","product","client","organisasi",
               "invoice_no","invoice_date","period","tahun",
               "invoice_amount","paid_amount","paid_date","notes"]

    header_fill   = PatternFill("solid", fgColor="1F4E79")
    ref_fill      = PatternFill("solid", fgColor="D6E4F0")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    ref_font      = Font(color="1F4E79", italic=True)
    center        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="BFBFBF")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [14,10,20,25,20,20,14,16,8,18,18,14,20]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20

    # Kolom referensi (pre-filled dari revenue_projects, tidak perlu diubah)
    ref_cols = {1,2,3,4,5,9}  # project_id,lob,product,client,organisasi,tahun

    for row_idx, p in enumerate(projects, 2):
        values = [
            p.get("project_id"), p.get("lob"), p.get("product"),
            p.get("client"), p.get("organisasi"),
            "",  # invoice_no
            "",  # invoice_date
            "",  # period
            p.get("tahun"),
            "",  # invoice_amount
            "",  # paid_amount
            "",  # paid_date
            "",  # notes
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx in ref_cols:
                cell.fill = ref_fill
                cell.font = ref_font
                cell.alignment = center

    # Sheet petunjuk
    ws2 = wb.create_sheet("Petunjuk")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 55
    notes_data = [
        ("Kolom", "Keterangan"),
        ("project_id", "ID project dari Revenue Tracker (format REV-XXXX). Jangan diubah."),
        ("lob", "Line of Business. Sudah terisi, bisa diubah jika perlu."),
        ("product", "Nama produk. Sudah terisi dari data project."),
        ("client", "Nama client. Sudah terisi dari data project."),
        ("organisasi", "Organisasi. Sudah terisi dari data project."),
        ("invoice_no", "Nomor invoice. Wajib diisi. Contoh: INV/2026/001"),
        ("invoice_date", "Tanggal terbit invoice. Format: YYYY-MM-DD"),
        ("period", "Periode tagihan. Contoh: January 2026"),
        ("tahun", "Tahun. Sudah terisi otomatis."),
        ("invoice_amount", "Nilai invoice (angka, tanpa titik/koma)."),
        ("paid_amount", "Nilai yang sudah dibayar (0 jika belum)."),
        ("paid_date", "Tanggal pembayaran. Format: YYYY-MM-DD. Kosongkan jika belum bayar."),
        ("notes", "Catatan opsional."),
    ]
    title_font = Font(bold=True, color="1F4E79")
    for r_idx, (col_name, desc) in enumerate(notes_data, 1):
        ws2.cell(row=r_idx, column=1, value=col_name).font = title_font if r_idx == 1 else Font(bold=True)
        ws2.cell(row=r_idx, column=2, value=desc)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"template_invoice_{cur_year}.xlsx"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/v1/revenue/projects/{pid}/monthly", tags=["Revenue"])
def project_monthly_detail(pid: str, user: dict = Depends(require_menu("rev_tracker"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM revenue_monthly WHERE project_id=%s ORDER BY month_num", (pid,))
    monthly = [_norm(dict(r)) for r in c.fetchall()]
    total_target = sum(float(r['target'] or 0) for r in monthly)
    total_actual = sum(float(r['actual'] or 0) for r in monthly)
    for r in monthly:
        t = float(r['target'] or 0)
        a = float(r['actual'] or 0)
        r['ach_pct'] = round(a / max(t, 1) * 100, 1) if t else 0
    conn.close()
    return {"monthly": monthly, "total_target": total_target, "total_actual": total_actual}


@app.post("/api/v1/revenue/import-won", status_code=201, tags=["Revenue"])
def revenue_import_won(payload: ImportWonPayload, user: dict = Depends(require_menu("rev_tracker"))):
    import datetime as _dt
    conn = get_conn(); c = conn.cursor()
    created = []
    for item in payload.items:
        # Derive tahun & start_month dari tgl_penagihan_pertama
        try:
            billing_date = _dt.date.fromisoformat(item.tgl_penagihan_pertama[:10]) if item.tgl_penagihan_pertama else None
        except Exception:
            billing_date = None
        yr          = billing_date.year  if billing_date else _dt.date.today().year
        start_month = billing_date.month if billing_date else _dt.date.today().month

        target = float(item.revenue_target or 0)
        st, rk = auto_status_risk(target, 0.0, item.type, yr)
        pid = next_rev_id()
        notes_txt = f"Dari Pipeline CRM: {item.lead_id}. {item.notes}".strip(". ")
        c.execute("""INSERT INTO revenue_projects
            (project_id,lob,pic,product,client,organisasi,kategori,type,tahun,
             revenue_target,actual_revenue,status,risk_level,notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (project_id) DO NOTHING""",
            (pid, item.lob, item.owner, item.product, item.client, item.organisasi,
             item.kategori, item.type, yr, target, 0, st, rk, notes_txt))

        # Generate revenue_monthly berdasarkan type
        type_lower = item.type.lower()
        if 'bulanan' in type_lower:
            # Gunakan bulanan_targets jika dikirim, fallback ke generate kosong
            bulanan_map = {bt.month: bt.target for bt in item.bulanan_targets} if item.bulanan_targets else {}
            for m in range(start_month, 13):
                month_target = float(bulanan_map.get(m, 0))
                c.execute("""
                    INSERT INTO revenue_monthly (project_id, month_num, month_name, target, actual, termin_no)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (project_id, month_num) DO NOTHING
                """, (pid, m, MONTHS_EN[m-1], month_target, 0, m - start_month + 1))
        elif 'termin' in type_lower and item.termins:
            # Buat record per termin dengan target proporsional
            for i, t in enumerate(item.termins):
                m = max(1, min(12, int(t.month)))
                termin_target = round(target * float(t.pct) / 100, 2)
                c.execute("""
                    INSERT INTO revenue_monthly (project_id, month_num, month_name, target, actual, termin_no)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (project_id, month_num) DO UPDATE SET target = EXCLUDED.target, termin_no = EXCLUDED.termin_no
                """, (pid, m, MONTHS_EN[m-1], termin_target, 0, i + 1))
        else:
            # One Time: 1 record di bulan penagihan pertama
            c.execute("""
                INSERT INTO revenue_monthly (project_id, month_num, month_name, target, actual, termin_no)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON CONFLICT (project_id, month_num) DO NOTHING
            """, (pid, start_month, MONTHS_EN[start_month-1], target, 0, 1))

        created.append(pid)
    conn.commit(); conn.close()
    return {"message": f"{len(created)} proyek berhasil diimport.", "created": created}


# ═══════════════════════════════════════════════════════════════════════════════
# REVENUE MONTHLY
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/revenue/monthly", tags=["Revenue"])
def revenue_monthly_api(
    month: int = Query(0), tahun: int = Query(0),
    owner: str = "", kategori: str = "",
    user: dict = Depends(require_menu("rev_monthly"))
):
    cur_year    = tahun if tahun else date.today().year
    cur_month_now = date.today().month if cur_year == date.today().year else 12
    conn = get_conn()
    c = conn.cursor()

    # ── Summary: satu baris per bulan (12 baris) ──────────────────────────────
    c.execute("""
        SELECT m.month_num, m.month_name,
               COALESCE(SUM(m.target), 0)                          AS target,
               COALESCE(SUM(m.actual), 0)                          AS actual
        FROM revenue_monthly m
        JOIN revenue_projects p ON m.project_id = p.project_id
        WHERE p.is_active = 1 AND p.tahun = %s
        GROUP BY m.month_num, m.month_name
        ORDER BY m.month_num
    """, (cur_year,))
    monthly_rows = {r['month_num']: _norm(dict(r)) for r in c.fetchall()}

    # Collection per bulan dari invoices
    c.execute("""
        SELECT EXTRACT(MONTH FROM invoice_date)::int AS month_num,
               COALESCE(SUM(invoice_amount), 0)      AS invoiced,
               COALESCE(SUM(paid_amount), 0)         AS collected
        FROM invoices WHERE tahun = %s
        GROUP BY 1
    """, (cur_year,))
    inv_rows = {r['month_num']: _norm(dict(r)) for r in c.fetchall()}

    # Project count per bulan
    c.execute("""
        SELECT m.month_num, COUNT(DISTINCT m.project_id) AS cnt
        FROM revenue_monthly m
        JOIN revenue_projects p ON m.project_id = p.project_id
        WHERE p.is_active = 1 AND p.tahun = %s AND m.target > 0
        GROUP BY m.month_num
    """, (cur_year,))
    proj_count = {r['month_num']: r['cnt'] for r in c.fetchall()}

    QUARTERS = {1:'Q1',2:'Q1',3:'Q1',4:'Q2',5:'Q2',6:'Q2',7:'Q3',8:'Q3',9:'Q3',10:'Q4',11:'Q4',12:'Q4'}
    MONTH_NAMES_ID = ['Januari','Februari','Maret','April','Mei','Juni',
                      'Juli','Agustus','September','Oktober','November','Desember']

    summary = []
    for mn in range(1, 13):
        row = monthly_rows.get(mn, {})
        inv = inv_rows.get(mn, {})
        target     = float(row.get('target', 0) or 0)
        actual     = float(row.get('actual', 0) or 0)
        collected  = float(inv.get('collected', 0) or 0)
        invoiced   = float(inv.get('invoiced', 0) or 0)
        outstanding = max(0, invoiced - collected)
        ach_pct    = round(actual / max(target, 1) * 100, 1) if target else 0
        is_past    = mn < cur_month_now
        is_current = mn == cur_month_now
        if ach_pct >= 80:    status = 'On Track'
        elif ach_pct >= 50:  status = 'At Risk'
        else:                status = 'Critical' if (is_past or is_current) else 'Upcoming'
        summary.append({
            'month_num':     mn,
            'month_name':    MONTH_NAMES_ID[mn-1],
            'quarter':       QUARTERS[mn],
            'target':        target,
            'actual':        actual,
            'collection':    collected,
            'outstanding':   outstanding,
            'ach_pct':       ach_pct,
            'status':        status,
            'is_past':       is_past,
            'is_current':    is_current,
            'project_count': proj_count.get(mn, 0),
        })

    # Grand totals
    grand_target = sum(r['target']     for r in summary)
    grand_actual = sum(r['actual']     for r in summary)
    grand_coll   = sum(r['collection'] for r in summary)
    grand_out    = sum(r['outstanding']for r in summary)
    grand_ach    = round(grand_actual / max(grand_target, 1) * 100, 1) if grand_target else 0

    # ── Detail rows: proyek per bulan jika month dipilih ──────────────────────
    detail_rows = []
    if month:
        month_name_en = MONTHS_EN[month - 1]
        q = """
            SELECT p.project_id, p.product, p.client, p.pic, p.kategori,
                   p.organisasi, p.status, p.risk_level,
                   m.target,
                   COALESCE((SELECT SUM(i.paid_amount) FROM invoices i
                             WHERE i.project_id=p.project_id AND i.tahun=%s
                               AND EXTRACT(MONTH FROM i.invoice_date)=%s), 0) AS actual
            FROM revenue_monthly m
            JOIN revenue_projects p ON m.project_id = p.project_id
            WHERE m.month_num = %s AND p.is_active = 1 AND p.tahun = %s
        """
        params = [cur_year, month, month, cur_year]
        if owner:    q += " AND p.pic=%s"; params.append(owner)
        if kategori: q += " AND p.kategori=%s"; params.append(kategori)
        c.execute(q + " ORDER BY p.project_id", params)
        for r in c.fetchall():
            r = _norm(dict(r))
            t = float(r['target'] or 0)
            a = float(r['actual'] or 0)
            r['ach_pct'] = round(a / max(t, 1) * 100, 1) if t else 0
            detail_rows.append(r)

    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]
    conn.close()

    return {
        "summary":      summary,
        "detail_rows":  detail_rows,
        "grand_target": grand_target,
        "grand_actual": grand_actual,
        "grand_coll":   grand_coll,
        "grand_out":    grand_out,
        "grand_ach":    grand_ach,
        "cur_year":     cur_year,
        "years":        years,
    }


@app.get("/api/v1/revenue/project-monthly", tags=["Revenue"])
def revenue_project_monthly_api(
    tahun: int = Query(0), organisasi: str = "", kategori: str = "",
    search: str = "",
    user: dict = Depends(require_menu("rev_monthly"))
):
    cur_year = tahun if tahun else date.today().year
    cur_month = date.today().month if cur_year == date.today().year else 12
    MONTH_NAMES_ID = ['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Agu','Sep','Okt','Nov','Des']
    conn = get_conn(); c = conn.cursor()

    # Semua proyek aktif tahun ini
    q = "SELECT project_id, product, client, pic, kategori, organisasi, revenue_target FROM revenue_projects WHERE is_active=1 AND tahun=%s"
    params = [cur_year]
    if organisasi: q += " AND organisasi=%s"; params.append(organisasi)
    if kategori:   q += " AND kategori=%s";   params.append(kategori)
    if search:     q += " AND (product LIKE %s OR client LIKE %s)"; params += [f"%{search}%"]*2
    q += " ORDER BY organisasi, product"
    c.execute(q, params)
    projects = [_norm(dict(r)) for r in c.fetchall()]

    if not projects:
        conn.close()
        return {"projects": [], "months": MONTH_NAMES_ID, "cur_year": cur_year, "cur_month": cur_month,
                "grand_target": 0, "grand_actual": 0, "month_totals": [], "years": [cur_year]}

    pids = [p['project_id'] for p in projects]
    placeholders = ','.join(['%s'] * len(pids))

    # Monthly target & actual per proyek
    c.execute(f"""
        SELECT project_id, month_num, COALESCE(target,0) AS target, COALESCE(actual,0) AS actual
        FROM revenue_monthly WHERE project_id IN ({placeholders})
        ORDER BY project_id, month_num
    """, pids)
    monthly_map: dict = {}
    for r in c.fetchall():
        monthly_map.setdefault(r['project_id'], {})[r['month_num']] = {
            'target': float(r['target'] or 0), 'actual': float(r['actual'] or 0)
        }

    # Enrich projects
    for p in projects:
        months = []
        for mn in range(1, 13):
            m = monthly_map.get(p['project_id'], {}).get(mn, {})
            t = m.get('target', 0); a = m.get('actual', 0)
            months.append({'month_num': mn, 'target': t, 'actual': a,
                           'ach': round(a / max(t, 1) * 100, 1) if t else 0})
        p['months'] = months
        p['total_target'] = sum(m['target'] for m in months)
        p['total_actual'] = sum(m['actual'] for m in months)
        p['total_ach']    = round(p['total_actual'] / max(p['total_target'], 1) * 100, 1) if p['total_target'] else 0

    # Month totals (aggregate semua proyek)
    month_totals = []
    for mn in range(1, 13):
        t = sum(p['months'][mn-1]['target'] for p in projects)
        a = sum(p['months'][mn-1]['actual'] for p in projects)
        month_totals.append({'month_num': mn, 'label': MONTH_NAMES_ID[mn-1],
                             'target': t, 'actual': a,
                             'ach': round(a / max(t, 1) * 100, 1) if t else 0,
                             'is_past': mn < cur_month, 'is_current': mn == cur_month})

    grand_target = sum(p['total_target'] for p in projects)
    grand_actual = sum(p['total_actual'] for p in projects)

    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]
    c.execute("SELECT DISTINCT organisasi FROM revenue_projects WHERE is_active=1 ORDER BY organisasi")
    org_list = [r['organisasi'] for r in c.fetchall() if r['organisasi']]
    conn.close()

    return {"projects": projects, "months": MONTH_NAMES_ID, "month_totals": month_totals,
            "grand_target": grand_target, "grand_actual": grand_actual,
            "grand_ach": round(grand_actual / max(grand_target, 1) * 100, 1) if grand_target else 0,
            "cur_year": cur_year, "cur_month": cur_month, "years": years, "org_list": org_list}


class MonthlyUpsert(BaseModel):
    project_id: str
    month_num: int
    month_name: str = ""
    target: float = 0
    actual: float = 0
    termin_no: int = 0

@app.post("/api/v1/revenue/monthly/upsert", tags=["Revenue"])
def monthly_upsert(payload: MonthlyUpsert, user: dict = Depends(require_menu("rev_tracker"))):
    conn = get_conn(); c = conn.cursor()
    month_name = payload.month_name or MONTHS_EN[payload.month_num - 1]
    termin_no  = payload.termin_no or payload.month_num
    c.execute("""
        INSERT INTO revenue_monthly (project_id, month_num, month_name, target, actual, termin_no)
        VALUES (%s,%s,%s,%s,%s,%s)
        ON CONFLICT (project_id, month_num) DO UPDATE
        SET target=EXCLUDED.target, actual=EXCLUDED.actual, month_name=EXCLUDED.month_name
    """, (payload.project_id, payload.month_num, month_name,
          payload.target, payload.actual, termin_no))
    conn.commit(); conn.close()
    return {"message": "OK"}


# ═══════════════════════════════════════════════════════════════════════════════
# INVOICES
# ═══════════════════════════════════════════════════════════════════════════════

class InvoiceCreate(BaseModel):
    project_id: str = ""
    client: str = ""
    product: str = ""
    owner: str = ""
    invoice_no: str = ""
    invoice_date: str = ""
    period: str = ""
    invoice_amount: float = 0
    paid_amount: float = 0
    paid_date: str = ""
    notes: str = ""
    tahun: int = 0


class InvoicePay(BaseModel):
    paid_amount: float
    paid_date: str = ""


@app.get("/api/v1/revenue/invoices", tags=["Revenue"])
def invoice_list(
    period: str = "", status: str = "", search: str = "",
    tahun: int = Query(0), project_id: str = "",
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=200),
    user: dict = Depends(require_menu("rev_invoice"))
):
    import math as _math
    cur_year = tahun if tahun else date.today().year
    is_sales = user.get("role_id") == 3
    conn = get_conn()
    c = conn.cursor()
    base = "FROM invoices i LEFT JOIN revenue_projects p ON i.project_id=p.project_id WHERE i.tahun=%s"
    params = [cur_year]
    if is_sales: base += " AND p.pic=%s"; params.append(user["nama"])
    if period:     base += " AND i.period=%s"; params.append(period)
    if project_id: base += " AND i.project_id=%s"; params.append(project_id)
    if status == 'Lunas':   base += " AND i.paid_amount >= i.invoice_amount AND i.invoice_amount > 0"
    elif status == 'Belum': base += " AND (i.paid_amount < i.invoice_amount)"
    if search: base += " AND (i.client LIKE %s OR i.product LIKE %s OR i.invoice_no LIKE %s)"; params += [f"%{search}%"]*3
    c.execute(f"SELECT COUNT(*) cnt {base}", params)
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(f"SELECT i.*, p.product as proj_product, p.client as proj_client {base} ORDER BY i.invoice_date DESC LIMIT %s OFFSET %s", params + [per_page, offset])
    invoices = [_norm(dict(r)) for r in c.fetchall()]
    for inv in invoices:
        if (inv['paid_amount'] or 0) >= (inv['invoice_amount'] or 0) and (inv['invoice_amount'] or 0) > 0:
            inv['display_status'] = 'Lunas'
        elif not inv['paid_date']:
            inv['display_status'] = 'Belum Dibayar'
        else:
            inv['display_status'] = 'Partial'

    c.execute("SELECT SUM(invoice_amount) t, SUM(paid_amount) p FROM invoices WHERE tahun=%s", (cur_year,))
    totals = _norm(dict(c.fetchone()))
    c.execute("SELECT DISTINCT period FROM invoices WHERE period IS NOT NULL ORDER BY period")
    periods = [r['period'] for r in c.fetchall()]
    c.execute("SELECT project_id, product, client FROM revenue_projects WHERE is_active=1 ORDER BY project_id")
    rev_projects = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT DISTINCT tahun FROM invoices ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]
    conn.close()

    return {
        "invoices": invoices, "periods": periods, "rev_projects": rev_projects,
        "years": years, "cur_year": cur_year,
        "total_invoice": float(totals['t'] or 0), "total_paid": float(totals['p'] or 0),
        "total": total, "page": page, "per_page": per_page,
        "total_pages": _math.ceil(total/per_page) if total else 1,
    }


@app.post("/api/v1/revenue/invoices", status_code=201, tags=["Revenue"])
def invoice_create(payload: InvoiceCreate, user: dict = Depends(require_menu("rev_invoice"))):
    conn = get_conn()
    c = conn.cursor()
    yr = payload.tahun if payload.tahun else date.today().year
    product = payload.product; client = payload.client
    if payload.project_id:
        c.execute("SELECT product, client FROM revenue_projects WHERE project_id=%s", (payload.project_id,))
        row = c.fetchone()
        if row:
            if not product: product = row['product']
            if not client:  client  = row['client']
    c.execute("""INSERT INTO invoices
        (project_id,lob,product,client,invoice_no,invoice_date,period,
         invoice_amount,paid_amount,paid_date,notes,tahun)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (payload.project_id or None, 'DCSS', product, client,
         payload.invoice_no, payload.invoice_date or None, payload.period,
         payload.invoice_amount, payload.paid_amount,
         payload.paid_date or None, payload.notes, yr))
    conn.commit()
    if payload.project_id:
        sync_project_status(payload.project_id)
    conn.close()
    return {"message": "Invoice berhasil dibuat."}


@app.put("/api/v1/revenue/invoices/{iid}", tags=["Revenue"])
def invoice_update(iid: int, payload: InvoiceCreate, user: dict = Depends(require_menu("rev_invoice"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT project_id FROM invoices WHERE id=%s", (iid,))
    row = c.fetchone()
    if not row:
        conn.close(); raise HTTPException(404, "Invoice tidak ditemukan")
    pid = row['project_id']
    yr = payload.tahun if payload.tahun else date.today().year
    c.execute("""UPDATE invoices SET
        project_id=%s, invoice_no=%s, invoice_date=%s, period=%s,
        invoice_amount=%s, paid_amount=%s, paid_date=%s, tahun=%s
        WHERE id=%s""",
        (payload.project_id or None, payload.invoice_no,
         payload.invoice_date or None, payload.period,
         payload.invoice_amount, payload.paid_amount,
         payload.paid_date or None, yr, iid))
    conn.commit()
    for p in set(filter(None, [pid, payload.project_id])):
        sync_project_status(p)
    conn.close()
    return {"message": "Invoice berhasil diperbarui."}


@app.delete("/api/v1/revenue/invoices/{iid}", tags=["Revenue"])
def invoice_delete(iid: int, user: dict = Depends(require_menu("rev_invoice"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT project_id FROM invoices WHERE id=%s", (iid,))
    row = c.fetchone()
    if not row:
        conn.close(); raise HTTPException(404, "Invoice tidak ditemukan")
    pid = row['project_id']
    c.execute("DELETE FROM invoices WHERE id=%s", (iid,))
    conn.commit()
    if pid:
        sync_project_status(pid)
    conn.close()
    return {"message": "Invoice berhasil dihapus."}


@app.post("/api/v1/revenue/invoices/{iid}/pay", tags=["Revenue"])
def invoice_pay(iid: int, payload: InvoicePay, user: dict = Depends(require_menu("rev_invoice"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT project_id FROM invoices WHERE id=%s", (iid,))
    row = c.fetchone()
    pid = row['project_id'] if row else None
    c.execute("UPDATE invoices SET paid_amount=%s, paid_date=%s WHERE id=%s",
              (payload.paid_amount, payload.paid_date or None, iid))
    conn.commit()
    if pid:
        sync_project_status(pid)
    conn.close()
    return {"message": "Pembayaran berhasil diperbarui."}


# ═══════════════════════════════════════════════════════════════════════════════
# KPI PROSPECTING
# ═══════════════════════════════════════════════════════════════════════════════

class KpiUpdate(BaseModel):
    q1_actual: float = 0
    q2_actual: float = 0
    q3_actual: float = 0
    q4_actual: float = 0


@app.get("/api/v1/revenue/kpi", tags=["Revenue"])
def kpi_list(tahun: int = Query(0), user: dict = Depends(require_menu("rev_kpi"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM kpi_prospecting ORDER BY sort_order")
    kpis = [_norm(dict(r)) for r in c.fetchall()]
    from collections import defaultdict
    grouped = defaultdict(list)
    for k in kpis:
        grouped[k['kpi_category']].append(k)
    conn.close()
    return {"kpis": kpis, "grouped": dict(grouped)}


@app.put("/api/v1/revenue/kpi/{kid}", tags=["Revenue"])
def kpi_update(kid: int, payload: KpiUpdate, user: dict = Depends(require_menu("rev_kpi"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""UPDATE kpi_prospecting SET q1_actual=%s,q2_actual=%s,q3_actual=%s,q4_actual=%s WHERE id=%s""",
              (payload.q1_actual, payload.q2_actual, payload.q3_actual, payload.q4_actual, kid))
    conn.commit(); conn.close()
    return {"message": "KPI berhasil diperbarui."}


# ── KPI Prospecting CRUD lengkap (dipakai kpi.vue) ───────────────────────────

class KpiProspectingCreate(BaseModel):
    kpi_category: str
    kpi_name: str
    unit: str = "Count"
    is_auto: bool = False
    target_annual: float = 0
    q1_target: float = 0; q2_target: float = 0; q3_target: float = 0; q4_target: float = 0
    q1_actual: float = 0; q2_actual: float = 0; q3_actual: float = 0; q4_actual: float = 0
    tahun: int = 2026
    sort_order: int = 0


class KpiProspectingUpdate(BaseModel):
    kpi_category: str = None
    kpi_name: str = None
    unit: str = None
    is_auto: bool = None
    target_annual: float = None
    q1_target: float = None; q2_target: float = None; q3_target: float = None; q4_target: float = None
    q1_actual: float = None; q2_actual: float = None; q3_actual: float = None; q4_actual: float = None


@app.get("/api/v1/kpi/prospecting/years", tags=["KPI"])
def kpi_years(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT DISTINCT tahun FROM kpi_prospecting ORDER BY tahun DESC")
    years = [r["tahun"] for r in c.fetchall()]
    conn.close()
    if not years:
        years = [2026]
    return years


@app.get("/api/v1/kpi/prospecting", tags=["KPI"])
def kpi_prospecting_list(tahun: int = Query(0), user: dict = Depends(get_current_user)):
    from datetime import date as _date
    yr = tahun or _date.today().year
    cur_q = ((_date.today().month - 1) // 3) + 1
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM kpi_prospecting WHERE tahun=%s ORDER BY sort_order, id", (yr,))
    data = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"data": data, "cur_q": cur_q, "tahun": yr}


@app.post("/api/v1/kpi/prospecting", tags=["KPI"])
def kpi_prospecting_create(payload: KpiProspectingCreate, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO kpi_prospecting
          (kpi_category,kpi_name,unit,is_auto,target_annual,
           q1_target,q2_target,q3_target,q4_target,
           q1_actual,q2_actual,q3_actual,q4_actual,tahun,sort_order)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (payload.kpi_category, payload.kpi_name, payload.unit, payload.is_auto,
          payload.target_annual,
          payload.q1_target, payload.q2_target, payload.q3_target, payload.q4_target,
          payload.q1_actual, payload.q2_actual, payload.q3_actual, payload.q4_actual,
          payload.tahun, payload.sort_order))
    new_id = c.fetchone()["id"]
    conn.commit(); conn.close()
    return {"id": new_id, "message": "KPI berhasil ditambahkan."}


@app.put("/api/v1/kpi/prospecting/{kid}", tags=["KPI"])
def kpi_prospecting_update(kid: int, payload: KpiProspectingUpdate, user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    # Ambil data sekarang
    c.execute("SELECT * FROM kpi_prospecting WHERE id=%s", (kid,))
    existing = c.fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="KPI tidak ditemukan")
    # Non-admin hanya boleh update aktual, bukan target
    # (lock enforcement ada di frontend; backend hanya filter kolom)
    updates = {}
    for field in ["kpi_category","kpi_name","unit","is_auto","target_annual",
                  "q1_target","q2_target","q3_target","q4_target",
                  "q1_actual","q2_actual","q3_actual","q4_actual"]:
        val = getattr(payload, field)
        if val is not None:
            updates[field] = val
    if not updates:
        conn.close()
        return {"message": "Tidak ada perubahan."}
    set_clause = ", ".join(f"{k}=%s" for k in updates)
    c.execute(f"UPDATE kpi_prospecting SET {set_clause} WHERE id=%s",
              list(updates.values()) + [kid])
    conn.commit(); conn.close()
    return {"message": "KPI berhasil diperbarui."}


@app.delete("/api/v1/kpi/prospecting/{kid}", tags=["KPI"])
def kpi_prospecting_delete(kid: int, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM kpi_prospecting WHERE id=%s", (kid,))
    conn.commit(); conn.close()
    return {"message": "KPI berhasil dihapus."}


# ═══════════════════════════════════════════════════════════════════════════════
# ENTERTAINMENT CLAIMS
# ═══════════════════════════════════════════════════════════════════════════════

def _next_claim_no(c) -> str:
    from datetime import date as _date
    prefix = f"ENT-{_date.today().strftime('%Y%m')}-"
    c.execute("SELECT claim_no FROM entertainment_claims WHERE claim_no LIKE %s ORDER BY id DESC LIMIT 1", (prefix + '%',))
    row = c.fetchone()
    seq = int(row['claim_no'].split('-')[-1]) + 1 if row else 1
    return f"{prefix}{seq:03d}"


class ClaimCreate(BaseModel):
    lead_id:    Optional[str] = None
    tgl_klaim:  str
    nama_klien: str
    lokasi:     str = ""
    lat:        Optional[float] = None
    lng:        Optional[float] = None
    jumlah:     float = 0
    keterangan: str = ""


class ClaimApprove(BaseModel):
    action:  str   # "Approved" | "Rejected"
    catatan: str = ""


class ClaimSettingUpdate(BaseModel):
    limit_per_bulan: float = 0


@app.get("/api/v1/entertain/settings", tags=["Entertain"])
def entertain_settings_get(user: dict = Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM entertainment_settings ORDER BY id LIMIT 1")
    row = c.fetchone()
    conn.close()
    return _norm(dict(row)) if row else {"limit_per_bulan": 0}


@app.put("/api/v1/entertain/settings", tags=["Entertain"])
def entertain_settings_update(payload: ClaimSettingUpdate, user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    c.execute("UPDATE entertainment_settings SET limit_per_bulan=%s, updated_at=NOW()",
              (payload.limit_per_bulan,))
    conn.commit(); conn.close()
    return {"message": "Setting berhasil disimpan."}


@app.get("/api/v1/entertain/claims", tags=["Entertain"])
def entertain_claims(
    status: str = Query(""),
    user_id: int = Query(0),
    bulan: int = Query(0),
    tahun: int = Query(0),
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=500),
    user: dict = Depends(require_menu("entertain"))
):
    from datetime import date as _date
    import math
    conn = get_conn(); c = conn.cursor()
    yr = tahun or _date.today().year
    role_id = user.get("role_id", 3)

    wheres = ["1=1"]
    params: list = []

    # Sales hanya lihat milik sendiri
    if role_id == 3:
        wheres.append("ec.user_id = %s"); params.append(user["id"])
    elif user_id:
        wheres.append("ec.user_id = %s"); params.append(user_id)

    if status:
        wheres.append("ec.status = %s"); params.append(status)
    if bulan:
        wheres.append("EXTRACT(MONTH FROM ec.tgl_klaim) = %s"); params.append(bulan)
    wheres.append("EXTRACT(YEAR FROM ec.tgl_klaim) = %s"); params.append(yr)

    where_sql = " AND ".join(wheres)

    c.execute(f"""SELECT COUNT(*) cnt FROM entertainment_claims ec WHERE {where_sql}""", params)
    total = int(c.fetchone()["cnt"])

    offset = (page - 1) * per_page
    c.execute(f"""
        SELECT ec.*, u.nama AS sales_nama,
               l.nama_company AS lead_nama
        FROM entertainment_claims ec
        JOIN users u ON u.id = ec.user_id
        LEFT JOIN leads l ON l.lead_id = ec.lead_id
        WHERE {where_sql}
        ORDER BY ec.submitted_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    claims = [_norm(dict(r)) for r in c.fetchall()]

    # Summary cards (filtered by period)
    c.execute("""
        SELECT COALESCE(SUM(jumlah),0) AS total_bulan,
               COALESCE(SUM(jumlah),0) AS total_amount,
               COALESCE(SUM(jumlah) FILTER (WHERE status='Approved'),0) AS approved_amount,
               COUNT(*) FILTER (WHERE status='Pending')  AS pending,
               COUNT(*) FILTER (WHERE status='Approved') AS approved,
               COUNT(*) FILTER (WHERE status='Rejected') AS rejected
        FROM entertainment_claims
        WHERE EXTRACT(YEAR FROM tgl_klaim)=%s
          AND (%s = 0 OR EXTRACT(MONTH FROM tgl_klaim)=%s)
          AND (%s = 3 AND user_id=%s OR %s != 3)
    """, (yr, bulan, bulan, role_id, user["id"], role_id))
    summary = _norm(dict(c.fetchone()))

    # Pending total (year-wide, no bulan filter) — for badge on dashboard
    c.execute("""
        SELECT COUNT(*) AS pending_total
        FROM entertainment_claims
        WHERE status='Pending'
          AND EXTRACT(YEAR FROM tgl_klaim)=%s
          AND (%s = 3 AND user_id=%s OR %s != 3)
    """, (yr, role_id, user["id"], role_id))
    summary["pending_total"] = int(c.fetchone()["pending_total"])

    # Limit: dari entertain_limit user (jika sales), atau default settings
    if role_id == 3:
        c.execute("SELECT entertain_limit FROM users WHERE id=%s", (user["id"],))
        row = c.fetchone()
        limit = float(row['entertain_limit'] or 0) if row else 0
    else:
        c.execute("SELECT limit_per_bulan FROM entertainment_settings LIMIT 1")
        setting = c.fetchone()
        limit = float(setting['limit_per_bulan'] or 0) if setting else 0

    conn.close()
    return {
        "claims": claims, "summary": summary, "limit_per_bulan": limit,
        "total": total, "page": page, "per_page": per_page,
        "total_pages": math.ceil(total / per_page) if total else 1,
    }


@app.get("/api/v1/entertain/claims/{cid}", tags=["Entertain"])
def entertain_claim_detail(cid: int, user: dict = Depends(require_menu("entertain_claims"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        SELECT ec.*, u.nama AS sales_nama, l.nama_company AS lead_nama
        FROM entertainment_claims ec
        JOIN users u ON u.id = ec.user_id
        LEFT JOIN leads l ON l.lead_id = ec.lead_id
        WHERE ec.id = %s
    """, (cid,))
    claim = c.fetchone()
    if not claim:
        conn.close(); raise HTTPException(404, "Klaim tidak ditemukan.")
    # Cek akses: sales hanya lihat milik sendiri
    if user.get("role_id") == 3 and claim["user_id"] != user["id"]:
        conn.close(); raise HTTPException(403, "Akses ditolak.")

    c.execute("""
        SELECT ea.*, u.nama AS approver_nama
        FROM entertainment_approvals ea
        JOIN users u ON u.id = ea.approver_id
        WHERE ea.claim_id = %s ORDER BY ea.created_at
    """, (cid,))
    approvals = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"claim": _norm(dict(claim)), "approvals": approvals}


@app.post("/api/v1/entertain/claims", tags=["Entertain"])
def entertain_claim_create(payload: ClaimCreate, user: dict = Depends(require_menu("entertain"))):
    from datetime import date as _date
    conn = get_conn(); c = conn.cursor()

    # Cek limit bulan berjalan
    tgl = _date.fromisoformat(payload.tgl_klaim)
    c.execute("""
        SELECT COALESCE(SUM(jumlah),0) AS total
        FROM entertainment_claims
        WHERE user_id=%s AND status='Approved'
          AND EXTRACT(YEAR FROM tgl_klaim)=%s
          AND EXTRACT(MONTH FROM tgl_klaim)=%s
    """, (user["id"], tgl.year, tgl.month))
    total_approved = float(c.fetchone()['total'] or 0)
    c.execute("SELECT entertain_limit FROM users WHERE id=%s", (user["id"],))
    u_row = c.fetchone()
    limit = float(u_row['entertain_limit'] or 0) if u_row else 0
    if not limit:
        c.execute("SELECT limit_per_bulan FROM entertainment_settings LIMIT 1")
        s_row = c.fetchone()
        limit = float(s_row['limit_per_bulan'] or 0) if s_row else 0
    limit_warning = bool(limit and (total_approved + payload.jumlah) > limit)

    # Insert dulu tanpa claim_no, ambil ID, lalu update claim_no berdasarkan ID
    # untuk menghindari race condition pada sequence generation
    c.execute("""
        INSERT INTO entertainment_claims
          (claim_no, user_id, lead_id, tgl_klaim, nama_klien, lokasi,
           lat, lng, jumlah, keterangan, status, limit_warning)
        VALUES ('TEMP',%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pending',%s) RETURNING id
    """, (user["id"], payload.lead_id, payload.tgl_klaim,
          payload.nama_klien, payload.lokasi, payload.lat, payload.lng,
          payload.jumlah, payload.keterangan, limit_warning))
    new_id = c.fetchone()['id']
    claim_no = f"ENT-{_date.today().strftime('%Y%m')}-{new_id:04d}"
    c.execute("UPDATE entertainment_claims SET claim_no=%s WHERE id=%s", (claim_no, new_id))
    conn.commit(); conn.close()
    return {"id": new_id, "claim_no": claim_no,
            "limit_warning": limit_warning,
            "message": "Klaim berhasil disubmit."}


@app.put("/api/v1/entertain/claims/{cid}", tags=["Entertain"])
def entertain_claim_update(cid: int, payload: ClaimCreate, user: dict = Depends(require_menu("entertain"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM entertainment_claims WHERE id=%s", (cid,))
    claim = c.fetchone()
    if not claim:
        conn.close(); raise HTTPException(404, "Klaim tidak ditemukan.")
    if claim['user_id'] != user["id"]:
        conn.close(); raise HTTPException(403, "Akses ditolak.")
    if claim['status'] != 'Pending':
        conn.close(); raise HTTPException(400, "Hanya klaim Pending yang dapat diedit.")
    c.execute("""
        UPDATE entertainment_claims
        SET lead_id=%s, tgl_klaim=%s, nama_klien=%s, lokasi=%s,
            lat=%s, lng=%s, jumlah=%s, keterangan=%s, updated_at=NOW()
        WHERE id=%s
    """, (payload.lead_id, payload.tgl_klaim, payload.nama_klien,
          payload.lokasi, payload.lat, payload.lng,
          payload.jumlah, payload.keterangan, cid))
    conn.commit(); conn.close()
    return {"message": "Klaim berhasil diupdate."}


@app.patch("/api/v1/entertain/claims/{cid}/cancel", tags=["Entertain"])
def entertain_claim_cancel(cid: int, user: dict = Depends(require_menu("entertain"))):
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM entertainment_claims WHERE id=%s", (cid,))
    claim = c.fetchone()
    if not claim:
        conn.close(); raise HTTPException(404, "Klaim tidak ditemukan.")
    if claim['user_id'] != user["id"]:
        conn.close(); raise HTTPException(403, "Akses ditolak.")
    if claim['status'] != 'Pending':
        conn.close(); raise HTTPException(400, "Hanya klaim Pending yang dapat dibatalkan.")
    c.execute("UPDATE entertainment_claims SET status='Cancelled', updated_at=NOW() WHERE id=%s", (cid,))
    conn.commit(); conn.close()
    return {"message": "Klaim dibatalkan."}


@app.post("/api/v1/entertain/claims/{cid}/approve", tags=["Entertain"])
def entertain_claim_approve(cid: int, payload: ClaimApprove, user: dict = Depends(require_menu("entertain_approval"))):
    if payload.action not in ("Approved", "Rejected"):
        raise HTTPException(400, "Action harus 'Approved' atau 'Rejected'.")
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT * FROM entertainment_claims WHERE id=%s", (cid,))
    claim = c.fetchone()
    if not claim:
        conn.close(); raise HTTPException(404, "Klaim tidak ditemukan.")
    if claim['status'] != 'Pending':
        conn.close(); raise HTTPException(400, "Hanya klaim Pending yang dapat diproses.")
    c.execute("""
        INSERT INTO entertainment_approvals (claim_id, approver_id, action, catatan)
        VALUES (%s,%s,%s,%s)
    """, (cid, user["id"], payload.action, payload.catatan))
    c.execute("UPDATE entertainment_claims SET status=%s, updated_at=NOW() WHERE id=%s",
              (payload.action, cid))
    conn.commit(); conn.close()
    return {"message": f"Klaim {payload.action}."}


@app.post("/api/v1/entertain/claims/{cid}/photo", tags=["Entertain"])
async def entertain_upload_photo(
    cid: int,
    file: UploadFile = File(...),
    user: dict = Depends(require_menu("entertain"))
):
    import shutil, uuid
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT user_id, status FROM entertainment_claims WHERE id=%s", (cid,))
    claim = c.fetchone()
    if not claim:
        conn.close(); raise HTTPException(404, "Klaim tidak ditemukan.")
    if user.get("role_id") == 3 and claim['user_id'] != user["id"]:
        conn.close(); raise HTTPException(403, "Akses ditolak.")

    ext = os.path.splitext(file.filename or "")[1].lower() or ".jpg"
    fname = f"{uuid.uuid4().hex}{ext}"
    upload_dir = os.path.join(os.path.dirname(__file__), "static", "uploads", "entertain")
    os.makedirs(upload_dir, exist_ok=True)
    with open(os.path.join(upload_dir, fname), "wb") as f:
        shutil.copyfileobj(file.file, f)
    photo_url = f"uploads/entertain/{fname}"
    c.execute("UPDATE entertainment_claims SET foto_bukti=%s, updated_at=NOW() WHERE id=%s",
              (photo_url, cid))
    conn.commit(); conn.close()
    return {"foto_bukti": photo_url, "url": f"/storage/{photo_url}"}


@app.get("/api/v1/entertain/rekap", tags=["Entertain"])
def entertain_rekap(
    tahun: int = Query(0),
    bulan: int = Query(0),
    user: dict = Depends(require_menu("entertain_approval"))
):
    from datetime import date as _date
    yr = tahun or _date.today().year
    conn = get_conn(); c = conn.cursor()

    # Per sales per bulan (chart bulanan — selalu tahunan)
    c.execute("""
        SELECT u.nama AS sales_nama,
               EXTRACT(MONTH FROM ec.tgl_klaim)::int AS bulan,
               COUNT(*) AS jumlah_klaim,
               COALESCE(SUM(ec.jumlah),0) AS total
        FROM entertainment_claims ec
        JOIN users u ON u.id = ec.user_id
        WHERE ec.status = 'Approved'
          AND EXTRACT(YEAR FROM ec.tgl_klaim) = %s
        GROUP BY u.nama, bulan ORDER BY u.nama, bulan
    """, (yr,))
    per_sales = [_norm(dict(r)) for r in c.fetchall()]

    # Grand total per sales + limit individu, filter bulan jika ada
    c.execute("""
        SELECT u.id AS user_id, u.nama AS sales_nama,
               u.entertain_limit, u.join_date,
               COUNT(ec_all.id) AS jumlah_klaim,
               COALESCE(SUM(ec_all.jumlah),0) AS total,
               COALESCE(SUM(CASE WHEN ec_all.status='Approved' THEN ec_all.jumlah ELSE 0 END),0) AS total_approved
        FROM users u
        LEFT JOIN entertainment_claims ec_all
          ON ec_all.user_id = u.id
          AND EXTRACT(YEAR FROM ec_all.tgl_klaim) = %s
          AND (%s = 0 OR EXTRACT(MONTH FROM ec_all.tgl_klaim) = %s)
          AND ec_all.status NOT IN ('Cancelled')
        WHERE u.is_active = 1
        GROUP BY u.id, u.nama, u.entertain_limit, u.join_date ORDER BY u.nama
    """, (yr, bulan, bulan))
    per_sales_total = [_norm(dict(r)) for r in c.fetchall()]

    conn.close()
    return {"per_sales": per_sales, "per_sales_total": per_sales_total, "tahun": yr, "bulan": bulan}


@app.get("/api/v1/entertain/sales-limits", tags=["Entertain"])
def entertain_sales_limits(user: dict = Depends(require_menu("entertain_approval"))):
    """Daftar sales aktif beserta limit entertain masing-masing."""
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        SELECT id, nama, entertain_limit
        FROM users WHERE is_active=1 ORDER BY nama
    """)
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows


@app.put("/api/v1/entertain/sales-limits/{uid}", tags=["Entertain"])
def entertain_sales_limit_update(uid: int, payload: dict, user: dict = Depends(require_admin)):
    """Update limit entertain untuk 1 sales."""
    limit = float(payload.get("entertain_limit", 0))
    conn = get_conn(); c = conn.cursor()
    c.execute("UPDATE users SET entertain_limit=%s WHERE id=%s", (limit, uid))
    conn.commit(); conn.close()
    return {"message": "Limit berhasil disimpan."}


@app.get("/api/v1/entertain/roi", tags=["Entertain"])
def entertain_roi(
    tahun: int = Query(0),
    sales: str = Query(""),
    user: dict = Depends(require_menu("entertain"))
):
    """
    Perbandingan entertain spend vs pipeline won vs revenue per sales, per tahun.
    Untuk Admin/Manager: semua sales (atau filter 1 sales).
    Untuk Sales: hanya diri sendiri.
    """
    from datetime import date as _date
    yr       = tahun or _date.today().year
    is_sales = user.get("role_id") == 3
    conn = get_conn(); c = conn.cursor()

    # Filter nama sales
    nama_filter = user["nama"] if is_sales else (sales or None)

    sf_u = "AND u.nama = %s" if nama_filter else ""
    p_u  = [nama_filter] if nama_filter else []

    # ── Per-sales entertain usage (tahun penuh) ───────────────────────────
    c.execute(f"""
        SELECT
            u.id          AS user_id,
            u.nama        AS sales_nama,
            u.entertain_limit,
            COALESCE(SUM(CASE WHEN ec.status='Approved'
                              AND EXTRACT(YEAR FROM ec.tgl_klaim)=%s
                         THEN ec.jumlah ELSE 0 END), 0) AS entertain_approved,
            COALESCE(SUM(CASE WHEN ec.status NOT IN ('Cancelled')
                              AND EXTRACT(YEAR FROM ec.tgl_klaim)=%s
                         THEN ec.jumlah ELSE 0 END), 0) AS entertain_diajukan,
            COUNT(CASE WHEN ec.status NOT IN ('Cancelled')
                            AND EXTRACT(YEAR FROM ec.tgl_klaim)=%s
                       THEN 1 END) AS jumlah_klaim
        FROM users u
        LEFT JOIN entertainment_claims ec ON ec.user_id = u.id
        WHERE u.is_active = 1 {sf_u}
        GROUP BY u.id, u.nama, u.entertain_limit
        ORDER BY u.nama
    """, [yr, yr, yr] + p_u)
    base = {r["user_id"]: dict(r) for r in c.fetchall()}

    # ── Pipeline Won dalam tahun yang sama ───────────────────────────────
    sf_l = "AND l.sales_owner = %s" if nama_filter else ""
    p_l  = [nama_filter] if nama_filter else []
    c.execute(f"""
        SELECT
            u.id AS user_id,
            COUNT(*) AS won_count,
            COALESCE(SUM(l.deal_value), 0) AS won_value
        FROM leads l
        JOIN users u ON u.nama = l.sales_owner
        WHERE l.stage = 'Won'
          AND EXTRACT(YEAR FROM COALESCE(l.deal_date, l.updated_at)) = %s
          {sf_l}
        GROUP BY u.id
    """, [yr] + p_l)
    for r in c.fetchall():
        if r["user_id"] in base:
            base[r["user_id"]]["won_count"] = int(r["won_count"])
            base[r["user_id"]]["won_value"] = float(r["won_value"])

    # ── Revenue actual dari revenue_projects ─────────────────────────────
    sf_r = "AND rp.pic = %s" if nama_filter else ""
    p_r  = [nama_filter] if nama_filter else []
    c.execute(f"""
        SELECT
            u.id AS user_id,
            COALESCE(SUM(rp.actual_revenue), 0) AS revenue_actual,
            COALESCE(SUM(rp.revenue_target), 0) AS revenue_target
        FROM revenue_projects rp
        JOIN users u ON u.nama = rp.pic
        WHERE rp.is_active = 1 AND rp.tahun = %s {sf_r}
        GROUP BY u.id
    """, [yr] + p_r)
    for r in c.fetchall():
        if r["user_id"] in base:
            base[r["user_id"]]["revenue_actual"] = float(r["revenue_actual"])
            base[r["user_id"]]["revenue_target"] = float(r["revenue_target"])

    # ── Tren entertain per bulan (semua sales atau 1 sales) ──────────────
    sf_t = "AND u.nama = %s" if nama_filter else ""
    p_t  = [nama_filter] if nama_filter else []
    c.execute(f"""
        SELECT
            EXTRACT(MONTH FROM ec.tgl_klaim)::int AS bulan,
            COALESCE(SUM(CASE WHEN ec.status='Approved' THEN ec.jumlah ELSE 0 END), 0) AS approved
        FROM entertainment_claims ec
        JOIN users u ON u.id = ec.user_id
        WHERE EXTRACT(YEAR FROM ec.tgl_klaim) = %s
          AND ec.status NOT IN ('Cancelled')
          {sf_t}
        GROUP BY bulan ORDER BY bulan
    """, [yr] + p_t)
    tren = {r["bulan"]: float(r["approved"]) for r in c.fetchall()}
    tren_monthly = [tren.get(m, 0) for m in range(1, 13)]

    conn.close()

    rows = list(base.values())
    for r in rows:
        r.setdefault("won_count", 0)
        r.setdefault("won_value", 0.0)
        r.setdefault("revenue_actual", 0.0)
        r.setdefault("revenue_target", 0.0)
        budget_tahunan = float(r["entertain_limit"] or 0) * 12
        r["budget_tahunan"] = budget_tahunan
        r["sisa_budget"]    = max(budget_tahunan - float(r["entertain_approved"]), 0)
        r["usage_pct"]      = (float(r["entertain_approved"]) / budget_tahunan * 100) if budget_tahunan else 0
        # ROI: setiap 1 Rupiah entertain menghasilkan berapa rupiah deal won
        spend = float(r["entertain_approved"])
        r["roi_pipeline"] = round(float(r["won_value"]) / spend, 1) if spend else None
        r["roi_revenue"]  = round(float(r["revenue_actual"]) / spend, 1) if spend else None

    return {
        "tahun": yr,
        "rows": rows,
        "tren_monthly": tren_monthly,
        "bulan_labels": ["Jan","Feb","Mar","Apr","Mei","Jun",
                         "Jul","Ags","Sep","Okt","Nov","Des"],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# BUDGET MONITORING
# ═══════════════════════════════════════════════════════════════════════════════

class BudgetCreate(BaseModel):
    perspektif_bsc: str
    category: str
    sub_category: str = ""
    budget_amount: float = 0
    actual_amount: float = 0
    month_num: int = 0
    status: str = "Planning"
    notes: str = ""
    tahun: int = 2026


class BudgetUpdate(BaseModel):
    perspektif_bsc: str = None
    category: str = None
    sub_category: str = None
    budget_amount: float = None
    actual_amount: float = None
    month_num: int = None
    status: str = None
    notes: str = None


def _budget_ach(budget: float, actual: float) -> float:
    if not budget:
        return 0.0
    return round(actual / budget * 100, 1)


@app.get("/api/v1/budget", tags=["Budget"])
def budget_list(tahun: int = Query(0), user: dict = Depends(require_menu("rev_budget"))):
    from datetime import date as _date
    yr = tahun or _date.today().year
    conn = get_conn()
    c = conn.cursor()

    # raw items
    c.execute("""SELECT * FROM budget_items WHERE tahun=%s AND deleted_at IS NULL
                 ORDER BY perspektif_bsc, category, id""", (yr,))
    raw = c.fetchall()
    items = []
    for r in raw:
        d = _norm(dict(r))
        d['ach_pct'] = _budget_ach(float(d.get('budget_amount') or 0), float(d.get('actual_amount') or 0))
        items.append(d)

    # summary per perspektif_bsc
    c.execute("""SELECT perspektif_bsc,
                        COALESCE(SUM(budget_amount),0) AS budget,
                        COALESCE(SUM(actual_amount),0) AS actual,
                        COUNT(*) AS item_count
                 FROM budget_items WHERE tahun=%s AND deleted_at IS NULL
                 GROUP BY perspektif_bsc ORDER BY perspektif_bsc""", (yr,))
    summary = []
    for r in c.fetchall():
        d = _norm(dict(r))
        d['ach_pct'] = _budget_ach(float(d.get('budget') or 0), float(d.get('actual') or 0))
        summary.append(d)

    # available years
    c.execute("SELECT DISTINCT tahun FROM budget_items ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall()] or [yr]

    # distinct perspectives
    c.execute("SELECT DISTINCT perspektif_bsc FROM budget_items WHERE tahun=%s AND deleted_at IS NULL ORDER BY 1", (yr,))
    perspectives = [r['perspektif_bsc'] for r in c.fetchall()]

    conn.close()
    return {"data": items, "summary": summary, "perspectives": perspectives, "years": years}


@app.post("/api/v1/budget", tags=["Budget"])
def budget_create(payload: BudgetCreate, user: dict = Depends(require_menu("rev_budget"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""INSERT INTO budget_items
                   (perspektif_bsc, category, sub_category, budget_amount, actual_amount,
                    month_num, status, notes, tahun)
                 VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
              (payload.perspektif_bsc, payload.category, payload.sub_category,
               payload.budget_amount, payload.actual_amount,
               payload.month_num, payload.status, payload.notes, payload.tahun))
    new_id = c.fetchone()['id']
    conn.commit(); conn.close()
    return {"id": new_id, "message": "Budget item berhasil ditambahkan."}


@app.put("/api/v1/budget/{bid}", tags=["Budget"])
def budget_update(bid: int, payload: BudgetUpdate, user: dict = Depends(require_menu("rev_budget"))):
    conn = get_conn()
    c = conn.cursor()
    updates = {k: v for k, v in payload.dict().items() if v is not None}
    if not updates:
        conn.close()
        return {"message": "Tidak ada perubahan."}
    set_clause = ", ".join(f"{k}=%s" for k in updates)
    c.execute(f"UPDATE budget_items SET {set_clause}, updated_at=NOW() WHERE id=%s",
              list(updates.values()) + [bid])
    conn.commit(); conn.close()
    return {"message": "Budget item berhasil diperbarui."}


@app.delete("/api/v1/budget/{bid}", tags=["Budget"])
def budget_delete(bid: int, user: dict = Depends(require_menu("rev_budget"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE budget_items SET deleted_at=NOW() WHERE id=%s", (bid,))
    conn.commit(); conn.close()
    return {"message": "Budget item berhasil dihapus."}


# Keep old endpoint for backward compat
@app.get("/api/v1/revenue/budget", tags=["Revenue"])
def budget_monitoring_api(tahun: int = Query(0), user: dict = Depends(require_menu("rev_budget"))):
    return budget_list(tahun=tahun, user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER DATA
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/roles", tags=["Master"])
def roles_list(user: dict = Depends(require_menu("roles"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM roles ORDER BY id")
    roles = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT role_id, menu_key FROM role_menus")
    role_menus: dict = {}
    for r in c.fetchall():
        role_menus.setdefault(r['role_id'], []).append(r['menu_key'])
    conn.close()
    for role in roles:
        role['menus'] = role_menus.get(role['id'], [])
    return {"roles": roles, "all_menus": ALL_MENUS}


@app.get("/api/v1/users", tags=["Master"])
def users_list(user: dict = Depends(require_menu("users"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.id, u.nama, u.email, u.role_id, u.is_active, u.location_tracking_enabled, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.id""")
    users = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT * FROM roles ORDER BY id")
    roles = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"users": users, "roles": roles}


@app.get("/api/v1/sales", tags=["Master"])
def sales_list_api(user: dict = Depends(get_current_user)):
    """Daftar sales = users beserta role mereka."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.id, u.nama, u.email, u.is_active, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.nama""")
    sales = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"sales": sales}


@app.get("/api/v1/menus", tags=["Master"])
def menus_list(user: dict = Depends(get_current_user)):
    allowed = list(get_user_menus(user["role_id"]))
    nav = [m for m in ALL_MENUS if m["key"] in allowed]
    return {"all_menus": ALL_MENUS, "allowed_menus": allowed, "nav_menus": nav}


# ── /master/* alias routes ─────────────────────────────────────────────────────

@app.get("/api/v1/master/roles", tags=["Master"])
def master_roles(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM roles ORDER BY id")
    roles = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT role_id, menu_key FROM role_menus")
    role_menus_map: dict = {}
    for r in c.fetchall():
        role_menus_map.setdefault(r['role_id'], []).append(r['menu_key'])
    conn.close()
    for role in roles:
        role['menus'] = role_menus_map.get(role['id'], [])
    return roles  # array langsung

@app.get("/api/v1/master/roles/{role_id}/menus", tags=["Master"])
def master_role_menus_get(role_id: int, user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT menu_key FROM role_menus WHERE role_id=%s", (role_id,))
    keys = [r["menu_key"] for r in c.fetchall()]
    conn.close()
    return keys  # array string langsung

@app.put("/api/v1/master/roles/{role_id}/menus", tags=["Master"])
def master_role_menus_update(role_id: int, payload: dict, user: dict = Depends(require_admin)):
    menus = payload.get("menus", [])
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM role_menus WHERE role_id=%s", (role_id,))
    for key in menus:
        c.execute("INSERT INTO role_menus (role_id, menu_key) VALUES (%s,%s) ON CONFLICT DO NOTHING", (role_id, key))
    conn.commit(); conn.close()
    return {"message": "Menu role berhasil diperbarui."}

@app.get("/api/v1/master/users", tags=["Master"])
def master_users(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.id, u.nama, u.email, u.role_id, u.is_active, u.location_tracking_enabled, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.id""")
    users = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return users  # array langsung

@app.post("/api/v1/master/users", tags=["Master"])
def master_user_create(payload: dict, user: dict = Depends(require_admin)):
    import hashlib
    nama  = payload.get("nama", "").strip()
    email = payload.get("email", "").strip()
    role_id = int(payload.get("role_id", 3))
    raw_pw = payload.get("password", "password")
    pw_hash = hashlib.sha256(raw_pw.encode()).hexdigest()
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE email=%s", (email,))
    if c.fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Email sudah terdaftar.")
    c.execute("""INSERT INTO users (nama, email, password, role_id, is_active)
                 VALUES (%s,%s,%s,%s,1) RETURNING id""", (nama, email, pw_hash, role_id))
    new_id = c.fetchone()["id"]
    conn.commit(); conn.close()
    return {"id": new_id, "message": "User berhasil ditambahkan."}

@app.put("/api/v1/master/users/{uid}", tags=["Master"])
def master_user_update(uid: int, payload: dict, user: dict = Depends(require_admin)):
    import hashlib
    conn = get_conn()
    c = conn.cursor()
    updates = {}
    if "nama" in payload: updates["nama"] = payload["nama"]
    if "email" in payload: updates["email"] = payload["email"]
    if "role_id" in payload: updates["role_id"] = int(payload["role_id"])
    if "is_active" in payload: updates["is_active"] = int(payload["is_active"])
    if "location_tracking_enabled" in payload: updates["location_tracking_enabled"] = bool(payload["location_tracking_enabled"])
    if "entertain_limit" in payload: updates["entertain_limit"] = float(payload["entertain_limit"] or 0)
    if "join_date" in payload and payload["join_date"]: updates["join_date"] = payload["join_date"]
    if "password" in payload and payload["password"]:
        updates["password"] = hashlib.sha256(payload["password"].encode()).hexdigest()
    if not updates:
        conn.close()
        return {"message": "Tidak ada perubahan."}
    set_clause = ", ".join(f"{k}=%s" for k in updates)
    c.execute(f"UPDATE users SET {set_clause} WHERE id=%s", list(updates.values()) + [uid])
    conn.commit(); conn.close()
    return {"message": "User berhasil diperbarui."}

@app.delete("/api/v1/master/users/{uid}", tags=["Master"])
def master_user_delete(uid: int, user: dict = Depends(require_admin)):
    if uid == user["id"]:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun sendiri.")
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=%s", (uid,))
    conn.commit(); conn.close()
    return {"message": "User berhasil dihapus."}

@app.get("/api/v1/master/sales", tags=["Master"])
def master_sales(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.id, u.nama, u.email, u.is_active, u.role_id, u.entertain_limit, u.join_date, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.nama""")
    sales = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return sales  # array langsung

@app.get("/api/v1/master/menus", tags=["Master"])
def master_menus(user: dict = Depends(get_current_user)):
    return ALL_MENUS  # array langsung (semua menu definition)

@app.get("/api/v1/master/products", tags=["Master"])
def master_products_list(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM products ORDER BY nama")
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows  # array langsung

@app.get("/api/v1/master/products/dropdown", tags=["Master"])
def master_products_dropdown(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, kode, nama FROM products WHERE is_active=TRUE ORDER BY nama")
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows

@app.post("/api/v1/master/products", tags=["Master"])
def master_product_create(payload: dict, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""INSERT INTO products (kode, nama, kategori, deskripsi, is_active)
                 VALUES (%s,%s,%s,%s,%s) RETURNING id""",
              (payload.get("kode",""), payload.get("nama",""), payload.get("kategori",""),
               payload.get("deskripsi",""), bool(payload.get("is_active", True))))
    new_id = c.fetchone()["id"]
    conn.commit(); conn.close()
    return {"id": new_id, "message": "Produk berhasil ditambahkan."}

@app.get("/api/v1/master/products/export", tags=["Master"])
def master_product_export(user: dict = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT kode, nama, kategori, deskripsi, is_active, created_at FROM products ORDER BY kategori, kode")
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Produk"

    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="334155")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["Kode", "Nama Produk", "Kategori", "Deskripsi", "Status", "Tgl Dibuat"]
    col_widths = [14, 36, 26, 48, 12, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    kat_colors = {
        "Software Product":             "EFF6FF",
        "System Development":           "F5F3FF",
        "Managed Services":             "ECFDF5",
        "Professional Services":        "FFFBEB",
        "Infrastructure & Third-Party": "FEF2F2",
        "Membership":                   "F8FAFC",
    }

    for ri, row in enumerate(rows, 2):
        kode      = row["kode"];      nama      = row["nama"]
        kategori  = row["kategori"];  deskripsi = row["deskripsi"]
        is_active = row["is_active"]; created_at= row["created_at"]
        bg       = kat_colors.get(kategori or "", "FFFFFF")
        row_fill = PatternFill("solid", fgColor=bg)
        row_data = [
            kode, nama, kategori or "—", deskripsi or "—",
            "Aktif" if is_active else "Non-aktif",
            str(created_at)[:10] if created_at else "—",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill; cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 4))
            if ci == 5:
                cell.font = Font(color="166534" if is_active else "991B1B", bold=True)
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=master_produk.xlsx"},
    )

@app.put("/api/v1/master/products/{pid}", tags=["Master"])
def master_product_update(pid: int, payload: dict, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""UPDATE products SET kode=%s, nama=%s, kategori=%s, deskripsi=%s, is_active=%s
                 WHERE id=%s""",
              (payload.get("kode",""), payload.get("nama",""), payload.get("kategori",""),
               payload.get("deskripsi",""), bool(payload.get("is_active", True)), pid))
    conn.commit(); conn.close()
    return {"message": "Produk berhasil diperbarui."}

@app.delete("/api/v1/master/products/{pid}", tags=["Master"])
def master_product_delete(pid: int, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM products WHERE id=%s", (pid,))
    conn.commit(); conn.close()
    return {"message": "Produk berhasil dihapus."}


@app.get("/api/v1/master/organizations", tags=["Master"])
def master_org_list(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM organizations WHERE deleted_at IS NULL ORDER BY nama")
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows  # array langsung

@app.get("/api/v1/master/organizations/dropdown", tags=["Master"])
def master_org_dropdown(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, kode, nama FROM organizations WHERE is_active=1 AND deleted_at IS NULL ORDER BY nama")
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows

@app.get("/api/v1/master/organizations/lob", tags=["Master"])
def master_org_lob(user: dict = Depends(get_current_user)):
    """Organisasi tanpa parent — digunakan sebagai pilihan LOB."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT kode, nama FROM organizations WHERE parent_id IS NULL AND is_active=1 AND deleted_at IS NULL ORDER BY nama")
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows

@app.post("/api/v1/master/organizations", tags=["Master"])
def master_org_create(payload: dict, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""INSERT INTO organizations (kode, nama, parent_id, head, is_active)
                 VALUES (%s,%s,%s,%s,1) RETURNING id""",
              (payload.get("kode",""), payload.get("nama",""),
               payload.get("parent_id") or None, payload.get("head","")))
    new_id = c.fetchone()["id"]
    conn.commit(); conn.close()
    return {"id": new_id, "message": "Organisasi berhasil ditambahkan."}

@app.put("/api/v1/master/organizations/{oid}", tags=["Master"])
def master_org_update(oid: int, payload: dict, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""UPDATE organizations SET kode=%s, nama=%s, parent_id=%s, head=%s, is_active=%s
                 WHERE id=%s""",
              (payload.get("kode",""), payload.get("nama",""),
               payload.get("parent_id") or None, payload.get("head",""),
               int(payload.get("is_active", 1)), oid))
    conn.commit(); conn.close()
    return {"message": "Organisasi berhasil diperbarui."}

@app.delete("/api/v1/master/organizations/{oid}", tags=["Master"])
def master_org_delete(oid: int, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE organizations SET deleted_at=NOW() WHERE id=%s", (oid,))
    conn.commit(); conn.close()
    return {"message": "Organisasi berhasil dihapus."}


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT CSV
# ═══════════════════════════════════════════════════════════════════════════════

PIPELINE_TEMPLATE_HEADERS = [
    "lead_id","nama_company","organisasi","product","contact_person","phone","email",
    "segmen","sub_segmen","source","stage","prioritas","tgl_masuk","propose_value",
    "deal_value","probability","exp_close_date","sales_owner","next_fu_date",
    "last_fu_date","last_fu_notes","fu_count","days_in_stage","remarks",
]

REVENUE_TEMPLATE_HEADERS = [
    "project_id","lob","pic","product","client","organisasi","kategori","type",
    "tahun","target_month","revenue_target","actual_revenue","target_invoice_date",
    "invoice_date","payment_date","notes",
]

INVOICE_TEMPLATE_HEADERS = [
    "project_id","lob","product","client","organisasi","invoice_no","invoice_date",
    "period","tahun","invoice_amount","paid_amount","paid_date","notes",
]

from fastapi.responses import StreamingResponse
import csv, io


def _csv_template(headers: list[str], rows: list[list] = None) -> StreamingResponse:
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(headers)
    if rows:
        for r in rows:
            w.writerow(r)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template.csv"},
    )


def _clean(v, default=None):
    if v is None: return default
    s = str(v).strip()
    return None if s in ('', 'nan', 'NaT', 'None') else s

def _numval(v, default=0.0):
    try:
        f = float(str(v).replace(',', ''))
        return 0.0 if f != f else f
    except: return default

def _datestr(v):
    if not v: return None
    s = str(v).strip()
    if s in ('', 'nan', 'NaT', 'None'): return None
    try:
        import pandas as pd
        ts = pd.Timestamp(s)
        return ts.strftime('%Y-%m-%d') if not pd.isna(ts) else None
    except: return s[:10] if len(s) >= 10 else None

def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
    text = content.decode('utf-8-sig').replace('\r\n', '\n').replace('\r', '\n')
    lines = [l for l in text.split('\n') if l.strip() and not l.strip().startswith('#')]
    if not lines:
        return [], []
    first = lines[0]
    sep = ';' if ';' in first and ',' not in first else ','
    reader = csv.DictReader(lines, delimiter=sep)
    rows = list(reader)
    return list(reader.fieldnames or []), rows


@app.get("/api/v1/import/template/{type}", tags=["Import"])
def import_template(type: str, tahun: int = Query(0), user=Depends(get_current_user)):
    if type == "pipeline":
        sample = [["LD-001","PT. Contoh","PKP IT Konsultan","SW-001","Budi","08123456789",
                   "budi@contoh.com","Swasta","BUMN","Referral","New","Warm","2026-01-15",
                   "50000000","","","2026-06-30","Sales A","","","","0","0",""]]
        return _csv_template(PIPELINE_TEMPLATE_HEADERS, sample)
    elif type == "revenue":
        sample = [["1","DCSS","Sales A","SW-001","PT. Client","PT. Client","Managed Services",
                   "Recurring","2026","2026-01-01","100000000","0","2026-01-10",
                   "","","Catatan opsional"]]
        return _csv_template(REVENUE_TEMPLATE_HEADERS, sample)
    elif type == "invoice":
        cur_year = tahun if tahun else date.today().year
        conn = get_conn(); c = conn.cursor()
        c.execute("""
            SELECT project_id, lob, product, client, organisasi, tahun
            FROM revenue_projects
            WHERE is_active=1 AND tahun=%s
            ORDER BY project_id
        """, (cur_year,))
        rows = c.fetchall()
        conn.close()
        if rows:
            sample = [
                [r["project_id"], r["lob"], r["product"], r["client"], r["organisasi"],
                 "", "", "", r["tahun"], "", "", "", ""]
                for r in rows
            ]
        else:
            sample = [["REV-0001","DCSS","SW-001","PT. Client","PT. Client","INV/2026/001",
                       "2026-01-15","January 2026","2026","50000000","50000000","2026-01-20",""]]
        return _csv_template(INVOICE_TEMPLATE_HEADERS, sample)
    raise HTTPException(404, "Template tidak ditemukan")


@app.post("/api/v1/import/pipeline", tags=["Import"])
async def import_pipeline(
    file: UploadFile = File(...),
    clear_first: str = "0",
    user=Depends(get_current_user),
):
    content = await file.read()
    _, rows = _parse_csv(content)

    conn = get_conn()
    c    = conn.cursor()
    imported = updated = skipped = 0
    errors: list[str] = []

    try:
        if clear_first == "1":
            c.execute("DELETE FROM followups")
            c.execute("DELETE FROM leads")

        STAGE_PROB = {
            "new": 10, "in progress": 25, "demo scheduled": 40,
            "proposal sent": 60, "negotiation": 80, "won": 100,
            "on hold": 20, "lost": 0,
        }

        for i, row in enumerate(rows, 1):
            lid  = _clean(row.get("lead_id"))
            nama = _clean(row.get("nama_company"))
            if not lid or not nama:
                skipped += 1
                continue

            prob_raw = _clean(row.get("probability"))
            stage    = _clean(row.get("stage"), "New")
            prob     = float(prob_raw) if prob_raw else float(STAGE_PROB.get((stage or "").lower(), 10))
            pv       = _numval(row.get("propose_value"))
            dv       = _numval(row.get("deal_value"))
            wv       = dv * (prob / 100) if dv else pv * (prob / 100)

            c.execute("SELECT lead_id FROM leads WHERE lead_id=%s", (lid,))
            exists = c.fetchone()

            c.execute("""
                INSERT INTO leads
                  (lead_id,nama_company,organisasi,product,contact_person,phone,email,
                   segmen,sub_segmen,source,stage,prioritas,tgl_masuk,propose_value,
                   deal_value,probability,exp_close_date,weighted_value,sales_owner,
                   next_fu_date,last_fu_date,last_fu_notes,fu_count,days_in_stage,remarks)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (lead_id) DO UPDATE SET
                  nama_company=EXCLUDED.nama_company, organisasi=EXCLUDED.organisasi,
                  product=EXCLUDED.product, contact_person=EXCLUDED.contact_person,
                  phone=EXCLUDED.phone, email=EXCLUDED.email,
                  segmen=EXCLUDED.segmen, sub_segmen=EXCLUDED.sub_segmen,
                  source=EXCLUDED.source, stage=EXCLUDED.stage, prioritas=EXCLUDED.prioritas,
                  tgl_masuk=EXCLUDED.tgl_masuk, propose_value=EXCLUDED.propose_value,
                  deal_value=EXCLUDED.deal_value, probability=EXCLUDED.probability,
                  exp_close_date=EXCLUDED.exp_close_date, weighted_value=EXCLUDED.weighted_value,
                  sales_owner=EXCLUDED.sales_owner, next_fu_date=EXCLUDED.next_fu_date,
                  last_fu_date=EXCLUDED.last_fu_date, last_fu_notes=EXCLUDED.last_fu_notes,
                  fu_count=EXCLUDED.fu_count, days_in_stage=EXCLUDED.days_in_stage,
                  remarks=EXCLUDED.remarks
            """, (
                lid, nama,
                _clean(row.get("organisasi")),
                _clean(row.get("product")), _clean(row.get("contact_person")),
                _clean(row.get("phone")), _clean(row.get("email")),
                _clean(row.get("segmen")), _clean(row.get("sub_segmen")),
                _clean(row.get("source")), stage,
                _clean(row.get("prioritas"), "Warm"),
                _datestr(row.get("tgl_masuk")), pv, dv, prob,
                _datestr(row.get("exp_close_date")),
                wv, _clean(row.get("sales_owner")),
                _datestr(row.get("next_fu_date")), _datestr(row.get("last_fu_date")),
                _clean(row.get("last_fu_notes")),
                int(_numval(row.get("fu_count", 0))),
                int(_numval(row.get("days_in_stage", 0))),
                _clean(row.get("remarks")),
            ))
            if exists: updated += 1
            else: imported += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        errors.append(str(e))
    finally:
        conn.close()

    return {"imported": imported, "updated": updated, "skipped": skipped, "errors": errors}


@app.post("/api/v1/import/revenue", tags=["Import"])
async def import_revenue(
    file: UploadFile = File(...),
    clear_first: str = "0",
    user=Depends(get_current_user),
):
    from main import auto_status_risk
    content = await file.read()
    _, rows = _parse_csv(content)

    conn = get_conn()
    c    = conn.cursor()
    imported = updated = skipped = 0
    errors: list[str] = []

    try:
        if clear_first == "1":
            c.execute("DELETE FROM revenue_monthly")
            c.execute("DELETE FROM revenue_projects")

        for row in rows:
            pid_raw = _clean(row.get("project_id"))
            if not pid_raw: skipped += 1; continue
            if str(pid_raw).upper().startswith("REV-"):
                project_id = pid_raw.upper()
            else:
                try: project_id = f"REV-{int(float(pid_raw)):04d}"
                except: skipped += 1; continue

            rev_target = _numval(row.get("revenue_target"))
            rev_actual = 0  # actual_revenue dihitung dari invoice, bukan dari file Excel
            tahun_raw  = _clean(row.get("tahun"))
            tahun      = int(float(tahun_raw)) if tahun_raw else date.today().year
            type_raw   = _clean(row.get("type"))
            tgt_month  = _datestr(row.get("target_month"))
            status, risk = auto_status_risk(rev_target, rev_actual, type_raw, tahun, tgt_month)

            c.execute("SELECT project_id FROM revenue_projects WHERE project_id=%s", (project_id,))
            exists = c.fetchone()

            c.execute("""
                INSERT INTO revenue_projects
                  (project_id,lob,pic,product,client,organisasi,kategori,type,tahun,
                   target_month,revenue_target,actual_revenue,status,target_invoice_date,
                   invoice_date,payment_date,notes,risk_level)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (project_id) DO UPDATE SET
                  lob=EXCLUDED.lob, pic=EXCLUDED.pic, product=EXCLUDED.product,
                  client=EXCLUDED.client, organisasi=EXCLUDED.organisasi,
                  kategori=EXCLUDED.kategori, type=EXCLUDED.type,
                  tahun=EXCLUDED.tahun, target_month=EXCLUDED.target_month,
                  revenue_target=EXCLUDED.revenue_target,
                  status=EXCLUDED.status, target_invoice_date=EXCLUDED.target_invoice_date,
                  invoice_date=EXCLUDED.invoice_date, payment_date=EXCLUDED.payment_date,
                  notes=EXCLUDED.notes, risk_level=EXCLUDED.risk_level
            """, (
                project_id,
                _clean(row.get("lob"), "DCSS"), _clean(row.get("pic") or row.get("owner")),
                _clean(row.get("product")), _clean(row.get("client")),
                _clean(row.get("organisasi")),
                _clean(row.get("kategori")), type_raw, tahun, tgt_month,
                rev_target, rev_actual, status,
                _datestr(row.get("target_invoice_date")),
                _datestr(row.get("invoice_date")), _datestr(row.get("payment_date")),
                _clean(row.get("notes")), risk,
            ))
            if exists: updated += 1
            else: imported += 1

            # ── Generate revenue_monthly sesuai TYPE ──────────────────────────
            import datetime as _dt
            type_lower = (type_raw or '').lower()
            tid = _datestr(row.get("target_invoice_date"))
            try:
                start_month = _dt.date.fromisoformat(str(tid)[:10]).month if tid else None
            except Exception:
                start_month = None

            # Hapus semua row lama sebelum regenerate — mencegah sisa bulan lama
            # ketika target_invoice_date berubah (ON CONFLICT tidak menghapus row lama)
            c.execute("DELETE FROM revenue_monthly WHERE project_id=%s", (project_id,))

            if type_lower == 'bulanan' and start_month:
                num_months = 12 - start_month + 1
                per_month  = round(rev_target / num_months, 2) if num_months > 0 else 0
                for m in range(start_month, 13):
                    c.execute("""
                        INSERT INTO revenue_monthly (project_id, month_num, month_name, target, actual, termin_no)
                        VALUES (%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (project_id, month_num) DO UPDATE
                          SET target=EXCLUDED.target, termin_no=EXCLUDED.termin_no
                    """, (project_id, m, MONTHS_EN[m-1], per_month, 0, m - start_month + 1))
            elif type_lower in ('termin', 'tahunan') and start_month:
                c.execute("""
                    INSERT INTO revenue_monthly (project_id, month_num, month_name, target, actual, termin_no)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (project_id, month_num) DO UPDATE
                      SET target=EXCLUDED.target, termin_no=EXCLUDED.termin_no
                """, (project_id, start_month, MONTHS_EN[start_month-1], rev_target, 0, 1))
            # One Time → tidak insert revenue_monthly

        conn.commit()
    except Exception as e:
        conn.rollback()
        errors.append(str(e))
    finally:
        conn.close()

    return {"imported": imported, "updated": updated, "skipped": skipped, "errors": errors}


@app.post("/api/v1/import/invoice", tags=["Import"])
async def import_invoice(
    file: UploadFile = File(...),
    clear_first: str = "0",
    user=Depends(get_current_user),
):
    content = await file.read()
    _, rows = _parse_csv(content)

    conn = get_conn()
    c    = conn.cursor()
    imported = updated = skipped = 0
    synced_months = synced_projects = 0
    errors: list[str] = []
    affected_projects: set[str] = set()

    try:
        if clear_first == "1":
            c.execute("DELETE FROM invoices")

        for row in rows:
            pid_raw = _clean(row.get("project_id"))
            inv_no  = _clean(row.get("invoice_no"))
            if not pid_raw or not inv_no: skipped += 1; continue
            # Terima format "REV-0001" langsung, atau angka saja
            if str(pid_raw).upper().startswith("REV-"):
                project_id = pid_raw.upper()
            else:
                try: project_id = f"REV-{int(float(pid_raw)):04d}"
                except: skipped += 1; continue

            c.execute("SELECT id FROM revenue_projects WHERE project_id=%s", (project_id,))
            if not c.fetchone(): errors.append(f"Project {project_id} tidak ditemukan (baris invoice_no={inv_no})"); skipped += 1; continue

            c.execute("SELECT id FROM invoices WHERE invoice_no=%s AND project_id=%s", (inv_no, project_id))
            exists = c.fetchone()

            inv_amount  = _numval(row.get("invoice_amount"))
            paid_amount = _numval(row.get("paid_amount"))
            paid_date   = _datestr(row.get("paid_date"))

            status_raw = _clean(row.get("status"))
            if status_raw:
                inv_status = status_raw
            elif paid_amount >= inv_amount and inv_amount > 0:
                inv_status = "Lunas"
            elif paid_amount > 0:
                inv_status = "Partial"
            else:
                inv_status = "Unpaid"

            if exists:
                c.execute("""
                    UPDATE invoices SET
                      invoice_date=%s, period=%s, invoice_amount=%s,
                      paid_amount=%s, paid_date=%s, status=%s, notes=%s
                    WHERE invoice_no=%s AND project_id=%s
                """, (
                    _datestr(row.get("invoice_date")), _clean(row.get("period")),
                    inv_amount, paid_amount, paid_date, inv_status,
                    _clean(row.get("notes")), inv_no, project_id,
                ))
                updated += 1
            else:
                tahun_inv_raw = _clean(row.get("tahun"))
                tahun_inv = int(float(tahun_inv_raw)) if tahun_inv_raw else date.today().year
                c.execute("""
                    INSERT INTO invoices
                      (project_id,lob,product,client,organisasi,invoice_no,invoice_date,
                       period,tahun,invoice_amount,paid_amount,paid_date,notes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    project_id,
                    _clean(row.get("lob"), "DCSS"),
                    _clean(row.get("product")), _clean(row.get("client")),
                    _clean(row.get("organisasi")),
                    inv_no, _datestr(row.get("invoice_date")),
                    _clean(row.get("period")), tahun_inv,
                    inv_amount, paid_amount, paid_date, _clean(row.get("notes")),
                ))
                imported += 1
            affected_projects.add(project_id)

        # Recalculate actual revenue per project dari invoices
        # Gunakan invoice_date sebagai acuan bulan (bukan paid_date)
        for pid in affected_projects:
            c.execute("""
                SELECT EXTRACT(MONTH FROM invoice_date) AS m, SUM(paid_amount) AS total
                FROM invoices
                WHERE project_id=%s AND invoice_date IS NOT NULL AND paid_amount > 0
                GROUP BY EXTRACT(MONTH FROM invoice_date)
            """, (pid,))
            monthly = {int(r['m']): float(r['total']) for r in c.fetchall()}

            for month_num, total in monthly.items():
                mn = MONTHS_EN[month_num - 1]
                c.execute("""
                    UPDATE revenue_monthly SET actual=%s
                    WHERE project_id=%s AND month_num=%s
                """, (total, pid, month_num))
                if c.rowcount == 0:
                    c.execute("""
                        INSERT INTO revenue_monthly (project_id,month_num,month_name,target,actual,status)
                        VALUES (%s,%s,%s,0,%s,'Achieve')
                    """, (pid, month_num, mn, total))
                synced_months += 1

            c.execute("SELECT SUM(actual) FROM revenue_monthly WHERE project_id=%s", (pid,))
            total_actual = float(c.fetchone()['sum'] or 0)
            c.execute("""
                UPDATE revenue_projects SET actual_revenue=%s WHERE project_id=%s
            """, (total_actual, pid))
            synced_projects += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        errors.append(str(e))
    finally:
        conn.close()

    return {
        "imported": imported, "updated": updated, "skipped": skipped,
        "errors": errors, "synced_months": synced_months, "synced_projects": synced_projects,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ACTIVITY HEATMAP
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/activity-heatmap", tags=["Analytics"])
def activity_heatmap(
    tahun: int = 0,
    sales: str = "",
    user: dict = Depends(get_current_user)
):
    """
    Heatmap aktivitas per jam (0-23) dan per hari minggu (0=Sen..6=Min).
    Menggabungkan: follow_up_log, visit_logs, daily_reports.
    """
    from datetime import date as _date
    cur_year  = tahun or _date.today().year
    is_sales  = user.get("role_id") == 3
    conn = get_conn(); c = conn.cursor()

    # Resolusi filter sales
    filter_nama = user["nama"] if is_sales else (sales or None)

    # ── Follow-up log ──────────────────────────────────────────
    # tgl_fu is DATE only; filter via sales_owner (text column)
    fu_owner = ""
    fu_params: list = [cur_year]
    if filter_nama:
        fu_owner = "AND fu.sales_owner = %s"
        fu_params.append(filter_nama)

    # tgl_fu is DATE only — no reliable hour; use DOW from tgl_fu, skip hour dimension
    c.execute(f"""
        SELECT
            EXTRACT(ISODOW FROM tgl_fu)::int - 1 AS dow,
            COUNT(*) AS cnt
        FROM follow_up_log fu
        WHERE EXTRACT(YEAR FROM tgl_fu) = %s {fu_owner}
        GROUP BY 1
    """, fu_params)
    fu_by_dow_rows = c.fetchall()

    # ── Visit logs (check-in) ──────────────────────────────────
    v_owner = ""
    v_params: list = [cur_year]
    if filter_nama:
        v_owner  = "AND u.nama = %s"
        v_params.append(filter_nama)

    c.execute(f"""
        SELECT
            EXTRACT(ISODOW FROM v.checked_in_at)::int - 1 AS dow,
            EXTRACT(HOUR  FROM v.checked_in_at)::int      AS hour,
            COUNT(*) AS cnt
        FROM visit_logs v
        JOIN users u ON u.id = v.user_id
        WHERE EXTRACT(YEAR FROM v.checked_in_at) = %s {v_owner}
        GROUP BY 1, 2
    """, v_params)
    visit_rows = c.fetchall()

    # ── Daily reports (sent_at timestamp) ─────────────────────
    d_owner = ""
    d_params: list = [cur_year]
    if filter_nama:
        d_owner  = "AND u.nama = %s"
        d_params.append(filter_nama)

    c.execute(f"""
        SELECT
            EXTRACT(ISODOW FROM dr.sent_at)::int - 1 AS dow,
            EXTRACT(HOUR  FROM dr.sent_at)::int      AS hour,
            COUNT(*) AS cnt
        FROM daily_reports dr
        JOIN users u ON u.id = dr.user_id
        WHERE dr.sent_at IS NOT NULL
          AND EXTRACT(YEAR FROM dr.sent_at) = %s {d_owner}
        GROUP BY 1, 2
    """, d_params)
    dr_rows = c.fetchall()

    conn.close()

    # Grid 7×24 hanya dari sumber yang punya timestamp valid (visit_logs + daily_reports)
    grid: dict[tuple, int] = {}
    for row in list(visit_rows) + list(dr_rows):
        dow  = int(row["dow"] or 0)
        hour = int(row["hour"] or 0)
        grid[(dow, hour)] = grid.get((dow, hour), 0) + int(row["cnt"])

    # Follow-up hanya masuk ke by_dow (tidak punya jam)
    fu_dow_extra: dict[int, int] = {}
    for row in fu_by_dow_rows:
        d = int(row["dow"] or 0)
        fu_dow_extra[d] = fu_dow_extra.get(d, 0) + int(row["cnt"])

    # Flatten ke list + hitung max
    cells = [{"dow": d, "hour": h, "count": v} for (d, h), v in grid.items()]
    max_count = max((c["count"] for c in cells), default=1)

    # Agregasi per jam (hanya visit+report) dan per hari (visit+report+FU)
    by_hour = [sum(grid.get((d, h), 0) for d in range(7)) for h in range(24)]
    by_dow  = [sum(grid.get((d, h), 0) for h in range(24)) + fu_dow_extra.get(d, 0) for d in range(7)]

    return {
        "tahun": cur_year,
        "cells": cells,
        "max_count": max_count,
        "by_hour": by_hour,
        "by_dow":  by_dow,
        "dow_labels":  ["Sen","Sel","Rab","Kam","Jum","Sab","Min"],
        "hour_labels": [f"{h:02d}:00" for h in range(24)],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FIELD ACTIVITY
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/field-activity/stats", tags=["Field Activity"])
def field_activity_stats(user=Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    today = date.today().isoformat()
    c.execute("""
        SELECT
          COUNT(*) FILTER (WHERE DATE(checked_in_at) = %s) as today_visits,
          COUNT(*) FILTER (WHERE checked_out_at IS NULL AND DATE(checked_in_at) = %s) as active_now,
          COUNT(DISTINCT user_id) FILTER (WHERE DATE(checked_in_at) = %s) as active_sales,
          ROUND(AVG(duration_minutes) FILTER (WHERE duration_minutes IS NOT NULL AND DATE(checked_in_at) = %s)) as avg_duration,
          COUNT(*) FILTER (WHERE DATE(checked_in_at) >= (CURRENT_DATE - INTERVAL '6 days')) as week_count
        FROM visit_logs
    """, (today, today, today, today))
    row = dict(c.fetchone())
    conn.close()
    today_visits = int(row.get("today_visits") or 0)
    active_now   = int(row.get("active_now") or 0)
    active_sales = int(row.get("active_sales") or 0)
    return {
        "today_visits":  today_visits,
        "active_now":    active_now,
        "active_sales":  active_sales,
        "avg_duration":  int(row.get("avg_duration") or 0),
        "week_count":    int(row.get("week_count") or 0),
        # alias for frontend
        "today_count":   today_visits,
        "active_count":  active_sales,
    }


@app.get("/api/v1/field-activity/map", tags=["Field Activity"])
def field_activity_map(user=Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    today = date.today().isoformat()
    c.execute("""
        SELECT v.id, v.user_id, u.nama as sales_nama, v.lead_id,
               v.latitude, v.longitude, v.address, v.type,
               v.checked_in_at, v.checked_out_at, v.notes
        FROM visit_logs v
        LEFT JOIN users u ON v.user_id = u.id
        WHERE DATE(v.checked_in_at) = %s
          AND v.latitude IS NOT NULL AND v.longitude IS NOT NULL
        ORDER BY v.checked_in_at DESC
    """, (today,))
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows


@app.get("/api/v1/field-activity", tags=["Field Activity"])
def field_activity_list(
    date: str = None,
    user_id: int = None,
    type: str = None,
    page: int = 1,
    per_page: int = 20,
    current_user=Depends(get_current_user),
):
    conn = get_conn(); c = conn.cursor()
    is_sales = current_user.get("role_id") == 3
    where, params = [], []
    if date:
        where.append("DATE(v.checked_in_at) = %s"); params.append(date)
    if is_sales:
        where.append("v.user_id = %s"); params.append(current_user["id"])
    elif user_id:
        where.append("v.user_id = %s"); params.append(user_id)
    if type:
        where.append("v.type = %s"); params.append(type)

    where_str = ("WHERE " + " AND ".join(where)) if where else ""

    c.execute(f"SELECT COUNT(*) as cnt FROM visit_logs v {where_str}", params)
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page

    c.execute(f"""
        SELECT v.*, u.nama as sales_nama, l.nama_company
        FROM visit_logs v
        LEFT JOIN users u ON v.user_id = u.id
        LEFT JOIN leads l ON v.lead_id = l.lead_id
        {where_str}
        ORDER BY v.checked_in_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    import math as _math
    def _hav(la1, lo1, la2, lo2):
        R = 6371
        dlat = _math.radians(float(la2)-float(la1)); dlon = _math.radians(float(lo2)-float(lo1))
        a = _math.sin(dlat/2)**2 + _math.cos(_math.radians(float(la1)))*_math.cos(_math.radians(float(la2)))*_math.sin(dlon/2)**2
        return round(R*2*_math.asin(_math.sqrt(a)), 3)
    for r in rows:
        la1,lo1,la2,lo2 = r.get("latitude"),r.get("longitude"),r.get("checkout_latitude"),r.get("checkout_longitude")
        r["distance_km"] = _hav(la1,lo1,la2,lo2) if all(v is not None for v in [la1,lo1,la2,lo2]) else None
    return {"data": rows, "total": total, "pages": -(-total // per_page), "page": page}


class CheckinRequest(BaseModel):
    user_id: int
    lead_id: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    address: Optional[str] = None
    accuracy_m: Optional[int] = None
    notes: Optional[str] = None
    photo_base64: Optional[str] = None
    checked_in_at: Optional[str] = None


@app.post("/api/v1/field-activity/checkin", status_code=201, tags=["Field Activity"])
def field_activity_checkin(req: CheckinRequest, user=Depends(get_current_user)):
    import base64, uuid, os
    conn = get_conn(); c = conn.cursor()

    photo_url = None
    if req.photo_base64:
        try:
            header, data = req.photo_base64.split(",", 1)
            ext = "jpg"
            img_bytes = base64.b64decode(data)
            fname = f"checkin_{uuid.uuid4().hex}.{ext}"
            upload_dir = os.path.join(os.path.dirname(__file__), "static", "uploads", "checkin")
            os.makedirs(upload_dir, exist_ok=True)
            with open(os.path.join(upload_dir, fname), "wb") as f:
                f.write(img_bytes)
            photo_url = f"uploads/checkin/{fname}"
        except Exception:
            pass

    checked_in_at = req.checked_in_at or datetime.now().isoformat()

    actual_user_id = user["id"] if user.get("role_id") == 3 else (req.user_id or user["id"])
    c.execute("""
        INSERT INTO visit_logs
          (user_id, lead_id, type, latitude, longitude, address,
           accuracy_m, photo_url, notes, checked_in_at, created_at, updated_at)
        VALUES (%s,%s,'check_in',%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
        RETURNING id
    """, (
        actual_user_id, req.lead_id, req.latitude, req.longitude,
        req.address, req.accuracy_m, photo_url, req.notes, checked_in_at,
    ))
    new_id = c.fetchone()["id"]
    conn.commit(); conn.close()
    return {"id": new_id, "message": "Check-in berhasil."}


class CheckoutRequest(BaseModel):
    checked_out_at: Optional[str] = None
    checkout_latitude: Optional[float] = None
    checkout_longitude: Optional[float] = None
    notes: Optional[str] = None


@app.put("/api/v1/field-activity/{visit_id}/checkout", tags=["Field Activity"])
def field_activity_checkout(visit_id: int, req: CheckoutRequest, user=Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    checked_out_at = req.checked_out_at or datetime.now().isoformat()
    c.execute("""
        UPDATE visit_logs SET
          checked_out_at = %s,
          checkout_latitude = %s,
          checkout_longitude = %s,
          duration_minutes = EXTRACT(EPOCH FROM (%s::timestamp - checked_in_at)) / 60,
          updated_at = NOW()
        WHERE id = %s
    """, (checked_out_at, req.checkout_latitude, req.checkout_longitude, checked_out_at, visit_id))
    if c.rowcount == 0:
        conn.close(); raise HTTPException(404, "Data tidak ditemukan.")
    conn.commit(); conn.close()
    return {"message": "Check-out berhasil."}


@app.delete("/api/v1/field-activity/{visit_id}", tags=["Field Activity"])
def field_activity_delete(visit_id: int, user=Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute("DELETE FROM visit_logs WHERE id=%s", (visit_id,))
    if c.rowcount == 0:
        conn.close(); raise HTTPException(404, "Data tidak ditemukan.")
    conn.commit(); conn.close()
    return {"message": "Data dihapus."}


@app.get("/api/v1/field-activity/monitor", tags=["Field Activity"])
def field_activity_monitor(
    date_from: str = None,
    date_to: str = None,
    user_id: int = None,
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=200),
    current_user=Depends(get_current_user),
):
    import math as _math
    conn = get_conn(); c = conn.cursor()
    today = date.today().isoformat()
    date_from = date_from or today
    date_to   = date_to   or today
    is_sales = current_user.get("role_id") == 3

    where, params = ["DATE(v.checked_in_at) BETWEEN %s AND %s"], [date_from, date_to]
    if is_sales:
        where.append("v.user_id = %s"); params.append(current_user["id"])
    elif user_id:
        where.append("v.user_id = %s"); params.append(user_id)

    where_str = "WHERE " + " AND ".join(where)

    c.execute(f"SELECT COUNT(*) cnt FROM visit_logs v {where_str}", params)
    total_monitor = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page

    c.execute(f"""
        SELECT v.*, DATE(v.checked_in_at) as tgl, u.nama as sales_nama,
               l.nama_company as client_nama
        FROM visit_logs v
        LEFT JOIN users u ON v.user_id = u.id
        LEFT JOIN leads l ON v.lead_id = l.lead_id
        {where_str}
        ORDER BY v.checked_in_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    activities = [_norm(dict(r)) for r in c.fetchall()]

    # Hitung jarak check-in → check-out (Haversine)
    import math
    def _haversine(lat1, lon1, lat2, lon2):
        R = 6371
        dlat = math.radians(float(lat2) - float(lat1))
        dlon = math.radians(float(lon2) - float(lon1))
        a = math.sin(dlat/2)**2 + math.cos(math.radians(float(lat1))) * math.cos(math.radians(float(lat2))) * math.sin(dlon/2)**2
        return round(R * 2 * math.asin(math.sqrt(a)), 3)

    for a in activities:
        lat1, lon1 = a.get("latitude"), a.get("longitude")
        lat2, lon2 = a.get("checkout_latitude"), a.get("checkout_longitude")
        if all(v is not None for v in [lat1, lon1, lat2, lon2]):
            a["distance_km"] = _haversine(lat1, lon1, lat2, lon2)
        else:
            a["distance_km"] = None

    # Map points (hanya yang ada koordinat)
    map_points = [
        {"id": a["id"],
         "latitude": a["latitude"], "longitude": a["longitude"],
         "lat": a["latitude"], "lng": a["longitude"],
         "user_id": a["user_id"], "sales_nama": a["sales_nama"],
         "client_nama": a.get("client_nama"), "lead_id": a.get("lead_id"),
         "address": a["address"], "type": a["type"],
         "checked_in_at": a["checked_in_at"], "time": a["checked_in_at"],
         "duration_minutes": a.get("duration_minutes")}
        for a in activities if a.get("latitude") and a.get("longitude")
    ]

    # Summary aggregate (global)
    c.execute(f"""
        SELECT COUNT(*) as total_kunjungan,
               COUNT(DISTINCT v.user_id) as total_sales,
               COUNT(DISTINCT DATE(v.checked_in_at)) as total_hari,
               ROUND(AVG(v.duration_minutes) FILTER (WHERE v.duration_minutes IS NOT NULL)) as avg_durasi,
               COALESCE(SUM(v.duration_minutes) FILTER (WHERE v.duration_minutes IS NOT NULL), 0) as total_durasi,
               COUNT(*) FILTER (WHERE v.checked_out_at IS NULL) as on_going,
               COUNT(*) FILTER (WHERE v.checked_out_at IS NOT NULL) as completed
        FROM visit_logs v
        LEFT JOIN users u ON v.user_id = u.id
        {where_str}
    """, params)
    agg = _norm(dict(c.fetchone()))

    # Top sales + sales recap (per sales summary)
    c.execute(f"""
        SELECT u.id as user_id, u.nama as sales_nama,
               COUNT(*) as total_kunjungan,
               COUNT(DISTINCT DATE(v.checked_in_at)) as hari_aktif,
               COUNT(*) FILTER (WHERE v.checked_out_at IS NULL) as belum_checkout,
               COUNT(*) FILTER (WHERE v.checked_out_at IS NOT NULL) as completed,
               COALESCE(SUM(v.duration_minutes) FILTER (WHERE v.duration_minutes IS NOT NULL), 0) as total_durasi,
               ROUND(AVG(v.duration_minutes) FILTER (WHERE v.duration_minutes IS NOT NULL)) as avg_durasi
        FROM visit_logs v
        LEFT JOIN users u ON v.user_id = u.id
        {where_str}
        GROUP BY u.id, u.nama ORDER BY total_kunjungan DESC
    """, params)
    top_sales = [_norm(dict(r)) for r in c.fetchall()]
    sales_recap = top_sales

    # Daily recap (kunjungan per hari)
    c.execute(f"""
        SELECT DATE(v.checked_in_at) as tgl,
               u.nama as sales_nama, u.id as user_id,
               COUNT(*) as kunjungan
        FROM visit_logs v
        LEFT JOIN users u ON v.user_id = u.id
        {where_str}
        GROUP BY DATE(v.checked_in_at), u.id, u.nama
        ORDER BY tgl
    """, params)
    daily_recap = [_norm(dict(r)) for r in c.fetchall()]

    # Sales list for filter dropdown
    c.execute("SELECT id, nama FROM users WHERE role_id = 3 AND is_active = 1 ORDER BY nama")
    sales_list = [dict(r) for r in c.fetchall()]

    # Timeline (per visit, used for gantt-style view)
    timeline = [
        {
            "user_id":       a["user_id"],
            "sales_nama":    a["sales_nama"],
            "tgl":           str(a["checked_in_at"])[:10] if a.get("checked_in_at") else "",
            "checked_in_at":  a["checked_in_at"],
            "checked_out_at": a["checked_out_at"],
            "duration_minutes": a.get("duration_minutes"),
            "address":        a.get("address"),
            "type":           a.get("type"),
            "lead_id":        a.get("lead_id"),
            "nama_company":   a.get("nama_company"),
        }
        for a in activities
    ]

    conn.close()
    return {
        "activities":  activities,
        "map_points":  map_points,
        "summary":     agg,
        "top_sales":   top_sales,
        "sales_recap": sales_recap,
        "daily_recap": daily_recap,
        "timeline":    timeline,
        "sales_list":  sales_list,
        "date_from":   date_from,
        "date_to":     date_to,
        "total": total_monitor, "page": page, "per_page": per_page,
        "total_pages": _math.ceil(total_monitor/per_page) if total_monitor else 1,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SCHEMA MIGRATIONS (auto-run on startup)
# ═══════════════════════════════════════════════════════════════════════════════

def _run_migrations():
    conn = get_conn()
    c = conn.cursor()
    migrations = [
        "ALTER TABLE contacts ADD COLUMN IF NOT EXISTS foto TEXT",
        "ALTER TABLE leads ADD COLUMN IF NOT EXISTS loss_reason TEXT",
        "ALTER TABLE leads ADD COLUMN IF NOT EXISTS won_import_excluded BOOLEAN DEFAULT FALSE",
        "ALTER TABLE revenue_projects ADD COLUMN IF NOT EXISTS organisasi TEXT",
        "ALTER TABLE revenue_monthly ADD COLUMN IF NOT EXISTS termin_no INTEGER",
        """CREATE TABLE IF NOT EXISTS annual_targets (
            id             SERIAL PRIMARY KEY,
            tahun          INTEGER NOT NULL,
            bulan          INTEGER NOT NULL,
            organisasi     TEXT    NOT NULL,
            target_revenue NUMERIC(18,2) DEFAULT 0,
            created_at     TIMESTAMP DEFAULT NOW(),
            updated_at     TIMESTAMP DEFAULT NOW(),
            UNIQUE(tahun, bulan, organisasi)
        )""",
        "INSERT INTO role_menus (role_id, menu_key) SELECT id, 'rev_annual_target' FROM roles WHERE nama IN ('Admin','Manager') ON CONFLICT DO NOTHING",
        """CREATE TABLE IF NOT EXISTS annual_target_orgs (
            id         SERIAL PRIMARY KEY,
            tahun      INTEGER NOT NULL,
            organisasi TEXT    NOT NULL,
            UNIQUE(tahun, organisasi)
        )""",
        """CREATE TABLE IF NOT EXISTS sales_targets (
            id          SERIAL PRIMARY KEY,
            sales_nama  VARCHAR(100) NOT NULL,
            tahun       INTEGER NOT NULL,
            bulan       INTEGER NOT NULL,
            target_deal NUMERIC(15,2) DEFAULT 0,
            created_at  TIMESTAMP DEFAULT NOW(),
            updated_at  TIMESTAMP DEFAULT NOW(),
            UNIQUE(sales_nama, tahun, bulan)
        )""",
    ]
    for sql in migrations:
        try:
            c.execute(sql)
        except Exception as e:
            conn.rollback()
            debugPrint = lambda *a: None
    conn.commit()
    conn.close()

try:
    _run_migrations()
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# SALES TARGET INDIVIDU
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/sales-targets", tags=["Sales Target"])
def sales_targets_list(tahun: int = 0, user: dict = Depends(get_current_user)):
    from datetime import date as _date
    cur_year = tahun or _date.today().year
    is_sales = user.get("role_id") == 3
    conn = get_conn()
    c = conn.cursor()

    # Sales hanya lihat dirinya sendiri; admin/manager lihat semua
    if is_sales:
        all_sales = [user["nama"]]
    else:
        c.execute("SELECT DISTINCT nama FROM users WHERE is_active=1 ORDER BY nama")
        all_sales = [r['nama'] for r in c.fetchall()]

    # Target yang sudah diset
    if is_sales:
        c.execute("""
            SELECT sales_nama, bulan, target_deal
            FROM sales_targets WHERE tahun=%s AND sales_nama=%s ORDER BY bulan
        """, (cur_year, user["nama"]))
    else:
        c.execute("""
            SELECT sales_nama, bulan, target_deal
            FROM sales_targets WHERE tahun=%s ORDER BY sales_nama, bulan
        """, (cur_year,))
    target_rows = c.fetchall()
    targets = {}
    for r in target_rows:
        targets[(r['sales_nama'], r['bulan'])] = float(r['target_deal'])

    # Actual per sales per bulan
    if is_sales:
        c.execute("""
            SELECT sales_owner,
                   EXTRACT(MONTH FROM tgl_masuk)::integer as bulan,
                   COALESCE(SUM(deal_value), 0) as actual_deal,
                   COUNT(*) as won_count
            FROM leads
            WHERE stage='Won' AND EXTRACT(YEAR FROM tgl_masuk)=%s
              AND sales_owner=%s
            GROUP BY sales_owner, bulan
        """, (cur_year, user["nama"]))
    else:
        c.execute("""
            SELECT sales_owner,
                   EXTRACT(MONTH FROM tgl_masuk)::integer as bulan,
                   COALESCE(SUM(deal_value), 0) as actual_deal,
                   COUNT(*) as won_count
            FROM leads
            WHERE stage='Won' AND EXTRACT(YEAR FROM tgl_masuk)=%s
              AND sales_owner IS NOT NULL AND sales_owner <> ''
            GROUP BY sales_owner, bulan
        """, (cur_year,))
    actual_rows = c.fetchall()
    actuals = {}
    for r in actual_rows:
        actuals[(r['sales_owner'], r['bulan'])] = {
            'actual': float(r['actual_deal']),
            'won_count': r['won_count'],
        }

    # Susun matrix sales × 12 bulan
    BULAN = ['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Ags','Sep','Okt','Nov','Des']
    result = []
    for sales in all_sales:
        row = {'sales_nama': sales, 'bulan': []}
        ytd_target = 0; ytd_actual = 0
        for m in range(1, 13):
            t = targets.get((sales, m), 0)
            a_data = actuals.get((sales, m), {'actual': 0, 'won_count': 0})
            a = a_data['actual']
            ytd_target += t; ytd_actual += a
            row['bulan'].append({
                'bulan': m,
                'bulan_label': BULAN[m-1],
                'target': t,
                'actual': a,
                'won_count': a_data['won_count'],
                'achievement_pct': round(a / t * 100, 1) if t > 0 else None,
            })
        row['ytd_target'] = ytd_target
        row['ytd_actual'] = ytd_actual
        row['ytd_achievement_pct'] = round(ytd_actual / ytd_target * 100, 1) if ytd_target > 0 else None
        result.append(row)

    conn.close()
    return {"tahun": cur_year, "data": result}


class SalesTargetUpsert(BaseModel):
    sales_nama: str
    tahun: int
    bulan: int
    target_deal: float


@app.post("/api/v1/sales-targets", tags=["Sales Target"])
def sales_target_upsert(payload: SalesTargetUpsert, user: dict = Depends(get_current_user)):
    import traceback as _tb
    conn = get_conn()
    c = conn.cursor()
    try:
        c.execute("""
            INSERT INTO sales_targets (sales_nama, tahun, bulan, target_deal, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (sales_nama, tahun, bulan)
            DO UPDATE SET target_deal=EXCLUDED.target_deal, updated_at=NOW()
        """, (payload.sales_nama, payload.tahun, payload.bulan, payload.target_deal))
        conn.commit()
    except Exception as e:
        conn.rollback()
        _tb.print_exc()
        raise HTTPException(status_code=500, detail=f"DB error: {e}")
    finally:
        conn.close()
    return {"message": "Target berhasil disimpan."}


# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# ANNUAL TARGET (target revenue per bulan per organisasi)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/annual-targets/orgs", tags=["Annual Target"])
def annual_target_orgs_get(tahun: int = Query(0), user: dict = Depends(require_menu("rev_dashboard"))):
    """Semua organisasi + flag apakah dipilih untuk tahun ini."""
    cur_year = tahun if tahun else date.today().year
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT kode, nama FROM organizations WHERE is_active=1 AND deleted_at IS NULL ORDER BY nama")
    all_orgs = [dict(r) for r in c.fetchall()]
    c.execute("SELECT organisasi FROM annual_target_orgs WHERE tahun=%s", (cur_year,))
    selected = {r["organisasi"] for r in c.fetchall()}
    conn.close()
    for o in all_orgs:
        o["selected"] = o["kode"] in selected
    return {"tahun": cur_year, "orgs": all_orgs}


@app.post("/api/v1/annual-targets/orgs", tags=["Annual Target"])
def annual_target_orgs_save(payload: dict, user: dict = Depends(require_admin)):
    """Simpan pilihan organisasi yang punya target untuk tahun ini."""
    tahun    = int(payload.get("tahun", date.today().year))
    selected = payload.get("selected", [])  # list of kode
    conn = get_conn(); c = conn.cursor()
    try:
        c.execute("DELETE FROM annual_target_orgs WHERE tahun=%s", (tahun,))
        for kode in selected:
            c.execute("INSERT INTO annual_target_orgs (tahun, organisasi) VALUES (%s,%s) ON CONFLICT DO NOTHING", (tahun, kode))
        conn.commit()
    except Exception as e:
        conn.rollback(); conn.close()
        raise HTTPException(500, str(e))
    conn.close()
    return {"message": "Organisasi target berhasil disimpan."}


@app.get("/api/v1/annual-targets", tags=["Annual Target"])
def annual_targets_get(tahun: int = Query(0), user: dict = Depends(require_menu("rev_dashboard"))):
    cur_year = tahun if tahun else date.today().year
    conn = get_conn(); c = conn.cursor()
    # Hanya organisasi yang dipilih untuk tahun ini
    c.execute("SELECT organisasi FROM annual_target_orgs WHERE tahun=%s ORDER BY organisasi", (cur_year,))
    selected = [r["organisasi"] for r in c.fetchall()]
    # Nama organisasi
    c.execute("SELECT kode, nama FROM organizations WHERE is_active=1 AND deleted_at IS NULL")
    org_names = {r["kode"]: r["nama"] for r in c.fetchall()}
    lobs = selected
    # Ambil target yang sudah diinput
    c.execute("SELECT bulan, organisasi, target_revenue FROM annual_targets WHERE tahun=%s", (cur_year,))
    rows = c.fetchall()
    conn.close()
    data: dict = {}
    for r in rows:
        b = str(r["bulan"])
        if b not in data: data[b] = {}
        data[b][r["organisasi"]] = float(r["target_revenue"] or 0)
    return {"tahun": cur_year, "lobs": lobs, "org_names": org_names, "data": data}


@app.post("/api/v1/annual-targets", tags=["Annual Target"])
def annual_targets_save(payload: dict, user: dict = Depends(require_admin)):
    """
    payload: { tahun: int, items: [{bulan, organisasi, target_revenue}] }
    """
    tahun = int(payload.get("tahun", date.today().year))
    items = payload.get("items", [])
    conn = get_conn(); c = conn.cursor()
    try:
        for item in items:
            bulan   = int(item["bulan"])
            org     = str(item["organisasi"])
            target  = float(item.get("target_revenue") or 0)
            c.execute("""
                INSERT INTO annual_targets (tahun, bulan, organisasi, target_revenue, updated_at)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (tahun, bulan, organisasi) DO UPDATE
                  SET target_revenue=EXCLUDED.target_revenue, updated_at=NOW()
            """, (tahun, bulan, org, target))
        conn.commit()
    except Exception as e:
        conn.rollback(); conn.close()
        raise HTTPException(500, str(e))
    conn.close()
    return {"message": "Target berhasil disimpan."}


@app.get("/api/v1/annual-targets/summary", tags=["Annual Target"])
def annual_targets_summary(tahun: int = Query(0), user: dict = Depends(require_menu("rev_dashboard"))):
    """Perbandingan annual target vs realisasi per bulan dan per LOB."""
    cur_year = tahun if tahun else date.today().year
    conn = get_conn(); c = conn.cursor()

    # Target per bulan per LOB
    c.execute("""
        SELECT bulan, organisasi, SUM(target_revenue) AS target
        FROM annual_targets WHERE tahun=%s
        GROUP BY bulan, organisasi ORDER BY bulan, organisasi
    """, (cur_year,))
    target_rows = c.fetchall()

    # Realisasi per bulan per LOB (dari revenue_monthly join revenue_projects)
    c.execute("""
        SELECT m.month_num AS bulan, p.organisasi, SUM(m.actual) AS actual
        FROM revenue_monthly m
        JOIN revenue_projects p ON m.project_id = p.project_id
        WHERE p.tahun=%s AND p.is_active=1
        GROUP BY m.month_num, p.organisasi ORDER BY m.month_num, p.organisasi
    """, (cur_year,))
    actual_rows = c.fetchall()

    # Hanya organisasi yang dipilih untuk tahun ini
    c.execute("SELECT organisasi FROM annual_target_orgs WHERE tahun=%s ORDER BY organisasi", (cur_year,))
    lobs = [r["organisasi"] for r in c.fetchall()]
    c.execute("SELECT kode, nama FROM organizations WHERE is_active=1 AND deleted_at IS NULL")
    org_names = {r["kode"]: r["nama"] for r in c.fetchall()}

    conn.close()

    # Susun per bulan
    monthly = []
    for m in range(1, 13):
        entry = {"bulan": m, "bulan_nama": MONTHS_EN[m-1], "total_target": 0, "total_actual": 0, "by_lob": {}}
        for lob in lobs:
            entry["by_lob"][lob] = {"target": 0, "actual": 0}
        monthly.append(entry)

    for r in target_rows:
        m = int(r["bulan"]) - 1
        org = r["organisasi"] or ""
        t   = float(r["target"] or 0)
        monthly[m]["total_target"] += t
        if org in monthly[m]["by_lob"]:
            monthly[m]["by_lob"][org]["target"] = t

    for r in actual_rows:
        m = int(r["bulan"]) - 1
        org = r["organisasi"] or ""
        a   = float(r["actual"] or 0)
        if org in monthly[m]["by_lob"]:
            monthly[m]["by_lob"][org]["actual"] = a
            monthly[m]["total_actual"] += a  # hanya org yang selected

    # Totals per LOB
    lob_summary = {}
    for lob in lobs:
        lob_summary[lob] = {
            "target": sum(monthly[m]["by_lob"].get(lob, {}).get("target", 0) for m in range(12)),
            "actual": sum(monthly[m]["by_lob"].get(lob, {}).get("actual", 0) for m in range(12)),
        }

    grand_target = sum(r["total_target"] for r in monthly)
    grand_actual = sum(r["total_actual"] for r in monthly)

    cur_month = date.today().month if cur_year == date.today().year else 12
    ytd_target = sum(r["total_target"] for r in monthly if r["bulan"] <= cur_month)
    ytd_actual = sum(r["total_actual"] for r in monthly if r["bulan"] <= cur_month)
    ytd_ach    = round(ytd_actual / max(ytd_target, 1) * 100, 1) if ytd_target else 0
    ytd_gap    = ytd_target - ytd_actual

    # Realisasi vs Target per Kategori (Project / Recurring)
    conn2 = get_conn(); c2 = conn2.cursor()
    c2.execute("""
        SELECT COALESCE(kategori, 'Lainnya') AS kategori,
               COALESCE(SUM(revenue_target), 0) AS target,
               COALESCE(SUM(actual_revenue),  0) AS actual
        FROM revenue_projects
        WHERE tahun=%s AND is_active=1 AND deleted_at IS NULL
          AND project_status IN ('Active','Completed')
        GROUP BY kategori ORDER BY kategori
    """, (cur_year,))
    kat_rows = c2.fetchall()
    conn2.close()

    kategori_summary = [
        {
            "kategori": r["kategori"],
            "target":   float(r["target"]),
            "actual":   float(r["actual"]),
        }
        for r in kat_rows
    ]

    return {
        "tahun": cur_year, "lobs": lobs, "org_names": org_names,
        "monthly": monthly,
        "lob_summary": lob_summary,
        "kategori_summary": kategori_summary,
        "grand_target": grand_target,
        "grand_actual": grand_actual,
        "grand_ach": round(grand_actual / max(grand_target, 1) * 100, 1),
        "ytd_target": ytd_target,
        "ytd_actual": ytd_actual,
        "ytd_ach": ytd_ach,
        "ytd_gap": ytd_gap,
        "cur_month": cur_month,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/export/pipeline", tags=["Export"])
def export_pipeline(
    stage: str = "", segmen: str = "", sales: str = "",
    user: dict = Depends(require_menu("export"))
):
    import csv, io
    conn = get_conn()
    c = conn.cursor()
    is_sales = user.get("role_id") == 3
    query = "SELECT * FROM leads WHERE 1=1"
    params = []
    if is_sales:
        query += " AND sales_owner=%s"; params.append(user["nama"])
    else:
        if stage:  query += " AND stage=%s";       params.append(stage)
        if segmen: query += " AND segmen=%s";      params.append(segmen)
        if sales:  query += " AND sales_owner=%s"; params.append(sales)
    if is_sales and stage:  query += " AND stage=%s";  params.append(stage)
    if is_sales and segmen: query += " AND segmen=%s"; params.append(segmen)
    query += " ORDER BY id DESC"
    c.execute(query, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    HEADERS = [
        "lead_id","nama_company","product","contact_person","segmen","sub_segmen",
        "source","stage","prioritas","tgl_masuk","propose_value","deal_value",
        "probability","weighted_value","exp_close_date","sales_owner",
        "next_fu_date","last_fu_date","remarks","loss_reason",
    ]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for r in rows:
        w.writerow([r.get(h, "") or "" for h in HEADERS])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": 'attachment; filename="pipeline_export.csv"'},
    )


@app.get("/api/v1/export/daily-reports", tags=["Export"])
def export_daily_reports(
    month: str = "", sales: str = "",
    user: dict = Depends(get_current_user)
):
    import csv, io
    is_sales = user.get("role_id") == 3
    conn = get_conn()
    c = conn.cursor()
    query = """
        SELECT dr.*, u.nama as sales_nama
        FROM daily_reports dr
        LEFT JOIN users u ON u.id = dr.user_id
        WHERE 1=1
    """
    params = []
    if is_sales:
        query += " AND dr.user_id=%s"; params.append(user["id"])
    elif sales:
        query += " AND u.nama=%s"; params.append(sales)
    if month:
        query += " AND TO_CHAR(dr.report_date,'YYYY-MM')=%s"; params.append(month)
    query += " ORDER BY dr.report_date DESC"
    c.execute(query, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    HEADERS = ["report_date","sales_nama","status","visit_count","fu_count",
               "new_lead_count","mood","notes_obstacle","notes_plan","sent_at"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for r in rows:
        w.writerow([str(r.get(h, "") or "") for h in HEADERS])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": 'attachment; filename="daily_reports_export.csv"'},
    )


@app.get("/api/v1/export/analytics", tags=["Export"])
def export_analytics(tahun: int = 0, user: dict = Depends(get_current_user)):
    import csv, io
    from datetime import date as _date
    cur_year = tahun or _date.today().year
    is_sales = user.get("role_id") == 3
    conn = get_conn()
    c = conn.cursor()

    # Per-sales summary
    sf = " AND sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []
    c.execute(f"""
        SELECT
            sales_owner as sales,
            COUNT(*) as total_leads,
            COUNT(CASE WHEN stage='Won' THEN 1 END) as won,
            COUNT(CASE WHEN stage='Lost' THEN 1 END) as lost,
            COALESCE(SUM(CASE WHEN stage='Won' THEN deal_value ELSE 0 END), 0) as total_deal,
            COALESCE(SUM(propose_value), 0) as total_pipeline,
            COALESCE(AVG(probability), 0) as avg_prob
        FROM leads
        WHERE EXTRACT(YEAR FROM tgl_masuk)=%s
          AND sales_owner IS NOT NULL AND sales_owner <> ''{sf}
        GROUP BY sales_owner ORDER BY total_deal DESC
    """, [cur_year] + sp)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    HEADERS = ["sales","total_leads","won","lost","total_deal","total_pipeline",
               "win_rate_pct","avg_prob"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for r in rows:
        total = r['total_leads'] or 1
        r['win_rate_pct'] = round(r['won'] / total * 100, 1)
        w.writerow([str(r.get(h, "") or "") for h in HEADERS])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="analytics_{cur_year}.csv"'},
    )


# ── PDF Helpers ───────────────────────────────────────────────────────────────
def _pdf_header(pdf, title: str, subtitle: str = ""):
    from reportlab.lib import colors
    pdf.setFillColor(colors.HexColor("#0f1f36"))
    pdf.rect(0, pdf.pagesizes[1] - 60, pdf.pagesizes[0], 60, fill=1, stroke=0)  # type: ignore
    pdf.setFillColor(colors.HexColor("#ffffff"))
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(30, pdf.pagesizes[1] - 38, title)  # type: ignore
    if subtitle:
        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(colors.HexColor("#94a3b8"))
        pdf.drawString(30, pdf.pagesizes[1] - 52, subtitle)  # type: ignore


def _make_pdf(title: str, subtitle: str, headers: list, rows: list, col_widths: list):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet
    import io as _io

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()

    NAVY   = colors.HexColor("#0f1f36")
    ACCENT = colors.HexColor("#3b82f6")
    LIGHT  = colors.HexColor("#e2e8f0")
    ALTROW = colors.HexColor("#f1f5f9")

    para_style = styles["Normal"].clone("cell")
    para_style.fontSize = 7
    para_style.leading  = 9

    # Wrap long text in cells
    table_data = [headers]
    for row in rows:
        table_data.append([Paragraph(str(v or ""), para_style) for v in row])

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",   (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING",(0, 0), (-1, 0), 6),
        ("TOPPADDING",   (0, 0), (-1, 0), 6),
        # Data rows
        ("FONTNAME",     (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, ALTROW]),
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",   (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        # Accent line under header
        ("LINEBELOW",    (0, 0), (-1, 0), 1.5, ACCENT),
    ]))

    header_para = Paragraph(
        f'<font size="14" color="#0f1f36"><b>{title}</b></font><br/>'
        f'<font size="8" color="#64748b">{subtitle}</font>',
        styles["Normal"]
    )
    story = [header_para, Spacer(1, 6*mm), t]
    doc.build(story)
    buf.seek(0)
    return buf


@app.get("/api/v1/export/pipeline/pdf", tags=["Export"])
def export_pipeline_pdf(
    stage: str = "", sales: str = "",
    user: dict = Depends(require_menu("export"))
):
    from datetime import date as _date
    conn = get_conn(); c = conn.cursor()
    is_sales = user.get("role_id") == 3
    q = "SELECT * FROM leads WHERE 1=1"
    p = []
    if is_sales:
        q += " AND sales_owner=%s"; p.append(user["nama"])
    else:
        if stage: q += " AND stage=%s"; p.append(stage)
        if sales: q += " AND sales_owner=%s"; p.append(sales)
    q += " ORDER BY stage, id DESC"
    c.execute(q, p)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    headers = ["ID","Perusahaan","Produk","Stage","Prioritas","Sales","Propose (Jt)","Deal (Jt)","Prob%","Exp. Close"]
    mm = __import__('reportlab.lib.units', fromlist=['mm']).mm
    col_w = [18*mm, 45*mm, 30*mm, 22*mm, 18*mm, 22*mm, 22*mm, 22*mm, 14*mm, 22*mm]
    data  = []
    for r in rows:
        data.append([
            r["lead_id"],
            r["nama_company"],
            r["product"] or "",
            r["stage"],
            r["prioritas"] or "",
            r["sales_owner"] or "",
            f"{(r['propose_value'] or 0)/1e6:.1f}",
            f"{(r['deal_value'] or 0)/1e6:.1f}",
            f"{r['probability'] or 0:.0f}%",
            str(r["exp_close_date"] or ""),
        ])

    subtitle = f"Digenerate: {_date.today().isoformat()} · {len(rows)} leads"
    if stage: subtitle += f" · Stage: {stage}"
    if sales: subtitle += f" · Sales: {sales}"
    buf = _make_pdf("Pipeline Report — APEX CRM", subtitle, headers, data, col_w)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="pipeline_report.pdf"'})


@app.get("/api/v1/export/daily-reports/pdf", tags=["Export"])
def export_daily_reports_pdf(
    month: str = "", sales: str = "",
    user: dict = Depends(get_current_user)
):
    from datetime import date as _date
    is_sales = user.get("role_id") == 3
    conn = get_conn(); c = conn.cursor()
    q = """SELECT dr.report_date, u.nama as sales_nama, dr.status, dr.visit_count,
                  dr.fu_count, dr.new_lead_count, dr.mood, dr.notes_obstacle, dr.notes_plan
           FROM daily_reports dr LEFT JOIN users u ON u.id=dr.user_id WHERE 1=1"""
    p = []
    if is_sales:
        q += " AND dr.user_id=%s"; p.append(user["id"])
    elif sales:
        q += " AND u.nama=%s"; p.append(sales)
    if month:
        q += " AND TO_CHAR(dr.report_date,'YYYY-MM')=%s"; p.append(month)
    q += " ORDER BY dr.report_date DESC"
    c.execute(q, p)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    mm = __import__('reportlab.lib.units', fromlist=['mm']).mm
    headers = ["Tanggal","Sales","Status","Kunjungan","FU","Lead Baru","Mood","Kendala","Rencana"]
    col_w = [22*mm, 25*mm, 18*mm, 16*mm, 12*mm, 16*mm, 16*mm, 50*mm, 50*mm]
    data  = [[str(r["report_date"] or ""), r["sales_nama"] or "", r["status"] or "",
              str(r["visit_count"] or 0), str(r["fu_count"] or 0), str(r["new_lead_count"] or 0),
              r["mood"] or "", r["notes_obstacle"] or "", r["notes_plan"] or ""] for r in rows]

    label = month or _date.today().strftime("%Y-%m")
    buf = _make_pdf(f"Laporan Harian Sales — {label}", f"{len(rows)} record", headers, data, col_w)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laporan_harian_{label}.pdf"'})


@app.get("/api/v1/export/analytics/pdf", tags=["Export"])
def export_analytics_pdf(tahun: int = 0, user: dict = Depends(get_current_user)):
    from datetime import date as _date
    cur_year = tahun or _date.today().year
    is_sales = user.get("role_id") == 3
    conn = get_conn(); c = conn.cursor()
    sf = " AND sales_owner=%s" if is_sales else ""
    sp = [user["nama"]] if is_sales else []
    c.execute(f"""
        SELECT sales_owner as sales, COUNT(*) total, COUNT(CASE WHEN stage='Won' THEN 1 END) won,
               COUNT(CASE WHEN stage='Lost' THEN 1 END) lost,
               COALESCE(SUM(CASE WHEN stage='Won' THEN deal_value ELSE 0 END),0) total_deal,
               COALESCE(SUM(propose_value),0) total_pipeline,
               COALESCE(AVG(probability),0) avg_prob
        FROM leads WHERE EXTRACT(YEAR FROM tgl_masuk)=%s AND sales_owner IS NOT NULL AND sales_owner<>''{sf}
        GROUP BY sales_owner ORDER BY total_deal DESC
    """, [cur_year] + sp)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    mm = __import__('reportlab.lib.units', fromlist=['mm']).mm
    headers = ["Sales","Total Leads","Won","Lost","Win Rate","Total Deal (Jt)","Pipeline (Jt)","Avg Prob%"]
    col_w = [35*mm, 22*mm, 16*mm, 16*mm, 18*mm, 28*mm, 28*mm, 22*mm]
    data  = []
    for r in rows:
        wr = f"{r['won']/r['total']*100:.1f}%" if r["total"] else "0%"
        data.append([r["sales"], str(r["total"]), str(r["won"]), str(r["lost"]), wr,
                     f"{r['total_deal']/1e6:.1f}", f"{r['total_pipeline']/1e6:.1f}",
                     f"{r['avg_prob']:.1f}%"])
    buf = _make_pdf(f"Analytics Sales {cur_year} — APEX CRM",
                    f"Digenerate: {_date.today().isoformat()}", headers, data, col_w)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="analytics_{cur_year}.pdf"'})


# ═══════════════════════════════════════════════════════════════════════════════
# LOCATION TRACKING
# ═══════════════════════════════════════════════════════════════════════════════

def _ensure_location_tables():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS location_logs (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL,
            latitude    DOUBLE PRECISION NOT NULL,
            longitude   DOUBLE PRECISION NOT NULL,
            accuracy_m  INTEGER,
            speed_kmh   VARCHAR(10),
            recorded_at TIMESTAMP NOT NULL DEFAULT NOW(),
            created_at  TIMESTAMP DEFAULT NOW()
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS location_settings (
            id                       SERIAL PRIMARY KEY,
            location_tracking_enabled BOOLEAN DEFAULT FALSE,
            updated_at               TIMESTAMP DEFAULT NOW()
        )
    """)
    c.execute("SELECT COUNT(*) as cnt FROM location_settings")
    if c.fetchone()["cnt"] == 0:
        c.execute("INSERT INTO location_settings (location_tracking_enabled) VALUES (FALSE)")
    conn.commit(); conn.close()

try:
    _ensure_location_tables()
except Exception:
    pass


@app.get("/api/v1/location/settings", tags=["Location"])
def location_settings(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT location_tracking_enabled FROM location_settings LIMIT 1")
    row = c.fetchone()
    conn.close()
    return {"location_tracking_enabled": bool(row["location_tracking_enabled"]) if row else False}


@app.post("/api/v1/location", tags=["Location"])
def location_post(payload: dict, user: dict = Depends(get_current_user)):
    uid = user["id"]
    conn = get_conn()
    c = conn.cursor()
    recorded_at = payload.get("recorded_at") or datetime.utcnow().isoformat()
    c.execute("""
        INSERT INTO location_logs (user_id, latitude, longitude, accuracy_m, speed_kmh, recorded_at)
        VALUES (%s,%s,%s,%s,%s,%s)
    """, (uid, payload.get("latitude"), payload.get("longitude"),
          payload.get("accuracy_m"), payload.get("speed_kmh"), recorded_at))
    conn.commit(); conn.close()
    return {"message": "Lokasi diterima."}


@app.get("/api/v1/location/me", tags=["Location"])
def location_me(limit: int = 20, user: dict = Depends(get_current_user)):
    uid = user["id"]
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        SELECT * FROM location_logs WHERE user_id=%s
        ORDER BY recorded_at DESC LIMIT %s
    """, (uid, limit))
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"locations": rows}


@app.get("/api/v1/location/team", tags=["Location"])
def location_team(user: dict = Depends(get_current_user)):
    from datetime import datetime as _dt, timedelta as _td
    conn = get_conn()
    c = conn.cursor()

    # Posisi terakhir tiap user hari ini dari location_logs
    c.execute("""
        SELECT DISTINCT ON (ll.user_id)
            ll.user_id, u.nama as sales_nama, ll.latitude, ll.longitude,
            ll.accuracy_m, ll.speed_kmh, ll.recorded_at as last_seen
        FROM location_logs ll
        JOIN users u ON u.id = ll.user_id
        WHERE ll.recorded_at >= (CURRENT_DATE AT TIME ZONE 'Asia/Jakarta')
          AND ll.recorded_at <  ((CURRENT_DATE + 1) AT TIME ZONE 'Asia/Jakarta')
        ORDER BY ll.user_id, ll.recorded_at DESC
    """)
    loc_rows = {r["user_id"]: _norm(dict(r)) for r in c.fetchall()}

    # Tambah posisi dari check-in terbaru hari ini jika tidak ada di location_logs
    c.execute("""
        SELECT DISTINCT ON (v.user_id)
            v.user_id, u.nama as sales_nama,
            v.latitude, v.longitude, v.address,
            COALESCE(v.checked_out_at, v.checked_in_at) as last_seen
        FROM visit_logs v
        JOIN users u ON u.id = v.user_id
        WHERE v.checked_in_at >= (CURRENT_DATE AT TIME ZONE 'Asia/Jakarta')
          AND v.checked_in_at <  ((CURRENT_DATE + 1) AT TIME ZONE 'Asia/Jakarta')
          AND v.latitude IS NOT NULL
        ORDER BY v.user_id, COALESCE(v.checked_out_at, v.checked_in_at) DESC
    """)
    for r in c.fetchall():
        uid = r["user_id"]
        if uid not in loc_rows:
            loc_rows[uid] = _norm(dict(r))

    # Semua user aktif
    c.execute("SELECT id, nama FROM users WHERE is_active=1 ORDER BY nama")
    all_users = {r["id"]: r["nama"] for r in c.fetchall()}

    # Cari siapa yang sedang check-in (belum checkout) hari ini
    c.execute("""
        SELECT DISTINCT ON (user_id) user_id
        FROM visit_logs
        WHERE DATE(checked_in_at AT TIME ZONE 'Asia/Jakarta') = CURRENT_DATE AT TIME ZONE 'Asia/Jakarta'
          AND checked_out_at IS NULL
        ORDER BY user_id, checked_in_at DESC
    """)
    open_checkin_ids = {r["user_id"] for r in c.fetchall()}
    conn.close()

    active, offline = [], []
    active_ids = set()

    for uid, loc in loc_rows.items():
        is_online = uid in open_checkin_ids
        entry = {**loc, "is_online": is_online}
        if is_online:
            active.append(entry)
        else:
            offline.append(entry)
        active_ids.add(uid)

    for uid, nama in all_users.items():
        if uid not in active_ids:
            offline.append({"user_id": uid, "sales_nama": nama, "is_online": False})

    return {
        "active": active,
        "offline": offline,
        "active_count": len(active),
        "offline_count": len(offline),
    }


@app.get("/api/v1/location/team/trails", tags=["Location"])
def location_team_trails(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()

    # Trail dari location_logs hari ini
    c.execute("""
        SELECT ll.user_id, u.nama as sales_nama,
               ll.latitude as lat, ll.longitude as lng,
               ll.recorded_at as ts
        FROM location_logs ll
        JOIN users u ON u.id = ll.user_id
        WHERE ll.recorded_at >= (CURRENT_DATE AT TIME ZONE 'Asia/Jakarta')
          AND ll.recorded_at <  ((CURRENT_DATE + 1) AT TIME ZONE 'Asia/Jakarta')
          AND ll.latitude IS NOT NULL
        ORDER BY ll.user_id, ll.recorded_at ASC
        LIMIT 500
    """)
    loc_rows = c.fetchall()

    # Trail dari visit_logs hari ini (check-in/out sebagai titik)
    c.execute("""
        SELECT v.user_id, u.nama as sales_nama,
               v.latitude as lat, v.longitude as lng,
               v.checked_in_at as ts
        FROM visit_logs v
        JOIN users u ON u.id = v.user_id
        WHERE DATE(v.checked_in_at) = CURRENT_DATE
          AND v.latitude IS NOT NULL
        ORDER BY v.user_id, v.checked_in_at ASC
    """)
    visit_rows = c.fetchall()
    conn.close()

    # Gabungkan dan kelompokkan per user
    from collections import defaultdict
    trails: dict = defaultdict(lambda: {"sales_nama": "", "points": []})
    for r in list(loc_rows) + list(visit_rows):
        uid = r["user_id"]
        trails[uid]["user_id"] = uid
        trails[uid]["sales_nama"] = r["sales_nama"]
        trails[uid]["points"].append({
            "lat": float(r["lat"]),
            "lng": float(r["lng"]),
            "ts": str(r["ts"]),
        })

    # Sort points per user by timestamp
    result = []
    for uid, t in trails.items():
        t["points"].sort(key=lambda p: p["ts"])
        result.append(t)

    return {"trails": result}


# ═══════════════════════════════════════════════════════════════════════════════
# CONTACTS — edit & delete
# ═══════════════════════════════════════════════════════════════════════════════

class ContactUpdate(BaseModel):
    lead_id: Optional[str] = None
    nama_company: Optional[str] = ""
    nama_contact: str = ""
    jabatan: Optional[str] = ""
    dept: Optional[str] = ""
    role: Optional[str] = ""
    no_hp: Optional[str] = None
    email: Optional[str] = None
    telepon: Optional[str] = None
    linkedin: Optional[str] = None
    preferensi_kontak: Optional[str] = None
    catatan: Optional[str] = None


@app.put("/api/v1/contacts/{contact_id}", tags=["Contacts"])
def contact_update(contact_id: int, payload: ContactUpdate,
                   user: dict = Depends(require_menu("contacts"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM contacts WHERE id=%s", (contact_id,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Kontak tidak ditemukan.")
    # Jika lead_id berubah, ambil nama_company dari lead
    nama_co = payload.nama_company or ""
    if payload.lead_id:
        c.execute("SELECT nama_company FROM leads WHERE lead_id=%s", (payload.lead_id,))
        row = c.fetchone()
        if row:
            nama_co = row["nama_company"]
    c.execute("""UPDATE contacts SET
        lead_id=%s, nama_company=%s, nama_contact=%s,
        jabatan=%s, dept=%s, role=%s, no_hp=%s,
        email=%s, telepon=%s, linkedin=%s, preferensi_kontak=%s, catatan=%s
        WHERE id=%s""",
        (payload.lead_id or None, nama_co, payload.nama_contact,
         payload.jabatan, payload.dept, payload.role,
         payload.no_hp or None, payload.email or None, payload.telepon or None,
         payload.linkedin or None, payload.preferensi_kontak or None,
         payload.catatan or None, contact_id))
    conn.commit(); conn.close()
    return {"message": "Kontak berhasil diperbarui."}


@app.delete("/api/v1/contacts/{contact_id}", tags=["Contacts"])
def contact_delete(contact_id: int, user: dict = Depends(require_menu("contacts"))):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM contacts WHERE id=%s", (contact_id,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Kontak tidak ditemukan.")
    c.execute("DELETE FROM contacts WHERE id=%s", (contact_id,))
    conn.commit(); conn.close()
    return {"message": "Kontak berhasil dihapus."}


@app.post("/api/v1/contacts/{contact_id}/foto", tags=["Contacts"])
async def contact_upload_foto(
    contact_id: int,
    file: UploadFile = File(...),
    user: dict = Depends(require_menu("contacts"))
):
    import shutil, uuid
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT id, foto FROM contacts WHERE id=%s", (contact_id,))
    row = c.fetchone()
    if not row:
        conn.close(); raise HTTPException(404, "Kontak tidak ditemukan.")

    # Hapus foto lama jika ada
    if row["foto"]:
        old_path = os.path.join(os.path.dirname(__file__), "static", row["foto"])
        if os.path.exists(old_path):
            os.remove(old_path)

    ext = os.path.splitext(file.filename or "")[1].lower() or ".jpg"
    fname = f"{uuid.uuid4().hex}{ext}"
    upload_dir = os.path.join(os.path.dirname(__file__), "static", "uploads", "contacts")
    os.makedirs(upload_dir, exist_ok=True)
    with open(os.path.join(upload_dir, fname), "wb") as f:
        shutil.copyfileobj(file.file, f)
    foto_path = f"uploads/contacts/{fname}"
    c.execute("UPDATE contacts SET foto=%s, updated_at=NOW() WHERE id=%s", (foto_path, contact_id))
    conn.commit(); conn.close()
    return {"foto": foto_path}


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTS — master data untuk dropdown pipeline
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/products", tags=["Master Data"])
def products_list(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    # Ambil produk unik dari leads yang pernah diinput
    c.execute("""
        SELECT DISTINCT TRIM(product) as product
        FROM leads
        WHERE product IS NOT NULL AND TRIM(product) <> ''
        ORDER BY 1
    """)
    rows = [r["product"] for r in c.fetchall()]
    conn.close()
    return {"products": rows}


@app.post("/api/v1/products", status_code=201, tags=["Master Data"])
def product_create(payload: dict, user: dict = Depends(get_current_user)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Nama produk tidak boleh kosong.")
    conn = get_conn()
    c = conn.cursor()
    # Simpan ke tabel master_products jika ada, atau kembalikan saja
    try:
        c.execute("INSERT INTO master_products (name) VALUES (%s) ON CONFLICT DO NOTHING", (name,))
        conn.commit()
    except Exception:
        conn.rollback()
    conn.close()
    return {"message": "Produk ditambahkan.", "name": name}


# ═══════════════════════════════════════════════════════════════════════════════
# DAILY REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def _ensure_daily_report_table():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS daily_reports (
            id           SERIAL PRIMARY KEY,
            user_id      INTEGER NOT NULL,
            report_date  DATE NOT NULL DEFAULT CURRENT_DATE,
            status       VARCHAR(10) NOT NULL DEFAULT 'draft',
            visit_count  INTEGER DEFAULT 0,
            fu_count     INTEGER DEFAULT 0,
            new_lead_count INTEGER DEFAULT 0,
            notes_obstacle TEXT,
            notes_plan     TEXT,
            mood           VARCHAR(10),
            sent_at        TIMESTAMP,
            created_at     TIMESTAMP DEFAULT NOW(),
            updated_at     TIMESTAMP DEFAULT NOW(),
            UNIQUE(user_id, report_date)
        )
    """)
    conn.commit(); conn.close()

try:
    _ensure_daily_report_table()
except Exception:
    pass


@app.get("/api/v1/daily-report/summary", tags=["Daily Report"])
def daily_report_summary(date: str = "", user: dict = Depends(get_current_user)):
    from datetime import date as _date
    target = date or _date.today().isoformat()
    uid = user["id"]
    conn = get_conn()
    c = conn.cursor()

    # Hitung kunjungan hari ini
    c.execute("""
        SELECT COUNT(*) as cnt FROM visit_logs
        WHERE user_id=%s AND DATE(checked_in_at)=%s
    """, (uid, target))
    visit_count = (c.fetchone() or {}).get("cnt", 0)

    # Hitung FU hari ini
    c.execute("""
        SELECT COUNT(*) as cnt FROM follow_up_log
        WHERE sales_owner=%s AND DATE(tgl_fu)=%s
    """, (user.get("nama", ""), target))
    fu_count = (c.fetchone() or {}).get("cnt", 0)

    # Hitung lead baru hari ini
    c.execute("""
        SELECT COUNT(*) as cnt FROM leads
        WHERE sales_owner=%s AND DATE(tgl_masuk)=%s
    """, (user.get("nama", ""), target))
    new_lead_count = (c.fetchone() or {}).get("cnt", 0)

    # Detail kunjungan
    c.execute("""
        SELECT vl.id, vl.lead_id, l.nama_company, vl.checked_in_at, vl.checked_out_at,
               vl.address, vl.notes
        FROM visit_logs vl
        LEFT JOIN leads l ON l.lead_id = vl.lead_id
        WHERE vl.user_id=%s AND DATE(vl.checked_in_at)=%s
        ORDER BY vl.checked_in_at
    """, (uid, target))
    details = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()

    return {
        "date": target,
        "visit_count": visit_count,
        "fu_count": fu_count,
        "new_lead_count": new_lead_count,
        "visit_details": details,
    }


@app.get("/api/v1/daily-report", tags=["Daily Report"])
def daily_report_list(month: str = "",
                      user_id: int = None,
                      page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=100),
                      user: dict = Depends(get_current_user)):
    import math
    from datetime import date as _date
    uid = user["id"]
    role = user.get("role", "sales")
    conn = get_conn()
    c = conn.cursor()

    where, params = [], []
    if role == "sales":
        where.append("dr.user_id=%s"); params.append(uid)
    elif user_id:
        where.append("dr.user_id=%s"); params.append(user_id)
    if month:
        where.append("TO_CHAR(dr.report_date,'YYYY-MM')=%s"); params.append(month)

    where_str = ("WHERE " + " AND ".join(where)) if where else ""
    c.execute(f"SELECT COUNT(*) cnt FROM daily_reports dr {where_str}", params)
    total = int(c.fetchone()["cnt"])
    offset = (page - 1) * per_page
    c.execute(f"""
        SELECT dr.*, u.nama as sales_nama
        FROM daily_reports dr
        LEFT JOIN users u ON u.id = dr.user_id
        {where_str}
        ORDER BY dr.report_date DESC LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    reports = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"reports": reports, "total": total, "page": page, "per_page": per_page, "total_pages": math.ceil(total/per_page) if total else 1}


@app.post("/api/v1/daily-report", status_code=201, tags=["Daily Report"])
def daily_report_create(payload: dict, user: dict = Depends(get_current_user)):
    from datetime import date as _date
    uid = user["id"]
    report_date = payload.get("report_date") or _date.today().isoformat()
    conn = get_conn()
    c = conn.cursor()
    # Cek apakah sudah ada laporan hari ini
    c.execute("SELECT id FROM daily_reports WHERE user_id=%s AND report_date=%s",
              (uid, report_date))
    existing = c.fetchone()
    if existing:
        conn.close()
        raise HTTPException(status_code=422,
                            detail={"message": "Laporan hari ini sudah ada.",
                                    "report_id": existing["id"]})
    c.execute("""
        INSERT INTO daily_reports
            (user_id, report_date, status, visit_count, fu_count, new_lead_count,
             notes_obstacle, notes_plan, mood)
        VALUES (%s,%s,'draft',%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (uid, report_date,
          payload.get("visit_count", 0),
          payload.get("fu_count", 0),
          payload.get("new_lead_count", 0),
          payload.get("notes_obstacle"),
          payload.get("notes_plan"),
          payload.get("mood")))
    rid = c.fetchone()["id"]
    conn.commit(); conn.close()
    return {"report_id": rid, "message": "Laporan berhasil dibuat."}


@app.put("/api/v1/daily-report/{report_id}", tags=["Daily Report"])
def daily_report_update(report_id: int, payload: dict,
                        user: dict = Depends(get_current_user)):
    uid = user["id"]
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM daily_reports WHERE id=%s AND user_id=%s",
              (report_id, uid))
    if not c.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan.")
    c.execute("""
        UPDATE daily_reports SET
            notes_obstacle=%s, notes_plan=%s, mood=%s, updated_at=NOW()
        WHERE id=%s AND user_id=%s
    """, (payload.get("notes_obstacle"), payload.get("notes_plan"),
          payload.get("mood"), report_id, uid))
    conn.commit(); conn.close()
    return {"message": "Laporan berhasil diperbarui."}


@app.post("/api/v1/daily-report/{report_id}/send", tags=["Daily Report"])
def daily_report_send(report_id: int, user: dict = Depends(get_current_user)):
    uid = user["id"]
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id, status FROM daily_reports WHERE id=%s AND user_id=%s",
              (report_id, uid))
    row = c.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Laporan tidak ditemukan.")
    if row["status"] == "sent":
        conn.close()
        raise HTTPException(status_code=422, detail="Laporan sudah pernah dikirim.")
    c.execute("""
        UPDATE daily_reports SET status='sent', sent_at=NOW(), updated_at=NOW()
        WHERE id=%s
    """, (report_id,))
    conn.commit(); conn.close()
    return {"message": "Laporan berhasil dikirim ke manager."}


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYTICS PERSONAL
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/analytics/personal", tags=["Analytics"])
def analytics_personal(user: dict = Depends(get_current_user)):
    sales_name = user.get("nama", "")
    uid = user["id"]
    conn = get_conn()
    c = conn.cursor()

    from datetime import date as _date
    today = _date.today()
    cur_year = today.year
    cur_month = today.month

    # Summary keseluruhan
    c.execute("""
        SELECT
            COUNT(*) as total_leads,
            SUM(CASE WHEN stage='Won' THEN 1 ELSE 0 END) as won,
            SUM(CASE WHEN stage='Lost' THEN 1 ELSE 0 END) as lost,
            COALESCE(SUM(CASE WHEN stage='Won' THEN deal_value ELSE 0 END),0) as total_deal_won
        FROM leads WHERE sales_owner=%s
    """, (sales_name,))
    row = dict(c.fetchone() or {})
    total = row.get("total_leads") or 0
    won   = row.get("won") or 0
    lost  = row.get("lost") or 0
    total_deal_won = float(row.get("total_deal_won") or 0)
    win_rate = round((won / total * 100), 1) if total > 0 else 0

    # FU bulan ini
    c.execute("""
        SELECT COUNT(*) as cnt FROM follow_up_log
        WHERE sales_owner=%s AND EXTRACT(YEAR FROM tgl_fu)=%s AND EXTRACT(MONTH FROM tgl_fu)=%s
    """, (sales_name, cur_year, cur_month))
    total_fu_month = (c.fetchone() or {}).get("cnt", 0)

    # Monthly activity (12 bulan terakhir)
    c.execute("""
        SELECT TO_CHAR(tgl_fu,'YYYY-MM') as month,
               COUNT(*) as total_fu
        FROM follow_up_log
        WHERE sales_owner=%s
          AND tgl_fu >= (CURRENT_DATE - INTERVAL '11 months')::date
        GROUP BY 1 ORDER BY 1
    """, (sales_name,))
    monthly_activity = [dict(r) for r in c.fetchall()]

    # By stage
    c.execute("""
        SELECT stage, COUNT(*) as count, COALESCE(SUM(propose_value),0) as value
        FROM leads WHERE sales_owner=%s
        GROUP BY stage ORDER BY count DESC
    """, (sales_name,))
    by_stage = [dict(r) for r in c.fetchall()]

    # Pipeline value per bulan (deal yang closed tahun ini)
    c.execute("""
        SELECT TO_CHAR(tgl_masuk,'YYYY-MM') as month,
               COALESCE(SUM(deal_value),0) as value
        FROM leads
        WHERE sales_owner=%s AND stage='Won'
          AND EXTRACT(YEAR FROM tgl_masuk)=%s
        GROUP BY 1 ORDER BY 1
    """, (sales_name, cur_year))
    pipeline_value = [dict(r) for r in c.fetchall()]

    conn.close()
    return {
        "summary": {
            "total_leads": total,
            "won": won,
            "lost": lost,
            "win_rate": win_rate,
            "total_deal_won": total_deal_won,
            "total_fu_month": total_fu_month,
        },
        "monthly_activity": monthly_activity,
        "by_stage": by_stage,
        "pipeline_value": pipeline_value,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# HEALTH CHECK
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/health", tags=["System"])
def health():
    try:
        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT 1")
        conn.close()
        db_ok = True
    except Exception as e:
        db_ok = False
    return {
        "status": "ok" if db_ok else "degraded",
        "db": "connected" if db_ok else "error",
        "version": "2.0.0",
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/notifications", tags=["Notifications"])
def notif_list(
    unread_only: bool = False,
    limit: int = 20,
    user: dict = Depends(get_current_user),
):
    conn = get_conn(); c = conn.cursor()
    q = "SELECT * FROM notifications WHERE user_id=%s"
    params = [user["id"]]
    if unread_only:
        q += " AND read_at IS NULL"
    q += " ORDER BY created_at DESC LIMIT %s"
    params.append(min(limit, 50))  # cap maksimum 50
    c.execute(q, params)
    notifs = [_norm(dict(r)) for r in c.fetchall()]

    # Gunakan COUNT dari hasil fetch jika unread_only untuk menghindari query kedua
    if unread_only:
        unread_count = len(notifs)
    else:
        c.execute("SELECT COUNT(*) cnt FROM notifications WHERE user_id=%s AND read_at IS NULL", (user["id"],))
        unread_count = int(c.fetchone()["cnt"])
    conn.close()
    return {"notifications": notifs, "unread_count": unread_count}


@app.put("/api/v1/notifications/{nid}/read", tags=["Notifications"])
def notif_mark_read(nid: int, user: dict = Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute(
        "UPDATE notifications SET read_at=NOW() WHERE id=%s AND user_id=%s AND read_at IS NULL",
        (nid, user["id"])
    )
    conn.commit(); conn.close()
    return {"message": "ok"}


@app.put("/api/v1/notifications/read-all", tags=["Notifications"])
def notif_read_all(user: dict = Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute(
        "UPDATE notifications SET read_at=NOW() WHERE user_id=%s AND read_at IS NULL",
        (user["id"],)
    )
    conn.commit(); conn.close()
    return {"message": "ok"}


@app.delete("/api/v1/notifications/{nid}", tags=["Notifications"])
def notif_delete(nid: int, user: dict = Depends(get_current_user)):
    conn = get_conn(); c = conn.cursor()
    c.execute("DELETE FROM notifications WHERE id=%s AND user_id=%s", (nid, user["id"]))
    conn.commit(); conn.close()
    return {"message": "ok"}


@app.post("/api/v1/notifications/generate", tags=["Notifications"])
def notif_generate(user: dict = Depends(get_current_user)):
    """Generate notifikasi otomatis: stale leads, overdue FU, closing soon."""
    _generate_notifications()
    return {"message": "Notifikasi berhasil di-generate."}


def _generate_notifications():
    """
    Generate notifikasi sistem untuk semua user yang relevan.
    Dipanggil saat server startup dan bisa di-trigger manual.
    Mencegah duplikat dengan cek lead_id + type + hari ini.
    """
    conn = get_conn(); c = conn.cursor()
    today = date.today()

    # Ambil semua user aktif
    c.execute("SELECT id FROM users WHERE is_active=1")
    user_ids = [r["id"] for r in c.fetchall()]

    # Ambil semua leads aktif
    c.execute("""
        SELECT l.*, u.id as uid
        FROM leads l
        LEFT JOIN users u ON l.sales_owner = u.nama
        WHERE l.stage NOT IN ('Won','Lost','On Hold')
    """)
    leads = [dict(r) for r in c.fetchall()]

    inserted = 0
    for lead in leads:
        lid        = lead["lead_id"]
        nama       = lead["nama_company"]
        prioritas  = (lead["prioritas"] or "").lower()
        stage      = lead["stage"] or ""
        owner_uid  = lead["uid"]
        last_fu    = lead.get("last_fu_date")
        exp_close  = lead.get("exp_close_date")

        # Hitung days since last FU
        ref_date = last_fu if last_fu else lead.get("tgl_masuk")
        if ref_date:
            if isinstance(ref_date, str):
                from datetime import datetime as _dt
                try: ref_date = _dt.fromisoformat(ref_date).date()
                except: ref_date = None
            if ref_date:
                days_since = (today - ref_date).days
            else:
                days_since = None
        else:
            days_since = None

        # Target user: owner lead, atau semua admin jika tidak ada owner
        targets = [owner_uid] if owner_uid else user_ids

        def already_notified(uid, ntype, lead_id):
            c.execute("""SELECT id FROM notifications
                WHERE user_id=%s AND type=%s AND lead_id=%s
                AND DATE(created_at)=%s""", (uid, ntype, lead_id, today))
            return c.fetchone() is not None

        def push(uid, ntype, title, body):
            nonlocal inserted
            if already_notified(uid, ntype, lid): return
            c.execute("""INSERT INTO notifications (user_id, type, title, body, lead_id, created_at, updated_at)
                VALUES (%s,%s,%s,%s,%s,NOW(),NOW())""", (uid, ntype, title, body, lid))
            inserted += 1

        for uid in targets:
            if not uid: continue

            # 1. Stale lead
            if days_since is not None:
                if prioritas == "hot" and days_since > 7:
                    push(uid, "stale",
                         f"🔴 Hot Lead Terbengkalai: {nama}",
                         f"{nama} ({stage}) sudah {days_since} hari tanpa follow-up. Hot lead butuh perhatian segera!")
                elif prioritas == "warm" and days_since > 14:
                    push(uid, "stale",
                         f"🟡 Warm Lead Perlu FU: {nama}",
                         f"{nama} ({stage}) sudah {days_since} hari tanpa follow-up.")
                elif days_since > 30:
                    push(uid, "stale",
                         f"⚪ Lead Tidak Aktif: {nama}",
                         f"{nama} sudah {days_since} hari tanpa aktivitas.")

            # 2. Overdue next FU
            next_fu = lead.get("next_fu_date")
            if next_fu:
                if isinstance(next_fu, str):
                    try:
                        from datetime import datetime as _dt
                        next_fu = _dt.fromisoformat(next_fu).date()
                    except: next_fu = None
                if next_fu and next_fu < today:
                    overdue_days = (today - next_fu).days
                    push(uid, "overdue",
                         f"⏰ Follow-Up Terlewat: {nama}",
                         f"Jadwal FU {nama} sudah {overdue_days} hari terlewat (target: {next_fu.isoformat()}).")

            # 3. Closing soon (dalam 14 hari)
            if exp_close:
                if isinstance(exp_close, str):
                    try:
                        from datetime import datetime as _dt
                        exp_close = _dt.fromisoformat(exp_close).date()
                    except: exp_close = None
                if exp_close:
                    days_to_close = (exp_close - today).days
                    if 0 <= days_to_close <= 14:
                        push(uid, "closing",
                             f"🎯 Mendekati Closing: {nama}",
                             f"Expected close date {nama} tinggal {days_to_close} hari lagi ({exp_close.isoformat()}). Pastikan pipeline siap!")

            # 4. FU jatuh tempo hari ini
            next_fu = lead.get("next_fu_date")
            if next_fu:
                if isinstance(next_fu, str):
                    try:
                        from datetime import datetime as _dt
                        next_fu = _dt.fromisoformat(next_fu).date()
                    except: next_fu = None
                if next_fu and next_fu == today:
                    push(uid, "reminder",
                         f"📅 Follow-Up Hari Ini: {nama}",
                         f"Jadwal follow-up {nama} ({stage}) adalah hari ini. Jangan terlewat!")

    # 5. Revenue project at risk / critical — notif ke semua admin & manager
    c.execute("""
        SELECT p.project_id, p.nama_project, p.pic, p.risk_level, p.status,
               p.revenue_target, p.actual_revenue,
               u.id as uid
        FROM revenue_projects p
        LEFT JOIN users u ON p.pic = u.nama
        WHERE p.is_active = 1
          AND p.tahun = %s
          AND p.status IN ('Critical', 'At Risk')
    """, (today.year,))
    rev_projects = [dict(r) for r in c.fetchall()]

    # Ambil semua admin & manager untuk notif proyek
    c.execute("SELECT id FROM users WHERE is_active=1 AND role_id IN (1,2)")
    admin_mgr_ids = [r["id"] for r in c.fetchall()]

    for rp in rev_projects:
        pid    = rp["project_id"]
        pnama  = rp["nama_project"]
        status = rp["status"]
        pic_uid = rp["uid"]
        gap_pct = 0
        if rp["revenue_target"] and rp["revenue_target"] > 0:
            gap_pct = round((1 - (rp["actual_revenue"] or 0) / rp["revenue_target"]) * 100, 1)

        targets_rev = list(set(admin_mgr_ids + ([pic_uid] if pic_uid else [])))

        def already_notified_proj(uid, ntype, proj_id):
            c.execute("""SELECT id FROM notifications
                WHERE user_id=%s AND type=%s AND body LIKE %s
                AND DATE(created_at)=%s""", (uid, ntype, f"%{proj_id}%", today))
            return c.fetchone() is not None

        for uid in targets_rev:
            if not uid: continue
            if already_notified_proj(uid, "warning", pid): continue
            if status == "Critical":
                c.execute("""INSERT INTO notifications (user_id, type, title, body, created_at, updated_at)
                    VALUES (%s,'warning',%s,%s,NOW(),NOW())""",
                    (uid,
                     f"🔴 Proyek Critical: {pnama}",
                     f"Proyek {pid} — {pnama} berstatus Critical dengan gap realisasi {gap_pct}%. Perlu tindakan segera."))
            else:
                c.execute("""INSERT INTO notifications (user_id, type, title, body, created_at, updated_at)
                    VALUES (%s,'warning',%s,%s,NOW(),NOW())""",
                    (uid,
                     f"🟡 Proyek At Risk: {pnama}",
                     f"Proyek {pid} — {pnama} berstatus At Risk dengan gap realisasi {gap_pct}%."))
            inserted += 1

    # 6. Target sales mendekati akhir bulan dengan gap besar
    days_left_month = (date(today.year, today.month % 12 + 1, 1) - today).days if today.month < 12 \
                      else (date(today.year + 1, 1, 1) - today).days
    if days_left_month <= 7:
        c.execute("""
            SELECT st.sales_nama, st.target_deal,
                   COALESCE(SUM(l.deal_value), 0) AS actual,
                   u.id AS uid
            FROM sales_targets st
            LEFT JOIN leads l ON l.sales_owner = st.sales_nama
                AND EXTRACT(YEAR  FROM l.deal_date) = %s
                AND EXTRACT(MONTH FROM l.deal_date) = %s
                AND l.stage = 'Won'
            LEFT JOIN users u ON u.nama = st.sales_nama
            WHERE st.tahun = %s AND st.bulan = %s AND st.target_deal > 0
            GROUP BY st.sales_nama, st.target_deal, u.id
        """, (today.year, today.month, today.year, today.month))
        target_rows = [dict(r) for r in c.fetchall()]

        for tr in target_rows:
            uid = tr["uid"]
            if not uid: continue
            gap = tr["target_deal"] - tr["actual"]
            if gap <= 0: continue
            gap_pct_t = round(gap / tr["target_deal"] * 100, 1)
            if gap_pct_t < 30: continue  # masih wajar
            c.execute("""SELECT id FROM notifications
                WHERE user_id=%s AND type='warning'
                AND title LIKE '%Target Bulan Ini%'
                AND DATE(created_at)=%s""", (uid, today))
            if c.fetchone(): continue
            c.execute("""INSERT INTO notifications (user_id, type, title, body, created_at, updated_at)
                VALUES (%s,'warning',%s,%s,NOW(),NOW())""",
                (uid,
                 f"⚠️ Target Bulan Ini Tertinggal",
                 f"Sisa {days_left_month} hari, target bulan ini masih kurang {gap_pct_t}% (Rp {gap:,.0f}). Kejar closing sekarang!"))
            inserted += 1

    conn.commit()
    conn.close()
    return inserted


# Jalankan generate notifikasi saat startup
try:
    _generate_notifications()
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# APP SETTINGS (target lock, dll)
# ═══════════════════════════════════════════════════════════════════════════════

def _ensure_app_settings():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            key        TEXT PRIMARY KEY,
            value      TEXT,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS lead_history (
            id          SERIAL PRIMARY KEY,
            lead_id     TEXT NOT NULL,
            field_name  TEXT NOT NULL,
            old_value   TEXT,
            new_value   TEXT,
            changed_by  TEXT,
            changed_at  TIMESTAMP DEFAULT NOW()
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_lead_history_lead ON lead_history(lead_id)")
    c.execute("""
        CREATE TABLE IF NOT EXISTS fu_templates (
            id          SERIAL PRIMARY KEY,
            nama        TEXT NOT NULL,
            hasil_fu    TEXT,
            metode_fu   TEXT,
            catatan     TEXT NOT NULL,
            created_by  TEXT,
            created_at  TIMESTAMP DEFAULT NOW()
        )
    """)
    conn.commit(); conn.close()

try:
    _ensure_app_settings()
except Exception:
    pass


class AppSettingUpdate(BaseModel):
    value: str


@app.get("/api/v1/app-settings", tags=["Settings"])
def get_app_settings(user: dict = Depends(get_current_user)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT key, value FROM app_settings")
    rows = {r["key"]: r["value"] for r in c.fetchall()}
    conn.close()
    return rows


@app.put("/api/v1/app-settings/{key}", tags=["Settings"])
def set_app_setting(key: str, payload: AppSettingUpdate, user: dict = Depends(require_admin)):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO app_settings (key, value, updated_at, updated_by)
        VALUES (%s, %s, NOW(), %s)
        ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value, updated_at=NOW(), updated_by=EXCLUDED.updated_by
    """, (key, payload.value, user["email"]))
    conn.commit(); conn.close()
    return {"key": key, "value": payload.value}


class TileUsagePayload(BaseModel):
    count: int

@app.post("/api/v1/tile-usage", tags=["Settings"])
def track_tile_usage(payload: TileUsagePayload, user: dict = Depends(get_current_user)):
    from datetime import date as _date
    key = f"tile_usage_{_date.today().strftime('%Y-%m')}"
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        INSERT INTO app_settings (key, value, updated_at, updated_by)
        VALUES (%s, %s, NOW(), %s)
        ON CONFLICT (key) DO UPDATE
        SET value     = (COALESCE(NULLIF(app_settings.value,'')::int, 0) + %s)::text,
            updated_at = NOW()
    """, (key, str(payload.count), user["email"], payload.count))
    conn.commit(); conn.close()
    return {"ok": True}

@app.get("/api/v1/tile-usage", tags=["Settings"])
def get_tile_usage(user: dict = Depends(require_admin)):
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        SELECT key, value FROM app_settings
        WHERE key LIKE 'tile_usage_%' ORDER BY key DESC LIMIT 6
    """)
    rows = {r["key"].replace("tile_usage_", ""): int(r["value"] or 0) for r in c.fetchall()}
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLEANSING (Admin only)
# ═══════════════════════════════════════════════════════════════════════════════

class ResetRequest(BaseModel):
    confirm_text: str          # harus "RESET"
    password: str              # password admin yang sedang login
    delete_config: list = []   # list nama tabel config yang ikut dihapus
    reset_kpi_actuals: bool = False     # zero-kan q*_actual di kpi_prospecting
    reset_budget_actuals: bool = False  # zero-kan actual_amount di budget_items

@app.get("/api/v1/admin/cleansing/preview", tags=["Admin"])
def cleansing_preview(user: dict = Depends(require_admin)):
    """Tampilkan jumlah record per tabel yang akan dihapus."""
    conn = get_conn(); c = conn.cursor()

    BUSINESS_TABLES = [
        "leads", "follow_up_log", "lead_history", "contacts",
        "visit_logs", "daily_reports", "location_logs", "notifications",
        "revenue_projects", "revenue_monthly", "invoices", "annual_targets", "win_loss",
        "entertainment_claims", "entertainment_approvals", "password_reset_otps",
    ]
    CONFIG_TABLES = [
        "app_settings", "location_settings", "entertainment_settings",
        "fu_templates", "organizations", "products", "sales_targets",
        "kpi_prospecting", "budget_items", "annual_target_orgs",
    ]

    result = {"business": [], "config": [], "users": {}}

    for tbl in BUSINESS_TABLES:
        c.execute(f"SELECT COUNT(*) AS n FROM {tbl}")
        result["business"].append({"table": tbl, "count": int(c.fetchone()["n"])})

    for tbl in CONFIG_TABLES:
        c.execute(f"SELECT COUNT(*) AS n FROM {tbl}")
        result["config"].append({"table": tbl, "count": int(c.fetchone()["n"])})

    c.execute("SELECT COUNT(*) AS total FROM users")
    result["users"]["total"] = int(c.fetchone()["total"])
    c.execute("SELECT COUNT(*) AS admin FROM users WHERE role_id = 1")
    result["users"]["admin"] = int(c.fetchone()["admin"])
    result["users"]["will_delete"] = result["users"]["total"] - result["users"]["admin"]

    c.execute("""SELECT COUNT(*) AS n FROM kpi_prospecting
                 WHERE COALESCE(q1_actual,0)+COALESCE(q2_actual,0)+
                       COALESCE(q3_actual,0)+COALESCE(q4_actual,0) > 0""")
    result["kpi_actuals_count"] = int(c.fetchone()["n"])

    c.execute("SELECT COUNT(*) AS n FROM budget_items WHERE COALESCE(actual_amount,0) > 0 AND deleted_at IS NULL")
    result["budget_actuals_count"] = int(c.fetchone()["n"])

    conn.close()
    return result


@app.post("/api/v1/admin/cleansing/reset", tags=["Admin"])
def cleansing_reset(req: ResetRequest, user: dict = Depends(require_admin)):
    """Reset data bisnis. Preserve: roles, role_menus, admin users."""
    if req.confirm_text != "RESET":
        raise HTTPException(status_code=400, detail="Konfirmasi tidak valid")

    # Verifikasi password admin
    conn = get_conn(); c = conn.cursor()
    c.execute("SELECT password FROM users WHERE id = %s", (user["id"],))
    row = c.fetchone()
    if not row or row["password"] != hash_pw(req.password):
        conn.close()
        raise HTTPException(status_code=403, detail="Password salah")

    BUSINESS_TABLES = [
        "entertainment_approvals", "entertainment_claims",
        "win_loss", "invoices", "revenue_monthly", "revenue_projects",
        "annual_targets",
        "location_logs", "visit_logs", "daily_reports",
        "notifications", "password_reset_otps",
        "lead_history", "follow_up_log", "contacts", "leads",
    ]
    CONFIG_TABLES = [
        "app_settings", "location_settings", "entertainment_settings",
        "fu_templates", "organizations", "products", "sales_targets",
        "kpi_prospecting", "budget_items", "annual_target_orgs",
    ]

    deleted_counts = {}
    try:
        # Hapus tabel bisnis
        for tbl in BUSINESS_TABLES:
            c.execute(f"SELECT COUNT(*) AS n FROM {tbl}")
            deleted_counts[tbl] = int(c.fetchone()["n"])
            c.execute(f"TRUNCATE TABLE {tbl} RESTART IDENTITY CASCADE")

        # Hapus tabel config yang dipilih (whitelist dari CONFIG_TABLES)
        for tbl in req.delete_config:
            if tbl in CONFIG_TABLES:
                c.execute(f"SELECT COUNT(*) AS n FROM {tbl}")
                deleted_counts[tbl] = int(c.fetchone()["n"])
                c.execute(f"TRUNCATE TABLE {tbl} RESTART IDENTITY CASCADE")

        # Hapus non-admin users jika dipilih
        if "users (non-admin)" in req.delete_config:
            c.execute("SELECT COUNT(*) AS n FROM users WHERE role_id != 1")
            deleted_counts["users (non-admin)"] = int(c.fetchone()["n"])
            c.execute("DELETE FROM users WHERE role_id != 1")

        # Zero-kan aktual KPI jika diminta
        if req.reset_kpi_actuals:
            c.execute("UPDATE kpi_prospecting SET q1_actual=0, q2_actual=0, q3_actual=0, q4_actual=0")
            deleted_counts["kpi_prospecting (actuals zeroed)"] = c.rowcount

        # Zero-kan actual budget jika diminta
        if req.reset_budget_actuals:
            c.execute("UPDATE budget_items SET actual_amount=0 WHERE actual_amount IS NOT NULL AND actual_amount > 0")
            deleted_counts["budget_items (actuals zeroed)"] = c.rowcount

        # Catat ke audit log
        import json as _json
        kept_config = [t for t in CONFIG_TABLES if t not in req.delete_config]
        c.execute("""
            INSERT INTO cleansing_logs (executed_by, keep_config, deleted)
            VALUES (%s, %s, %s)
        """, (user["email"], bool(kept_config), _json.dumps(deleted_counts)))

        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Reset gagal: {str(e)}")

    conn.close()
    return {
        "ok": True,
        "message": "Data berhasil direset",
        "deleted": deleted_counts,
        "kept": ["roles", "role_menus"] + kept_config + ["users (admin)"],
    }


@app.get("/api/v1/admin/cleansing/logs", tags=["Admin"])
def cleansing_logs(user: dict = Depends(require_admin)):
    """Riwayat eksekusi data cleansing."""
    conn = get_conn(); c = conn.cursor()
    c.execute("""
        SELECT id, executed_by, executed_at, keep_config, deleted
        FROM cleansing_logs
        ORDER BY executed_at DESC
        LIMIT 50
    """)
    rows = []
    for r in c.fetchall():
        rows.append({
            "id": r["id"],
            "executed_by": r["executed_by"],
            "executed_at": r["executed_at"].isoformat() if r["executed_at"] else None,
            "keep_config": r["keep_config"],
            "deleted": r["deleted"],
            "total_deleted": sum(v for v in (r["deleted"] or {}).values() if isinstance(v, int)),
        })
    conn.close()
    return {"logs": rows}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("api_app:app", host="0.0.0.0", port=8001, reload=True)
