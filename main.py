from fastapi import FastAPI, Request, Form, HTTPException, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import uvicorn
import os
from datetime import datetime, date
from database_pg import (get_conn, init_db, next_lead_id, next_fu_id,
                         hash_pw, ALL_MENUS, get_user_menus)


def _norm(d: dict) -> dict:
    """Konversi datetime.date/datetime ke ISO string agar kompatibel dengan Jinja2 template."""
    from datetime import date as _date, datetime as _dt
    from decimal import Decimal
    out = {}
    for k, v in d.items():
        if isinstance(v, (_date, _dt)):
            out[k] = v.isoformat()
        elif isinstance(v, Decimal):
            out[k] = float(v)
        else:
            out[k] = v
    return out


MONTHS_ID = ['Januari','Februari','Maret','April','Mei','Juni',
             'Juli','Agustus','September','Oktober','November','Desember']
MONTHS_EN = ['January','February','March','April','May','June',
             'July','August','September','October','November','December']

# ── Threshold stale lead (hari) ───────────────────────────────────────────────
STALE_HOT   = 7    # Hot lead tanpa FU > 7 hari  → URGENT
STALE_WARM  = 14   # Warm lead tanpa FU > 14 hari → WARNING
STALE_ANY   = 30   # Lead apapun tanpa FU > 30 hari → STALE

def compute_stale_flag(prioritas: str, days: float) -> str:
    """Hitung flag stale berdasarkan prioritas & hari tanpa FU."""
    if days is None:
        return "OK"
    if (prioritas or '').lower() == 'hot'  and days > STALE_HOT:
        return "URGENT"
    if (prioritas or '').lower() == 'warm' and days > STALE_WARM:
        return "WARNING"
    if days > STALE_ANY:
        return "STALE"
    return "OK"

def auto_status_risk(revenue_target, actual_revenue,
                     type_: str, tahun: int, target_month: str = None):
    revenue_target = float(revenue_target or 0)
    actual_revenue = float(actual_revenue or 0)
    """
    Hitung Status dan Risk Level secara otomatis.

    STATUS berdasarkan achievement% vs pace (bulan berjalan):
      - On Track  : ach% ≥ pace × 0.9  (dalam 10% dari jadwal)
      - At Risk   : ach% ≥ pace × 0.5  (setengah dari pace)
      - Critical  : ach% < pace × 0.5

    RISK LEVEL berdasarkan Type + pace gap:
      - Termin    : selalu HIGH sampai invoice pertama masuk
      - Tahunan   : HIGH jika belum bayar di bulan target
      - Bulanan   : MEDIUM jika ach < 70%, LOW jika ach ≥ 70%
      - One Time  : HIGH jika belum ada actual
    """
    if not revenue_target or revenue_target <= 0:
        return "Critical", "LOW"

    ach = actual_revenue / revenue_target  # 0.0 – 1.0+

    today = date.today()
    # Proporsi tahun yang sudah berlalu
    if today.year == tahun:
        months_elapsed = today.month
    elif today.year > tahun:
        months_elapsed = 12  # tahun sudah lewat
    else:
        months_elapsed = 0   # tahun belum mulai
    pace = months_elapsed / 12  # 0.0 – 1.0

    # Status
    if pace == 0:
        status = "Critical" if ach == 0 else "On Track"
    elif ach >= max(pace * 0.9, 0.95):
        status = "On Track"
    elif ach >= pace * 0.5:
        status = "At Risk"
    else:
        status = "Critical"

    # Jika sudah 100% achieved → paksa On Track
    if ach >= 1.0:
        status = "On Track"

    # Risk Level
    type_lower = (type_ or '').lower()
    gap = pace - ach  # seberapa jauh tertinggal dari pace

    if 'termin' in type_lower:
        risk = "HIGH" if actual_revenue == 0 else ("HIGH" if gap > 0.3 else "MEDIUM")
    elif 'tahunan' in type_lower:
        risk = "HIGH" if actual_revenue == 0 else ("MEDIUM" if gap > 0.2 else "LOW")
    elif 'bulanan' in type_lower:
        risk = "LOW" if ach >= 0.7 else ("MEDIUM" if ach >= 0.4 else "HIGH")
    elif 'one time' in type_lower:
        risk = "HIGH" if actual_revenue == 0 else "LOW"
    else:
        risk = "HIGH" if ach < 0.3 else ("MEDIUM" if ach < 0.7 else "LOW")

    return status, risk

def sync_project_status(project_id: str):
    """Sync actual_revenue dari invoice dan update status/risk otomatis."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM revenue_projects WHERE project_id=%s", (project_id,))
    p = c.fetchone()
    if not p:
        conn.close()
        return
    p = dict(p)

    # Actual sudah di-sync oleh trigger, baca nilai terkini
    c.execute("SELECT COALESCE(SUM(paid_amount),0) as total FROM invoices WHERE project_id=%s", (project_id,))
    actual = c.fetchone()['total']

    status, risk = auto_status_risk(
        float(p['revenue_target'] or 0), float(actual or 0), p['type'], int(p.get('tahun', 2026) or 2026), p.get('target_month')
    )
    c.execute("""UPDATE revenue_projects SET actual_revenue=%s, status=%s, risk_level=%s,
        updated_at=NOW() WHERE project_id=%s""",
        (actual, status, risk, project_id))
    conn.commit()
    conn.close()

def sync_all_status():
    """Sync status & risk untuk semua proyek aktif."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT project_id, revenue_target, actual_revenue, type, tahun, target_month FROM revenue_projects WHERE is_active=1")
    projects = [_norm(dict(r)) for r in c.fetchall()]
    for p in projects:
        c.execute("SELECT COALESCE(SUM(paid_amount),0) as total FROM invoices WHERE project_id=%s", (p['project_id'],))
        actual = c.fetchone()['total']
        status, risk = auto_status_risk(float(p['revenue_target'] or 0), float(actual or 0), p['type'], p.get('tahun', 2026))
        c.execute("UPDATE revenue_projects SET actual_revenue=%s, status=%s, risk_level=%s WHERE project_id=%s",
                     (actual, status, risk, p['project_id']))
    conn.commit()
    conn.close()

def next_rev_id():
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT project_id FROM revenue_projects ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    if row:
        try:
            num = int(row['project_id'].replace('REV-','')) + 1
        except:
            num = 9000
    else:
        num = 1
    return f"REV-{num:04d}"

app = FastAPI(title="CRM Leads Tracker - DCSS")
app.add_middleware(SessionMiddleware, secret_key="crm-dcss-pkp-2026-secret", max_age=86400)

BASE_DIR = os.path.dirname(__file__)
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

init_db()

# ── AUTH HELPERS ──────────────────────────────────────────────────────────────

def get_current_user(request: Request):
    uid = request.session.get("user_id")
    if not uid:
        return None
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.*, r.nama as role_nama
                 FROM users u LEFT JOIN roles r ON u.role_id = r.id
                 WHERE u.id=%s AND u.is_active=1""", (uid,))
    row = c.fetchone()
    conn.close()
    return dict(row) if row else None

def login_required(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=307,
                            headers={"Location": f"/login?next={request.url.path}"})
    return user

def build_ctx(request: Request, user: dict, extra: dict = None) -> dict:
    """Build template context with user info and allowed menus."""
    allowed = get_user_menus(user["role_id"]) if user.get("role_id") else set()
    nav_menus = [m for m in ALL_MENUS if m["key"] in allowed]
    ctx = {"request": request, "current_user": user,
           "nav_menus": nav_menus, "allowed_menus": allowed}
    if extra:
        ctx.update(extra)
    return ctx

def can_access(user: dict, menu_key: str) -> bool:
    allowed = get_user_menus(user["role_id"]) if user.get("role_id") else set()
    return menu_key in allowed

def guard(user: dict, menu_key: str):
    if not can_access(user, menu_key):
        raise HTTPException(403, "Anda tidak memiliki akses ke halaman ini.")

# ── LOGIN / LOGOUT ─────────────────────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, next: str = "/"):
    if get_current_user(request):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "next": next, "error": None})

@app.post("/login")
async def login_post(request: Request,
                     email: str = Form(...), password: str = Form(...),
                     next: str = Form("/")):
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE email=%s AND is_active=1", (email.strip(),))
    user = c.fetchone()
    conn.close()
    if user and user["password"] == hash_pw(password):
        request.session["user_id"] = user["id"]
        return RedirectResponse(next or "/", status_code=303)
    return templates.TemplateResponse("login.html",
        {"request": request, "next": next,
         "error": "Email atau password salah."})

@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)

# ── MASTER ROLES ──────────────────────────────────────────────────────────────

@app.get("/roles", response_class=HTMLResponse)
async def roles_list(request: Request, user=Depends(login_required)):
    guard(user, "roles")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT r.*, COUNT(u.id) as user_count FROM roles r LEFT JOIN users u ON u.role_id=r.id GROUP BY r.id ORDER BY r.id")
    roles = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT role_id, menu_key FROM role_menus")
    role_menus_map = {}
    for row in c.fetchall():
        role_menus_map.setdefault(row["role_id"], set()).add(row["menu_key"])
    conn.close()
    return templates.TemplateResponse("roles.html", build_ctx(request, user, {
        "roles": roles, "role_menus_map": role_menus_map, "all_menus": ALL_MENUS
    }))

@app.post("/roles/new")
async def role_create(request: Request, user=Depends(login_required),
                      nama: str = Form(...), deskripsi: str = Form("")):
    guard(user, "roles")
    conn = get_conn()
    c = conn.cursor()
    c.execute("INSERT INTO roles (nama, deskripsi) VALUES (%s,%s)", (nama.strip(), deskripsi.strip()))
    conn.commit()
    conn.close()
    return RedirectResponse("/roles", status_code=303)

@app.post("/roles/{rid}/edit")
async def role_update(rid: int, request: Request, user=Depends(login_required),
                      nama: str = Form(...), deskripsi: str = Form("")):
    guard(user, "roles")
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE roles SET nama=%s, deskripsi=%s WHERE id=%s", (nama.strip(), deskripsi.strip(), rid))
    conn.commit()
    conn.close()
    return RedirectResponse("/roles", status_code=303)

@app.post("/roles/{rid}/menus")
async def role_save_menus(rid: int, request: Request, user=Depends(login_required)):
    guard(user, "roles")
    form = await request.form()
    selected = {v for k, v in form.multi_items() if k == "menus"}
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM role_menus WHERE role_id=%s", (rid,))
    for key in selected:
        c.execute("INSERT INTO role_menus VALUES (%s,%s) ON CONFLICT DO NOTHING", (rid, key))
    conn.commit()
    conn.close()
    return RedirectResponse("/roles", status_code=303)

@app.post("/roles/{rid}/delete")
async def role_delete(rid: int, request: Request, user=Depends(login_required)):
    guard(user, "roles")
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM roles WHERE id=%s", (rid,))
    conn.commit()
    conn.close()
    return RedirectResponse("/roles", status_code=303)

# ── MASTER USERS ──────────────────────────────────────────────────────────────

@app.get("/users", response_class=HTMLResponse)
async def users_list(request: Request, user=Depends(login_required)):
    guard(user, "users")
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT u.*, r.nama as role_nama FROM users u
                 LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.id""")
    users = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT * FROM roles ORDER BY nama")
    roles = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return templates.TemplateResponse("users.html", build_ctx(request, user, {
        "users": users, "roles": roles
    }))

@app.post("/users/new")
async def user_create(request: Request, user=Depends(login_required),
                      nama: str = Form(...), email: str = Form(...),
                      password: str = Form("pkp2026"), role_id: int = Form(...)):
    guard(user, "users")
    conn = get_conn()
    c = conn.cursor()
    c.execute("INSERT INTO users (nama,email,password,role_id) VALUES (%s,%s,%s,%s)",
                 (nama.strip(), email.strip(), hash_pw(password), role_id))
    conn.commit()
    conn.close()
    return RedirectResponse("/users", status_code=303)

@app.post("/users/{uid}/edit")
async def user_update(uid: int, request: Request, user=Depends(login_required),
                      nama: str = Form(...), email: str = Form(...),
                      role_id: int = Form(...), is_active: int = Form(1)):
    guard(user, "users")
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE users SET nama=%s,email=%s,role_id=%s,is_active=%s WHERE id=%s",
                 (nama.strip(), email.strip(), role_id, is_active, uid))
    conn.commit()
    conn.close()
    return RedirectResponse("/users", status_code=303)

@app.post("/users/{uid}/reset-password")
async def user_reset_pw(uid: int, request: Request, user=Depends(login_required),
                         new_password: str = Form(...)):
    guard(user, "users")
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE users SET password=%s WHERE id=%s", (hash_pw(new_password), uid))
    conn.commit()
    conn.close()
    return RedirectResponse("/users", status_code=303)

@app.post("/users/{uid}/delete")
async def user_delete(uid: int, request: Request, user=Depends(login_required)):
    guard(user, "users")
    if uid == user["id"]:
        raise HTTPException(400, "Tidak bisa menghapus akun sendiri.")
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=%s", (uid,))
    conn.commit()
    conn.close()
    return RedirectResponse("/users", status_code=303)

# ── MASTER SALES (legacy route kept, now auth-protected) ─────────────────────

def get_sales_list():
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT u.*, r.nama as role_nama FROM users u LEFT JOIN roles r ON u.role_id=r.id ORDER BY u.nama")
    rows = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return rows

@app.get("/sales", response_class=HTMLResponse)
async def sales_list(request: Request, user=Depends(login_required)):
    guard(user, "sales")
    return templates.TemplateResponse("sales.html", build_ctx(request, user, {
        "sales_list": get_sales_list()
    }))

@app.post("/sales/new")
async def sales_create(request: Request, user=Depends(login_required),
                       nama: str = Form(...), email: str = Form(""),
                       role: str = Form("sales"), password: str = Form("pkp2026")):
    guard(user, "sales")
    # redirect to users route for proper creation
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM roles WHERE LOWER(nama)=%s", (role.lower(),))
    row = c.fetchone()
    rid = row["id"] if row else None
    c.execute("INSERT INTO users (nama,email,password,role_id) VALUES (%s,%s,%s,%s)",
                 (nama.strip(), email.strip(), hash_pw(password), rid))
    conn.commit()
    conn.close()
    return RedirectResponse("/sales", status_code=303)

@app.post("/sales/{uid}/edit")
async def sales_update(uid: int, request: Request, user=Depends(login_required),
                       nama: str = Form(...), email: str = Form(""), role: str = Form("sales")):
    guard(user, "sales")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT id FROM roles WHERE LOWER(nama)=%s", (role.lower(),))
    row = c.fetchone()
    rid = row["id"] if row else None
    c.execute("UPDATE users SET nama=%s,email=%s,role_id=%s WHERE id=%s", (nama.strip(), email.strip(), rid, uid))
    conn.commit()
    conn.close()
    return RedirectResponse("/sales", status_code=303)

@app.post("/sales/{uid}/delete")
async def sales_delete(uid: int, request: Request, user=Depends(login_required)):
    guard(user, "sales")
    if uid == user["id"]:
        raise HTTPException(400, "Tidak bisa menghapus akun sendiri.")
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=%s", (uid,))
    conn.commit()
    conn.close()
    return RedirectResponse("/sales", status_code=303)

# ── DASHBOARD ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, user=Depends(login_required)):
    conn = get_conn()
    c = conn.cursor()

    c.execute("SELECT COUNT(*) as total FROM leads")
    total = c.fetchone()["total"]

    c.execute("SELECT stage, COUNT(*) as cnt, SUM(propose_value) as total_val, AVG(propose_value) as avg_val FROM leads GROUP BY stage")
    by_stage = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("SELECT segmen, COUNT(*) as leads, SUM(propose_value) as propose_val, SUM(deal_value) as deal_val, SUM(CASE WHEN stage='Won' THEN 1 ELSE 0 END) as won FROM leads GROUP BY segmen")
    by_segmen = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage='Won'")
    won = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage='Lost'")
    lost = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage='New'")
    new = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage='In Progress'")
    in_progress = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage IN ('Proposal Sent','Negotiation')")
    proposal = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage IN ('On Hold','Lost')")
    on_hold_lost = c.fetchone()["cnt"]

    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE next_fu_date <= CURRENT_DATE AND stage NOT IN ('Won','Lost')")
    overdue = c.fetchone()["cnt"]

    c.execute("SELECT SUM(deal_value) as total FROM leads WHERE stage='Won'")
    total_deal = c.fetchone()["total"] or 0

    c.execute("SELECT SUM(propose_value) as total FROM leads WHERE stage NOT IN ('Won','Lost')")
    total_pipeline = c.fetchone()["total"] or 0

    # ── Stale Alert: leads aktif yang sudah lama tanpa FU ──────────────────────
    c.execute("""
        SELECT *,
          COALESCE(
            CAST((CURRENT_DATE - last_fu_date)::integer AS INTEGER),
            CAST((CURRENT_DATE - tgl_masuk)::integer AS INTEGER),
            999
          ) as days_without_fu
        FROM leads
        WHERE stage NOT IN ('Won','Lost')
        ORDER BY prioritas DESC, days_without_fu DESC
    """)
    all_active = [_norm(dict(r)) for r in c.fetchall()]

    stale_urgent  = [l for l in all_active if compute_stale_flag(l['prioritas'], l['days_without_fu']) == 'URGENT']
    stale_warning = [l for l in all_active if compute_stale_flag(l['prioritas'], l['days_without_fu']) == 'WARNING']
    stale_stale   = [l for l in all_active if compute_stale_flag(l['prioritas'], l['days_without_fu']) == 'STALE']

    conn.close()
    guard(user, "dashboard")
    return templates.TemplateResponse("dashboard.html", build_ctx(request, user, {
        "total": total, "won": won, "lost": lost, "new": new,
        "in_progress": in_progress, "proposal": proposal, "on_hold_lost": on_hold_lost,
        "by_stage": by_stage, "by_segmen": by_segmen,
        "overdue": overdue, "total_deal": total_deal, "total_pipeline": total_pipeline,
        "stale_urgent": stale_urgent, "stale_warning": stale_warning, "stale_stale": stale_stale,
        "STALE_HOT": STALE_HOT, "STALE_WARM": STALE_WARM, "STALE_ANY": STALE_ANY,
    }))

# ── PIPELINE ──────────────────────────────────────────────────────────────────

@app.get("/pipeline", response_class=HTMLResponse)
async def pipeline_list(request: Request, user=Depends(login_required),
                        stage: str = "", segmen: str = "", search: str = "", sales: str = ""):
    guard(user, "pipeline")
    conn = get_conn()
    c = conn.cursor()
    query = """SELECT *,
        COALESCE(
          (CURRENT_DATE - last_fu_date)::integer,
          (CURRENT_DATE - tgl_masuk)::integer,
          999
        ) as days_without_fu
        FROM leads WHERE 1=1"""
    params = []
    if stage:
        query += " AND stage=%s"
        params.append(stage)
    if segmen:
        query += " AND segmen=%s"
        params.append(segmen)
    if search:
        query += " AND (nama_company LIKE %s OR product LIKE %s OR contact_person LIKE %s)"
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]
    if sales:
        query += " AND sales_owner=%s"
        params.append(sales)
    query += " ORDER BY id DESC"
    c.execute(query, params)
    raw_leads = [_norm(dict(r)) for r in c.fetchall()]
    # Tambahkan stale_flag ke setiap lead
    for l in raw_leads:
        l['stale_flag'] = compute_stale_flag(l['prioritas'], l['days_without_fu']) \
                          if l['stage'] not in ('Won', 'Lost') else 'OK'
    leads = raw_leads
    conn.close()
    return templates.TemplateResponse("pipeline.html", build_ctx(request, user, {
        "leads": leads, "filter_stage": stage, "filter_segmen": segmen,
        "search": search, "filter_sales": sales, "sales_list": get_sales_list(),
        "today_str": str(date.today()),
    }))

@app.get("/pipeline/new", response_class=HTMLResponse)
async def pipeline_new(request: Request, user=Depends(login_required)):
    guard(user, "pipeline")
    return templates.TemplateResponse("lead_form.html", build_ctx(request, user, {
        "lead": None, "sales_list": get_sales_list(), "next_id": next_lead_id()
    }))

@app.post("/pipeline/new")
async def pipeline_create(request: Request, user=Depends(login_required),
    nama_company: str = Form(...), product: str = Form(""),
    contact_person: str = Form(""), segmen: str = Form(""),
    sub_segmen: str = Form(""), source: str = Form(""),
    stage: str = Form("New"), prioritas: str = Form("Warm"),
    tgl_masuk: str = Form(""), propose_value: str = Form("0"),
    deal_value: str = Form("0"), probability: str = Form("0"),
    exp_close_date: str = Form(""), sales_owner: str = Form(""),
    next_fu_date: str = Form(""), remarks: str = Form("")):
    guard(user, "pipeline")
    conn = get_conn()
    c = conn.cursor()
    lid = next_lead_id()
    pv = float(propose_value or 0)
    dv = float(deal_value or 0)
    prob = float(probability or 0)
    wv = dv * (prob / 100) if dv else pv * (prob / 100)
    c.execute("""INSERT INTO leads
        (lead_id,nama_company,product,contact_person,segmen,sub_segmen,source,stage,prioritas,
         tgl_masuk,propose_value,deal_value,probability,exp_close_date,weighted_value,
         sales_owner,next_fu_date,remarks)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (lid, nama_company, product, contact_person, segmen, sub_segmen, source, stage, prioritas,
         tgl_masuk or str(date.today()), pv, dv, prob, exp_close_date, wv,
         sales_owner, next_fu_date, remarks))
    conn.commit()
    conn.close()
    return RedirectResponse("/pipeline", status_code=303)

@app.get("/pipeline/{lead_id}/edit", response_class=HTMLResponse)
async def pipeline_edit(request: Request, lead_id: str, user=Depends(login_required)):
    guard(user, "pipeline")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM leads WHERE lead_id=%s", (lead_id,))
    lead = _norm(dict(c.fetchone()))
    conn.close()
    return templates.TemplateResponse("lead_form.html", build_ctx(request, user, {
        "lead": lead, "sales_list": get_sales_list()
    }))

@app.post("/pipeline/{lead_id}/edit")
async def pipeline_update(lead_id: str, request: Request, user=Depends(login_required),
    nama_company: str = Form(...), product: str = Form(""),
    contact_person: str = Form(""), segmen: str = Form(""),
    sub_segmen: str = Form(""), source: str = Form(""),
    stage: str = Form("New"), prioritas: str = Form("Warm"),
    tgl_masuk: str = Form(""), propose_value: str = Form("0"),
    deal_value: str = Form("0"), probability: str = Form("0"),
    exp_close_date: str = Form(""), sales_owner: str = Form(""),
    next_fu_date: str = Form(""), remarks: str = Form("")):
    guard(user, "pipeline")
    pv = float(propose_value or 0)
    dv = float(deal_value or 0)
    prob = float(probability or 0)
    wv = dv * (prob / 100) if dv else pv * (prob / 100)
    conn = get_conn()
    c = conn.cursor()
    c.execute("""UPDATE leads SET
        nama_company=%s,product=%s,contact_person=%s,segmen=%s,sub_segmen=%s,source=%s,
        stage=%s,prioritas=%s,tgl_masuk=%s,propose_value=%s,deal_value=%s,probability=%s,
        exp_close_date=%s,weighted_value=%s,sales_owner=%s,next_fu_date=%s,remarks=%s,
        updated_at=NOW() WHERE lead_id=%s""",
        (nama_company, product, contact_person, segmen, sub_segmen, source,
         stage, prioritas, tgl_masuk, pv, dv, prob, exp_close_date, wv,
         sales_owner, next_fu_date, remarks, lead_id))
    conn.commit()
    conn.close()
    return RedirectResponse("/pipeline", status_code=303)

@app.post("/pipeline/{lead_id}/delete")
async def pipeline_delete(lead_id: str, request: Request, user=Depends(login_required)):
    guard(user, "pipeline")
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM leads WHERE lead_id=%s", (lead_id,))
    conn.commit()
    conn.close()
    return RedirectResponse("/pipeline", status_code=303)

@app.get("/pipeline/{lead_id}", response_class=HTMLResponse)
async def pipeline_detail(request: Request, lead_id: str, user=Depends(login_required)):
    guard(user, "pipeline")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM leads WHERE lead_id=%s", (lead_id,))
    row = c.fetchone()
    if not row:
        raise HTTPException(404)
    lead = _norm(dict(row))
    c.execute("SELECT * FROM follow_up_log WHERE lead_id=%s ORDER BY tgl_fu DESC", (lead_id,))
    logs = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT * FROM contacts WHERE lead_id=%s", (lead_id,))
    contacts = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return templates.TemplateResponse("lead_detail.html", build_ctx(request, user, {
        "lead": lead, "logs": logs, "contacts": contacts, "sales_list": get_sales_list()
    }))

# ── TODAY / ACTIVITY DASHBOARD ────────────────────────────────────────────────

@app.get("/today", response_class=HTMLResponse)
async def today_dashboard(request: Request, user=Depends(login_required)):
    guard(user, "today")
    conn = get_conn()
    c = conn.cursor()
    today_str = str(date.today())
    c.execute("""SELECT l.*, (CURRENT_DATE - l.next_fu_date)::integer as days_late
        FROM leads l WHERE l.next_fu_date < %s AND l.stage NOT IN ('Won','Lost')
        AND l.next_fu_date IS NOT NULL ORDER BY l.next_fu_date ASC""", (today_str,))
    overdue = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("""SELECT l.* FROM leads l WHERE l.next_fu_date = %s AND l.stage NOT IN ('Won','Lost')
        ORDER BY l.prioritas DESC""", (today_str,))
    due_today = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("""SELECT l.*, (l.next_fu_date - CURRENT_DATE)::integer as days_until
        FROM leads l WHERE l.next_fu_date > %s AND l.next_fu_date <= (CURRENT_DATE + INTERVAL '7 days')
        AND l.stage NOT IN ('Won','Lost') ORDER BY l.next_fu_date ASC""", (today_str,))
    upcoming = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("""SELECT * FROM leads WHERE next_fu_date IS NULL
        AND stage NOT IN ('Won','Lost') ORDER BY prioritas DESC, propose_value DESC""")
    unscheduled = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT COUNT(*) as cnt FROM leads WHERE stage NOT IN ('Won','Lost')")
    total_active = c.fetchone()["cnt"]
    c.execute("SELECT COUNT(*) as cnt FROM follow_up_log WHERE tgl_fu = %s", (today_str,))
    fu_done_today = c.fetchone()["cnt"]

    # ── Stale leads untuk today view ────────────────────────────────────────
    c.execute("""
        SELECT *,
          COALESCE(
            CAST((CURRENT_DATE - last_fu_date)::integer AS INTEGER),
            CAST((CURRENT_DATE - tgl_masuk)::integer AS INTEGER),
            999
          ) as days_without_fu
        FROM leads
        WHERE stage NOT IN ('Won','Lost')
          AND next_fu_date IS NULL
        ORDER BY
          CASE prioritas WHEN 'Hot' THEN 1 WHEN 'Warm' THEN 2 ELSE 3 END,
          days_without_fu DESC
    """)
    stale_leads_raw = [_norm(dict(r)) for r in c.fetchall()]
    for l in stale_leads_raw:
        l['stale_flag'] = compute_stale_flag(l['prioritas'], l['days_without_fu'])

    stale_urgent  = [l for l in stale_leads_raw if l['stale_flag'] == 'URGENT']
    stale_warning = [l for l in stale_leads_raw if l['stale_flag'] == 'WARNING']
    stale_stale   = [l for l in stale_leads_raw if l['stale_flag'] == 'STALE']

    # ── Laporan harian hari ini dari APEX Mobile App ─────────────────────────
    conn2 = get_conn()
    c2 = conn2.cursor()
    try:
        c2.execute("""
            SELECT dr.*, u.nama as sales_nama
            FROM daily_reports dr
            LEFT JOIN users u ON dr.user_id = u.id
            WHERE dr.report_date = %s
            ORDER BY dr.submitted_at DESC NULLS LAST
        """, (today_str,))
        daily_reports = [_norm(dict(r)) for r in c2.fetchall()]
    except Exception:
        daily_reports = []
    finally:
        conn2.close()

    conn.close()
    return templates.TemplateResponse("today.html", build_ctx(request, user, {
        "today_str": today_str, "overdue": overdue, "due_today": due_today,
        "upcoming": upcoming, "unscheduled": unscheduled,
        "total_active": total_active, "fu_done_today": fu_done_today,
        "stale_urgent": stale_urgent, "stale_warning": stale_warning,
        "stale_stale": stale_stale,
        "STALE_HOT": STALE_HOT, "STALE_WARM": STALE_WARM, "STALE_ANY": STALE_ANY,
        "daily_reports": daily_reports,
    }))

# ── FOLLOW-UP LOG ─────────────────────────────────────────────────────────────

@app.get("/followup", response_class=HTMLResponse)
async def followup_list(request: Request, user=Depends(login_required),
                        lead_id: str = "", search: str = ""):
    guard(user, "followup")
    conn = get_conn()
    c = conn.cursor()
    query = "SELECT * FROM follow_up_log WHERE 1=1"
    params = []
    if lead_id:
        query += " AND lead_id=%s"; params.append(lead_id)
    if search:
        query += " AND (nama_company LIKE %s OR catatan_fu LIKE %s OR kontak LIKE %s)"
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]
    c.execute(query + " ORDER BY tgl_fu DESC", params)
    logs = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT lead_id, nama_company FROM leads ORDER BY nama_company")
    leads = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return templates.TemplateResponse("followup.html", build_ctx(request, user, {
        "logs": logs, "leads": leads, "filter_lead": lead_id,
        "search": search, "sales_list": get_sales_list()
    }))

@app.post("/followup/new")
async def followup_create(request: Request, user=Depends(login_required),
    lead_id: str = Form(...), tgl_fu: str = Form(...),
    metode_fu: str = Form(""), kontak: str = Form(""),
    hasil_fu: str = Form(""), catatan_fu: str = Form(""),
    stage_saat_fu: str = Form(""), next_action: str = Form(""),
    tgl_fu_berikut: str = Form(""), sales_owner: str = Form(""),
    status: str = Form("Done")):
    guard(user, "followup")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT nama_company FROM leads WHERE lead_id=%s", (lead_id,))
    row = c.fetchone()
    nama = row["nama_company"] if row else ""
    fid = next_fu_id()
    c.execute("""INSERT INTO follow_up_log
        (fu_id,lead_id,tgl_fu,nama_company,sales_owner,metode_fu,kontak,hasil_fu,
         catatan_fu,stage_saat_fu,next_action,tgl_fu_berikut,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (fid, lead_id, tgl_fu, nama, sales_owner, metode_fu, kontak, hasil_fu,
         catatan_fu, stage_saat_fu, next_action, tgl_fu_berikut, status))
    c.execute("""UPDATE leads SET last_fu_date=%s, last_fu_notes=%s, fu_count=fu_count+1,
        next_fu_date=%s, updated_at=NOW() WHERE lead_id=%s""",
        (tgl_fu, catatan_fu, tgl_fu_berikut or None, lead_id))
    conn.commit()
    conn.close()
    return RedirectResponse("/followup", status_code=303)

# ── FU SCHEDULE ───────────────────────────────────────────────────────────────

@app.get("/schedule", response_class=HTMLResponse)
async def fu_schedule(request: Request, user=Depends(login_required)):
    guard(user, "schedule")
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT l.*,
        CASE WHEN l.next_fu_date < CURRENT_DATE THEN 'Overdue'
             WHEN l.next_fu_date = CURRENT_DATE THEN 'Today'
             ELSE 'Upcoming' END as fu_status
        FROM leads l WHERE l.stage NOT IN ('Won','Lost')
        AND l.next_fu_date IS NOT NULL
        ORDER BY l.next_fu_date ASC""")
    schedule = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return templates.TemplateResponse("schedule.html", build_ctx(request, user, {"schedule": schedule}))

# ── WIN-LOSS ──────────────────────────────────────────────────────────────────

@app.get("/winloss", response_class=HTMLResponse)
async def winloss(request: Request, user=Depends(login_required)):
    guard(user, "winloss")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM win_loss ORDER BY id DESC")
    records = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT COUNT(*) as cnt, SUM(deal_value) as total FROM win_loss WHERE hasil='Won'")
    won_stats = _norm(dict(c.fetchone()))
    c.execute("SELECT COUNT(*) as cnt FROM win_loss WHERE hasil='Lost'")
    lost_stats = _norm(dict(c.fetchone()))
    conn.close()
    return templates.TemplateResponse("winloss.html", build_ctx(request, user, {
        "records": records, "won_stats": won_stats, "lost_stats": lost_stats
    }))

# ── CONTACTS ──────────────────────────────────────────────────────────────────

@app.get("/contacts", response_class=HTMLResponse)
async def contacts_list(request: Request, user=Depends(login_required), search: str = ""):
    guard(user, "contacts")
    conn = get_conn()
    c = conn.cursor()
    if search:
        c.execute("SELECT * FROM contacts WHERE nama_contact LIKE %s OR nama_company LIKE %s OR no_hp LIKE %s",
                  (f"%{search}%", f"%{search}%", f"%{search}%"))
    else:
        c.execute("SELECT * FROM contacts ORDER BY nama_company")
    ct = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return templates.TemplateResponse("contacts.html", build_ctx(request, user, {
        "contacts": ct, "search": search
    }))

@app.post("/contacts/new")
async def contact_create(request: Request, user=Depends(login_required),
    lead_id: str = Form(...), nama_contact: str = Form(...),
    jabatan: str = Form(""), dept: str = Form(""), role: str = Form(""),
    no_hp: str = Form(""), email: str = Form(""), telepon: str = Form(""),
    linkedin: str = Form(""), preferensi_kontak: str = Form(""), catatan: str = Form("")):
    guard(user, "contacts")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT nama_company FROM leads WHERE lead_id=%s", (lead_id,))
    row = c.fetchone()
    nama_co = row["nama_company"] if row else ""
    c.execute("""INSERT INTO contacts
        (lead_id,nama_company,nama_contact,jabatan,dept,role,no_hp,email,telepon,linkedin,preferensi_kontak,catatan)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (lead_id, nama_co, nama_contact, jabatan, dept, role, no_hp, email, telepon, linkedin, preferensi_kontak, catatan))
    conn.commit()
    conn.close()
    return RedirectResponse("/contacts", status_code=303)

# ── EXPORT ────────────────────────────────────────────────────────────────────

@app.get("/export")
async def export_excel(request: Request, user=Depends(login_required)):
    guard(user, "export")
    from fastapi.responses import FileResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM leads ORDER BY lead_id")
    leads = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT * FROM follow_up_log ORDER BY fu_id")
    logs = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pipeline"
    header_fill = PatternFill("solid", start_color="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)

    headers = ["Lead ID","Nama Company","Product","Contact Person","Segmen","Sub-segmen",
               "Source","Stage","Prioritas","Tgl Masuk","Propose Value","Deal Value",
               "Probability (%)","Exp. Close Date","Weighted Value","Sales Owner",
               "Next FU Date","Last FU Date","Last FU Notes","FU Count","Remarks"]
    keys = ["lead_id","nama_company","product","contact_person","segmen","sub_segmen",
            "source","stage","prioritas","tgl_masuk","propose_value","deal_value",
            "probability","exp_close_date","weighted_value","sales_owner",
            "next_fu_date","last_fu_date","last_fu_notes","fu_count","remarks"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, lead in enumerate(leads, 2):
        for col, key in enumerate(keys, 1):
            ws.cell(row=row, column=col, value=lead.get(key))

    ws2 = wb.create_sheet("Follow-Up Log")
    h2 = ["FU-ID","Lead ID","Tgl FU","Nama Company","Sales Owner","Metode FU",
          "Kontak","Hasil FU","Catatan FU","Stage Saat FU","Next Action","Tgl FU Berikut","Status"]
    k2 = ["fu_id","lead_id","tgl_fu","nama_company","sales_owner","metode_fu",
          "kontak","hasil_fu","catatan_fu","stage_saat_fu","next_action","tgl_fu_berikut","status"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for row, log in enumerate(logs, 2):
        for col, key in enumerate(k2, 1):
            ws2.cell(row=row, column=col, value=log.get(key))

    out = os.path.join(BASE_DIR, "data", "export.xlsx")
    wb.save(out)
    return FileResponse(out, filename=f"CRM_Leads_Export_{date.today()}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── INSIGHTS DASHBOARD ───────────────────────────────────────────────────────

@app.get("/insights", response_class=HTMLResponse)
async def insights(request: Request, user=Depends(login_required)):
    guard(user, "insights")
    conn = get_conn()
    c = conn.cursor()

    c.execute("SELECT * FROM leads")
    all_leads = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()

    total = len(all_leads)
    active = [l for l in all_leads if l["stage"] not in ("Won", "Lost")]
    won_leads = [l for l in all_leads if l["stage"] == "Won"]
    lost_leads = [l for l in all_leads if l["stage"] == "Lost"]
    onhold_leads = [l for l in all_leads if l["stage"] == "On Hold"]

    total_propose = sum(l["propose_value"] or 0 for l in all_leads)
    total_deal = sum(l["deal_value"] or 0 for l in won_leads)
    onhold_val = sum(l["propose_value"] or 0 for l in onhold_leads)
    lost_val = sum(l["propose_value"] or 0 for l in lost_leads)
    inactive = len(onhold_leads) + len(lost_leads)
    inactive_pct = round(inactive / total * 100) if total else 0
    won_pct = round(len(won_leads) / total * 100) if total else 0

    # Stage breakdown for funnel
    from collections import defaultdict
    stage_order = ["New", "In Progress", "Demo Scheduled", "Proposal Sent", "Negotiation", "Won", "On Hold", "Lost"]
    stage_colors = {
        "New": "#3b82f6", "In Progress": "#eab308", "Demo Scheduled": "#a855f7",
        "Proposal Sent": "#f97316", "Negotiation": "#f59e0b",
        "Won": "#22c55e", "On Hold": "#9ca3af", "Lost": "#ef4444"
    }
    stage_counts = defaultdict(lambda: {"cnt": 0, "val": 0})
    for l in all_leads:
        stage_counts[l["stage"]]["cnt"] += 1
        stage_counts[l["stage"]]["val"] += l["propose_value"] or 0
    stage_data = []
    for s in stage_order:
        if stage_counts[s]["cnt"] > 0:
            stage_data.append({
                "stage": s,
                "cnt": stage_counts[s]["cnt"],
                "val": stage_counts[s]["val"],
                "pct": round(stage_counts[s]["cnt"] / total * 100),
                "color": stage_colors.get(s, "#9ca3af")
            })

    # Action list — prioritized
    action_list = []
    rank = 0

    def add_action(lead, issue_type, issue_label, action):
        nonlocal rank
        rank += 1
        action_list.append({
            "rank": rank,
            "lead_id": lead["lead_id"],
            "nama_company": lead["nama_company"],
            "stage": lead["stage"],
            "segmen": lead["segmen"],
            "propose_value": lead["propose_value"] or 0,
            "issue_type": issue_type,
            "issue_label": issue_label,
            "action": action
        })

    # Rank 1-4: On Hold Hot/Warm dengan nilai besar
    for l in sorted(onhold_leads, key=lambda x: (-(x["propose_value"] or 0))):
        if l["prioritas"] in ("Hot", "Warm") and (l["propose_value"] or 0) >= 500_000_000:
            add_action(l, "frozen", "🧊 Membeku di On Hold",
                       f"Hubungi kembali minggu ini. Tetapkan deadline: lanjut atau tutup sebagai Lost.")

    # Rank berikutnya: Negotiation tanpa nilai
    for l in [x for x in all_leads if x["stage"] == "Negotiation"]:
        add_action(l, "no_value", "⚠️ Nego tanpa nilai",
                   "Tentukan deal value dan terms. Segera kirim proposal final.")

    # Leads aktif tanpa next FU
    for l in [x for x in active if not x.get("next_fu_date") and x["stage"] not in ("Won","Lost","On Hold")]:
        add_action(l, "no_fu", "📵 Tanpa jadwal FU",
                   "Set next FU date dan catat rencana kontak berikutnya.")

    # New leads tanpa nilai
    for l in [x for x in all_leads if x["stage"] == "New" and not (x["propose_value"] or 0)]:
        if l["nama_company"] not in [a["nama_company"] for a in action_list[:-1]]:
            add_action(l, "duplicate" if l["nama_company"] == "FIFGROUP" else "no_value",
                       "🔁 Duplikat/Stagnan" if l["nama_company"] == "FIFGROUP" else "❓ Belum dikualifikasi",
                       "Verifikasi apakah lead valid. Jika duplikat, hapus atau merge. Jika valid, kualifikasi dan tentukan nilai.")

    # Scorecard
    no_fu_active = len([l for l in active if not l.get("next_fu_date")])
    no_owner = len([l for l in all_leads if not l.get("sales_owner")])
    win_rate = round(len(won_leads) / total * 100) if total else 0
    multifinance_conv = round(
        sum(l["deal_value"] or 0 for l in all_leads if l["segmen"] == "Multifinance") /
        max(sum(l["propose_value"] or 0 for l in all_leads if l["segmen"] == "Multifinance"), 1) * 100, 1
    )

    scorecard = [
        {
            "label": "Win Rate Keseluruhan",
            "value": f"{win_rate}%",
            "icon": "🔴" if win_rate < 15 else "🟡",
            "status": "Kritis — Target min. 20%" if win_rate < 15 else "Perlu ditingkatkan",
            "color": "text-red-400" if win_rate < 15 else "text-yellow-300"
        },
        {
            "label": "Leads Tanpa Jadwal FU",
            "value": f"{no_fu_active}",
            "icon": "🔴" if no_fu_active > 5 else "🟡",
            "status": f"{no_fu_active} dari {len(active)} leads aktif",
            "color": "text-red-400" if no_fu_active > 5 else "text-yellow-300"
        },
        {
            "label": "Konversi Multifinance",
            "value": f"{multifinance_conv}%",
            "icon": "🔴" if multifinance_conv < 5 else "🟡",
            "status": "ROI sangat rendah" if multifinance_conv < 5 else "Perlu ditingkatkan",
            "color": "text-red-400" if multifinance_conv < 5 else "text-yellow-300"
        },
        {
            "label": "Pipeline Tidak Aktif",
            "value": f"{inactive_pct}%",
            "icon": "🔴" if inactive_pct > 35 else "🟡",
            "status": f"{inactive} leads On Hold + Lost",
            "color": "text-red-400" if inactive_pct > 35 else "text-yellow-300"
        },
    ]

    return templates.TemplateResponse("insights.html", build_ctx(request, user, {
        "total": total, "total_propose": total_propose, "total_deal": total_deal,
        "onhold_val": onhold_val, "onhold_count": len(onhold_leads),
        "lost_val": lost_val, "lost_leads": lost_leads,
        "inactive": inactive, "inactive_pct": inactive_pct,
        "won_count": len(won_leads), "won_pct": won_pct,
        "stage_data": stage_data, "action_list": action_list[:12], "scorecard": scorecard,
    }))

# ── API JSON (untuk chart) ─────────────────────────────────────────────────────

@app.get("/api/stats")
async def api_stats():
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT stage, COUNT(*) as cnt, SUM(propose_value) as val FROM leads GROUP BY stage")
    by_stage = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT segmen, COUNT(*) as cnt FROM leads GROUP BY segmen")
    by_segmen = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()
    return {"by_stage": by_stage, "by_segmen": by_segmen}

# ══════════════════════════════════════════════════════════════════════════════
# REVENUE LOB MODULES
# ══════════════════════════════════════════════════════════════════════════════

def revenue_summary():
    """Shared stats for revenue dashboard."""
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT SUM(revenue_target) as t, SUM(actual_revenue) as a FROM revenue_projects WHERE is_active=1")
    row = _norm(dict(c.fetchone()))
    total_target = row['t'] or 0
    total_actual = row['a'] or 0
    ach_pct = round(float(total_actual or 0) / float(total_target or 1) * 100, 1) if total_target else 0

    c.execute("SELECT SUM(revenue_target) as t, SUM(actual_revenue) as a FROM revenue_projects WHERE kategori='Recurring' AND is_active=1")
    r = _norm(dict(c.fetchone())); rec_target = float(r['t'] or 0); rec_actual = float(r['a'] or 0)
    c.execute("SELECT SUM(revenue_target) as t, SUM(actual_revenue) as a FROM revenue_projects WHERE kategori='Project' AND is_active=1")
    r = _norm(dict(c.fetchone())); prj_target = r['t'] or 0; prj_actual = r['a'] or 0

    c.execute("SELECT status, COUNT(*) as cnt FROM revenue_projects WHERE is_active=1 GROUP BY status")
    by_status = {r['status']: r['cnt'] for r in c.fetchall()}

    c.execute("SELECT COUNT(*) as cnt FROM revenue_projects WHERE is_active=1")
    total_projects = c.fetchone()['cnt']
    conn.close()
    return {
        "total_target": float(total_target or 0), "total_actual": float(total_actual or 0), "ach_pct": ach_pct,
        "rec_target": rec_target, "rec_actual": rec_actual,
        "prj_target": prj_target, "prj_actual": prj_actual,
        "by_status": by_status, "total_projects": total_projects,
    }

# ── B0: REVENUE INSIGHTS ──────────────────────────────────────────────────────

@app.get("/revenue/insights", response_class=HTMLResponse)
async def revenue_insights(request: Request, user=Depends(login_required), tahun: int = 0):
    guard(user, "rev_dashboard")
    cur_year = tahun if tahun else date.today().year
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]

    # Summary
    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    s = _norm(dict(c.fetchone()))
    total_target = s['t'] or 0; total_actual = s['a'] or 0
    ach_pct = round(float(total_actual or 0) / float(total_target or 1) * 100, 1) if total_target else 0
    gap = total_target - total_actual

    # By kategori
    c.execute("SELECT kategori, SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY kategori", (cur_year,))
    by_kat = {r['kategori']: dict(r) for r in c.fetchall()}
    project_target = float(by_kat.get('Project', {}).get('t', 0) or 0)
    project_actual = float(by_kat.get('Project', {}).get('a', 0) or 0)
    recurring_target = float(by_kat.get('Recurring', {}).get('t', 0) or 0)
    recurring_actual = float(by_kat.get('Recurring', {}).get('a', 0) or 0)
    project_ach = round(project_actual / project_target * 100, 1) if project_target else 0
    recurring_ach = round(recurring_actual / recurring_target * 100, 1) if recurring_target else 0

    # By status
    c.execute("SELECT status, COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY status", (cur_year,))
    by_status = {r['status']: r['cnt'] for r in c.fetchall()}
    critical_count = by_status.get('Critical', 0)
    c.execute("SELECT COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    total_projects = c.fetchone()['cnt']

    # By owner
    c.execute("SELECT owner, SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY owner", (cur_year,))
    by_owner = {r['owner']: dict(r) for r in c.fetchall()}
    ama = by_owner.get('AMA', {}); eiw = by_owner.get('EIW', {})
    ama_target = float(ama.get('t', 0) or 0); ama_actual = float(ama.get('a', 0) or 0)
    eiw_target = float(eiw.get('t', 0) or 0); eiw_actual = float(eiw.get('a', 0) or 0)
    ama_ach = round(ama_actual / ama_target * 100, 1) if ama_target else 0
    eiw_ach = round(eiw_actual / eiw_target * 100, 1) if eiw_target else 0

    # Zero actual projects
    c.execute("""SELECT project_id, product, client, owner, revenue_target, type, risk_level, kategori
        FROM revenue_projects WHERE actual_revenue=0 AND revenue_target>0 AND is_active=1 AND tahun=%s
        ORDER BY revenue_target DESC""", (cur_year,))
    zero_projects = [_norm(dict(r)) for r in c.fetchall()]
    zero_count = len(zero_projects)
    zero_actual_value = sum(float(p['revenue_target'] or 0) for p in zero_projects)

    # Monthly trend dari invoice
    c.execute("""SELECT m.month_name, m.month_num,
        SUM(m.target) target,
        COALESCE((SELECT SUM(i.paid_amount) FROM invoices i
            JOIN revenue_projects rp ON i.project_id=rp.project_id
            WHERE rp.is_active=1 AND i.tahun=%s AND i.period=m.month_name), 0) as actual,
        SUM(CASE WHEN m.status='Achieve' THEN 1 ELSE 0 END) achieved,
        COUNT(*) total
        FROM revenue_monthly m JOIN revenue_projects p ON m.project_id=p.project_id
        WHERE p.is_active=1 AND p.tahun=%s GROUP BY m.month_num, m.month_name ORDER BY m.month_num""", (cur_year, cur_year))
    monthly_data_raw = [_norm(dict(r)) for r in c.fetchall()]

    # Invoice outstanding untuk tahun ini
    c.execute("""SELECT COUNT(*) cnt, SUM(invoice_amount - paid_amount) outstanding
        FROM invoices WHERE paid_amount < invoice_amount AND tahun=%s""", (cur_year,))
    inv_row = _norm(dict(c.fetchone()))
    outstanding_invoices = inv_row['cnt'] or 0
    outstanding_amount = inv_row['outstanding'] or 0

    # Action list — proyek perlu intervensi
    action_list = []
    for p in zero_projects[:8]:
        if float(p['revenue_target'] or 0) >= 1_000_000_000:
            issue_label = "🔴 Nol Realisasi"
        else:
            issue_label = "⭕ Belum Tagih"
        action = (
            "Jadwalkan war room segera, tetapkan milestone Q3 dan target penagihan termin pertama."
            if p['type'] == 'Termin' else
            "Cek status kontrak, kirim invoice, follow-up pembayaran."
        )
        action_list.append({**p, 'issue': 'zero', 'issue_label': issue_label, 'action': action})

    # Scorecard
    scorecard = [
        {"label": "Achievement YTD", "value": f"{ach_pct}%",
         "icon": "🔴", "status": "Kritis — Target 100%", "color": "text-red-400"},
        {"label": "Revenue Project", "value": f"{project_ach}%",
         "icon": "🔴", "status": "Hampir tidak bergerak", "color": "text-red-400"},
        {"label": "Revenue Recurring", "value": f"{recurring_ach}%",
         "icon": "🟡" if recurring_ach < 80 else "🟢",
         "status": "Perlu diakselerasi" if recurring_ach < 80 else "On track",
         "color": "text-yellow-300" if recurring_ach < 80 else "text-green-400"},
        {"label": "AMA Achievement", "value": f"{ama_ach}%",
         "icon": "🔴", "status": f"Rp {ama_actual/1e6:.0f}Jt / Rp {ama_target/1e9:.1f}M", "color": "text-red-400"},
        {"label": "EIW Achievement", "value": f"{eiw_ach}%",
         "icon": "🟡" if eiw_ach < 80 else "🟢",
         "status": f"Rp {eiw_actual/1e9:.2f}M / Rp {eiw_target/1e9:.1f}M",
         "color": "text-yellow-300" if eiw_ach < 80 else "text-green-400"},
        {"label": "Proyek Critical", "value": f"{critical_count}/{total_projects}",
         "icon": "🔴", "status": f"{round(critical_count/max(total_projects,1)*100)}% dari total", "color": "text-red-400"},
        {"label": "Gap Kejar",  "value": f"Rp {gap/1e9:.1f}M",
         "icon": "🔴", "status": f"Harus dicapai {cur_year}", "color": "text-red-400"},
        {"label": "Invoice Outstanding", "value": f"Rp {outstanding_amount/1e6:.0f}Jt",
         "icon": "🟡", "status": f"{outstanding_invoices} invoice belum lunas", "color": "text-yellow-300"},
    ]

    conn.close()
    return templates.TemplateResponse("rev_insights.html", build_ctx(request, user, {
        "total_target": float(total_target or 0), "total_actual": float(total_actual or 0),
        "ach_pct": ach_pct, "gap": gap,
        "project_target": project_target, "project_actual": project_actual, "project_ach": project_ach,
        "recurring_target": recurring_target, "recurring_actual": recurring_actual, "recurring_ach": recurring_ach,
        "critical_count": critical_count, "total_projects": total_projects,
        "ama_target": ama_target, "ama_actual": ama_actual, "ama_ach": ama_ach,
        "eiw_target": eiw_target, "eiw_actual": eiw_actual, "eiw_ach": eiw_ach,
        "zero_projects": zero_projects, "zero_count": zero_count, "zero_actual_value": zero_actual_value,
        "monthly_data": monthly_data_raw, "monthly_achieve": monthly_data_raw,
        "outstanding_invoices": outstanding_invoices, "outstanding_amount": outstanding_amount,
        "action_list": action_list[:10], "scorecard": scorecard,
        "cur_year": cur_year, "years": years,
    }))

# ── B1: REVENUE DASHBOARD ─────────────────────────────────────────────────────

@app.get("/revenue", response_class=HTMLResponse)
async def revenue_dashboard(request: Request, user=Depends(login_required), tahun: int = 0):
    guard(user, "rev_dashboard")
    sync_all_status()
    cur_year = tahun if tahun else date.today().year
    conn = get_conn()
    c = conn.cursor()

    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    s = _norm(dict(c.fetchone())); tt = s['t'] or 0; ta = s['a'] or 0
    ach_pct = round(float(ta or 0)/float(tt or 1)*100,1) if tt else 0

    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE kategori='Recurring' AND is_active=1 AND tahun=%s", (cur_year,))
    r = _norm(dict(c.fetchone())); rec_target = float(r['t'] or 0); rec_actual = float(r['a'] or 0)
    c.execute("SELECT SUM(revenue_target) t, SUM(actual_revenue) a FROM revenue_projects WHERE kategori='Project' AND is_active=1 AND tahun=%s", (cur_year,))
    r = _norm(dict(c.fetchone())); prj_target = r['t'] or 0; prj_actual = r['a'] or 0
    c.execute("SELECT status, COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY status", (cur_year,))
    by_status = {r['status']: r['cnt'] for r in c.fetchall()}
    c.execute("SELECT COUNT(*) cnt FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    total_projects = c.fetchone()['cnt']

    c.execute("""SELECT m.month_num, m.month_name, SUM(m.target) total_target, SUM(m.actual) total_actual
        FROM revenue_monthly m JOIN revenue_projects p ON m.project_id=p.project_id
        WHERE p.is_active=1 AND p.tahun=%s GROUP BY m.month_num, m.month_name ORDER BY m.month_num""", (cur_year,))
    monthly_trend = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("""SELECT * FROM revenue_projects WHERE status IN ('Critical','At Risk') AND is_active=1 AND tahun=%s
        ORDER BY risk_level DESC, revenue_target DESC LIMIT 10""", (cur_year,))
    critical = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT owner, COUNT(*) cnt, SUM(revenue_target) target, SUM(actual_revenue) actual FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY owner", (cur_year,))
    by_owner = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT product, COUNT(*) cnt, SUM(revenue_target) target, SUM(actual_revenue) actual FROM revenue_projects WHERE is_active=1 AND tahun=%s GROUP BY product ORDER BY target DESC LIMIT 8", (cur_year,))
    by_product = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]
    conn.close()

    return templates.TemplateResponse("rev_dashboard.html", build_ctx(request, user, {
        "total_target": float(tt or 0), "total_actual": float(ta or 0), "ach_pct": ach_pct,
        "rec_target": rec_target, "rec_actual": rec_actual,
        "prj_target": prj_target, "prj_actual": prj_actual,
        "by_status": by_status, "total_projects": total_projects,
        "monthly_trend": monthly_trend, "critical": critical,
        "by_owner": by_owner, "by_product": by_product,
        "cur_year": cur_year, "years": years,
    }))

# ── B2: REVENUE TRACKER (Read-Only Monitoring) ────────────────────────────────

@app.get("/revenue/tracker", response_class=HTMLResponse)
async def revenue_tracker(request: Request, user=Depends(login_required),
                          owner: str = "", kategori: str = "", status: str = "",
                          search: str = "", tahun: int = 0, month: int = 0):
    guard(user, "rev_tracker")
    sync_all_status()
    cur_year  = tahun if tahun else date.today().year
    cur_month = month if month else 0  # 0 = semua bulan
    conn = get_conn()
    c = conn.cursor()

    # Actual per project = SUM(paid_amount) dari invoices, difilter bulan jika dipilih
    if cur_month:
        month_name_en = MONTHS_EN[cur_month - 1]
        invoice_sub = f"""COALESCE((SELECT SUM(paid_amount) FROM invoices i
            WHERE i.project_id=p.project_id AND i.tahun={cur_year} AND i.period='{month_name_en}'),0)"""
        invoice_count_sub = f"""COALESCE((SELECT COUNT(*) FROM invoices i
            WHERE i.project_id=p.project_id AND i.tahun={cur_year} AND i.period='{month_name_en}'),0)"""
    else:
        invoice_sub = """COALESCE((SELECT SUM(paid_amount) FROM invoices i WHERE i.project_id=p.project_id),0)"""
        invoice_count_sub = """COALESCE((SELECT COUNT(*) FROM invoices i WHERE i.project_id=p.project_id),0)"""

    q = f"""SELECT p.*,
        {invoice_sub} as invoice_actual,
        {invoice_count_sub} as invoice_count,
        CASE WHEN p.revenue_target > 0 THEN
            CAST({invoice_sub} AS FLOAT) / p.revenue_target
        ELSE 0 END as achievement_pct
        FROM revenue_projects p WHERE p.is_active=1 AND p.tahun=%s"""
    params = [cur_year]
    if owner:    q += " AND p.owner=%s"; params.append(owner)
    if kategori: q += " AND p.kategori=%s"; params.append(kategori)
    if status:   q += " AND p.status=%s"; params.append(status)
    if search:   q += " AND (p.product LIKE %s OR p.client LIKE %s)"; params += [f"%{search}%"]*2
    c.execute(q + " ORDER BY CAST(REPLACE(p.project_id,'REV-','') AS INTEGER)", params)
    projects = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("SELECT DISTINCT owner FROM revenue_projects WHERE owner IS NOT NULL ORDER BY owner")
    owners = [r['owner'] for r in c.fetchall()]
    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]

    # Summary for cur_year (total invoice actual)
    if cur_month:
        month_name_en = MONTHS_EN[cur_month - 1]
        c.execute("""SELECT COALESCE(SUM(i.paid_amount),0) a FROM invoices i
            JOIN revenue_projects p ON i.project_id=p.project_id
            WHERE p.is_active=1 AND i.tahun=%s AND i.period=%s""", (cur_year, month_name_en))
        ta = float(c.fetchone()['a'] or 0)
    else:
        c.execute("SELECT COALESCE(SUM(actual_revenue),0) a FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
        ta = float(c.fetchone()['a'] or 0)
    c.execute("SELECT COALESCE(SUM(revenue_target),0) t FROM revenue_projects WHERE is_active=1 AND tahun=%s", (cur_year,))
    tt = float(c.fetchone()['t'] or 0)
    conn.close()
    return templates.TemplateResponse("rev_tracker.html", build_ctx(request, user, {
        "projects": projects, "owners": owners,
        "filter_owner": owner, "filter_kategori": kategori,
        "filter_status": status, "search": search,
        "cur_year": cur_year, "cur_month": cur_month, "years": years,
        "months": [(i+1, m) for i, m in enumerate(MONTHS_EN)],
        "total_target": float(tt or 0), "total_actual": float(ta or 0),
        "ach_pct": round(float(ta or 0)/float(tt or 1)*100,1) if tt else 0,
        "total_projects": len(projects),
    }))

# Route untuk insert dari Pipeline (Win → Revenue)
@app.post("/revenue/tracker/from-pipeline")
async def rev_from_pipeline(request: Request, user=Depends(login_required),
    lead_id: str = Form(""), nama_company: str = Form(""),
    product: str = Form(""), sales_owner: str = Form(""),
    deal_value: str = Form("0"), tahun: int = Form(2026),
    notes: str = Form("")):
    guard(user, "rev_tracker")
    pid = next_rev_id()
    target = float(deal_value or 0)
    status, risk = auto_status_risk(float(target or 0), 0.0, 'One Time', int(tahun or 2026))
    conn = get_conn()
    c = conn.cursor()
    c.execute("""INSERT INTO revenue_projects
        (project_id,lob,owner,product,client,kategori,type,tahun,
         revenue_target,actual_revenue,status,risk_level,notes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (project_id) DO NOTHING""",
        (pid, 'DCSS', sales_owner, product, nama_company,
         'Project', 'One Time', tahun,
         target, 0, status, risk,
         f"Dari Pipeline CRM: {lead_id}. {notes}"))
    conn.commit(); conn.close()
    return RedirectResponse(f"/revenue/tracker?tahun={tahun}", status_code=303)

# Route insert dari Win/Loss ke Revenue Tracker
@app.post("/revenue/tracker/from-winloss")
async def rev_from_winloss(request: Request, user=Depends(login_required),
    lead_id: str = Form(""), nama_company: str = Form(""),
    product: str = Form(""), sales_owner: str = Form(""),
    deal_value: str = Form("0"), tahun: int = Form(2026),
    type_: str = Form("One Time"), kategori: str = Form("Project"),
    notes: str = Form("")):
    guard(user, "rev_tracker")
    pid = next_rev_id()
    target = float(deal_value or 0)
    status, risk = auto_status_risk(float(target or 0), 0.0, type_, int(tahun or 2026))
    conn = get_conn()
    c = conn.cursor()
    c.execute("""INSERT INTO revenue_projects
        (project_id,lob,owner,product,client,kategori,type,tahun,
         revenue_target,actual_revenue,status,risk_level,notes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (project_id) DO NOTHING""",
        (pid, 'DCSS', sales_owner, product, nama_company,
         kategori, type_, tahun, target, 0, status, risk,
         f"Dari Win/Loss CRM: {lead_id}. {notes}"))
    conn.commit(); conn.close()
    return RedirectResponse(f"/revenue/tracker?tahun={tahun}", status_code=303)

# ── B3: MONTHLY MONITORING (Read-Only) ───────────────────────────────────────

@app.get("/revenue/monthly", response_class=HTMLResponse)
async def revenue_monthly(request: Request, user=Depends(login_required),
                          month: int = 0, tahun: int = 0, owner: str = "", kategori: str = ""):
    guard(user, "rev_monthly")
    cur_year  = tahun if tahun else date.today().year
    cur_month = month if month else date.today().month
    conn = get_conn()
    c = conn.cursor()

    q = """SELECT p.project_id, p.product, p.client, p.owner, p.kategori,
               p.revenue_target, p.status, p.risk_level,
               m.month_num, m.month_name, m.target,
               COALESCE((SELECT SUM(i.paid_amount) FROM invoices i
                         WHERE i.project_id=p.project_id
                         AND i.tahun=%s AND i.period=m.month_name), 0) as actual,
               m.status as month_status,
               COALESCE((SELECT SUM(i.paid_amount) FROM invoices i
                         WHERE i.project_id=p.project_id
                         AND i.tahun=%s AND i.period=m.month_name), 0) as invoice_actual_month
           FROM revenue_monthly m
           JOIN revenue_projects p ON m.project_id=p.project_id
           WHERE m.month_num=%s AND p.is_active=1 AND p.tahun=%s"""
    params = [cur_year, cur_year, cur_month, cur_year]
    if owner:    q += " AND p.owner=%s"; params.append(owner)
    if kategori: q += " AND p.kategori=%s"; params.append(kategori)
    c.execute(q + " ORDER BY CAST(REPLACE(p.project_id,'REV-','') AS INTEGER)", params)
    rows = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("""SELECT SUM(m.target) t, SUM(m.actual) a
        FROM revenue_monthly m JOIN revenue_projects p ON m.project_id=p.project_id
        WHERE m.month_num=%s AND p.is_active=1 AND p.tahun=%s""", (cur_month, cur_year))
    ms = _norm(dict(c.fetchone()))

    # Invoice actual bulan ini
    c.execute("""SELECT COALESCE(SUM(paid_amount),0) inv_actual
        FROM invoices WHERE tahun=%s AND period=%s""",
        (cur_year, MONTHS_EN[cur_month-1]))
    inv_actual_month = c.fetchone()['inv_actual']

    c.execute("SELECT DISTINCT owner FROM revenue_projects WHERE owner IS NOT NULL ORDER BY owner")
    owners = [r['owner'] for r in c.fetchall()]
    c.execute("SELECT DISTINCT tahun FROM revenue_projects ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]
    conn.close()

    return templates.TemplateResponse("rev_monthly.html", build_ctx(request, user, {
        "rows": rows, "cur_month": cur_month, "cur_year": cur_year,
        "month_name": MONTHS_EN[cur_month-1],
        "month_target": float(ms['t'] or 0), "month_actual": float(inv_actual_month or 0),
        "inv_actual_month": inv_actual_month,
        "owners": owners, "filter_owner": owner, "filter_kategori": kategori,
        "months": [(i+1, m) for i, m in enumerate(MONTHS_EN)],
        "years": years,
    }))

# ── B4: INVOICE & PAYMENT ─────────────────────────────────────────────────────

@app.get("/revenue/invoice", response_class=HTMLResponse)
async def invoice_list(request: Request, user=Depends(login_required),
                       period: str = "", status: str = "", search: str = "",
                       tahun: int = 0, project_id: str = ""):
    guard(user, "rev_invoice")
    cur_year = tahun if tahun else date.today().year
    conn = get_conn()
    c = conn.cursor()
    q = "SELECT i.*, p.product as proj_product, p.client as proj_client FROM invoices i LEFT JOIN revenue_projects p ON i.project_id=p.project_id WHERE i.tahun=%s"
    params = [cur_year]
    if period:     q += " AND i.period=%s"; params.append(period)
    if project_id: q += " AND i.project_id=%s"; params.append(project_id)
    if status == 'Lunas':   q += " AND i.paid_amount >= i.invoice_amount AND i.invoice_amount > 0"
    elif status == 'Belum': q += " AND (i.paid_amount < i.invoice_amount)"
    if search: q += " AND (i.client LIKE %s OR i.product LIKE %s OR i.invoice_no LIKE %s)"; params += [f"%{search}%"]*3
    c.execute(q + " ORDER BY i.invoice_date DESC", params)
    invoices = [_norm(dict(r)) for r in c.fetchall()]

    c.execute("SELECT SUM(invoice_amount) t, SUM(paid_amount) p FROM invoices WHERE tahun=%s", (cur_year,))
    totals = _norm(dict(c.fetchone()))
    c.execute("SELECT DISTINCT period FROM invoices WHERE period IS NOT NULL ORDER BY period")
    periods = [r['period'] for r in c.fetchall()]
    c.execute("SELECT project_id, product, client FROM revenue_projects WHERE is_active=1 ORDER BY project_id")
    rev_projects = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT DISTINCT tahun FROM invoices ORDER BY tahun DESC")
    years = [r['tahun'] for r in c.fetchall() if r['tahun']]

    for inv in invoices:
        if (inv['paid_amount'] or 0) >= (inv['invoice_amount'] or 0) and (inv['invoice_amount'] or 0) > 0:
            inv['display_status'] = 'Lunas'
        elif not inv['paid_date']:
            inv['display_status'] = 'Belum Dibayar'
        else:
            inv['display_status'] = 'Partial'

    return templates.TemplateResponse("rev_invoice.html", build_ctx(request, user, {
        "invoices": invoices, "periods": periods,
        "filter_period": period, "filter_status": status, "search": search,
        "filter_project": project_id,
        "total_invoice": totals['t'] or 0, "total_paid": totals['p'] or 0,
        "rev_projects": rev_projects, "cur_year": cur_year, "years": years,
    }))

@app.post("/revenue/invoice/new")
async def invoice_create(request: Request, user=Depends(login_required),
    project_id: str = Form(""), client: str = Form(""), product: str = Form(""),
    owner: str = Form(""), invoice_no: str = Form(""),
    invoice_date: str = Form(""), period: str = Form(""),
    invoice_amount: str = Form("0"), paid_amount: str = Form("0"),
    paid_date: str = Form(""), notes: str = Form(""),
    tahun: int = Form(0)):
    guard(user, "rev_invoice")
    # Auto-fill client/product dari project jika ada
    conn = get_conn()
    c = conn.cursor()
    yr = tahun if tahun else date.today().year
    if project_id:
        c.execute("SELECT product, client, owner FROM revenue_projects WHERE project_id=%s", (project_id,))
        row = c.fetchone()
        if row:
            if not product: product = row['product']
            if not client:  client  = row['client']
            if not owner:   owner   = row['owner'] or owner
    c.execute("""INSERT INTO invoices
        (project_id,lob,owner,product,client,invoice_no,invoice_date,period,
         invoice_amount,paid_amount,paid_date,notes,tahun)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (project_id or None, 'DCSS', owner, product, client, invoice_no,
         invoice_date or None, period,
         float(invoice_amount or 0), float(paid_amount or 0),
         paid_date or None, notes, yr))
    conn.commit()
    # Sync actual ke revenue_projects
    if project_id:
        c.execute("SELECT COALESCE(SUM(paid_amount),0) total FROM invoices WHERE project_id=%s", (project_id,))
        actual = c.fetchone()['total']
        c.execute("SELECT revenue_target, type, tahun FROM revenue_projects WHERE project_id=%s", (project_id,))
        p = c.fetchone()
        if p:
            st, rk = auto_status_risk(float(p['revenue_target'] or 0), float(actual or 0), p['type'], int(p['tahun'] or 2026))
            c.execute("UPDATE revenue_projects SET actual_revenue=%s, status=%s, risk_level=%s, updated_at=NOW() WHERE project_id=%s",
                         (actual, st, rk, project_id))
            conn.commit()
    conn.close()
    return RedirectResponse(f"/revenue/invoice?tahun={yr}", status_code=303)

@app.post("/revenue/invoice/{iid}/pay")
async def invoice_pay(iid: int, request: Request, user=Depends(login_required),
                      paid_amount: str = Form("0"), paid_date: str = Form("")):
    guard(user, "rev_invoice")
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT project_id FROM invoices WHERE id=%s", (iid,))
    row = c.fetchone()
    pid = row['project_id'] if row else None
    c.execute("UPDATE invoices SET paid_amount=%s, paid_date=%s WHERE id=%s",
                 (float(paid_amount or 0), paid_date or None, iid))
    conn.commit()
    # Re-sync revenue_projects
    if pid:
        sync_project_status(pid)
    conn.close()
    return RedirectResponse("/revenue/invoice", status_code=303)

# ── B5: KPI PROSPECTING ───────────────────────────────────────────────────────

@app.get("/revenue/kpi", response_class=HTMLResponse)
async def kpi_dashboard(request: Request, user=Depends(login_required), tahun: int = 0):
    guard(user, "rev_kpi")
    cur_year = tahun if tahun else date.today().year
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT * FROM kpi_prospecting ORDER BY sort_order")
    kpis = [_norm(dict(r)) for r in c.fetchall()]
    from collections import defaultdict
    grouped = defaultdict(list)
    for k in kpis:
        grouped[k['kpi_category']].append(k)
    conn.close()
    return templates.TemplateResponse("rev_kpi.html", build_ctx(request, user, {
        "kpis": kpis, "grouped": dict(grouped), "cur_year": cur_year,
    }))

@app.post("/revenue/kpi/{kid}/update")
async def kpi_update(kid: int, request: Request, user=Depends(login_required),
    q1_actual: str = Form("0"), q2_actual: str = Form("0"),
    q3_actual: str = Form("0"), q4_actual: str = Form("0")):
    guard(user, "rev_kpi")
    conn = get_conn()
    c = conn.cursor()
    c.execute("""UPDATE kpi_prospecting SET q1_actual=%s,q2_actual=%s,q3_actual=%s,q4_actual=%s WHERE id=%s""",
        (float(q1_actual or 0), float(q2_actual or 0), float(q3_actual or 0), float(q4_actual or 0), kid))
    conn.commit(); conn.close()
    return RedirectResponse("/revenue/kpi", status_code=303)

# ── B6: BUDGET MONITORING ─────────────────────────────────────────────────────

@app.get("/revenue/budget", response_class=HTMLResponse)
async def budget_monitoring(request: Request, user=Depends(login_required), tahun: int = 0):
    guard(user, "rev_budget")
    cur_year = tahun if tahun else date.today().year
    conn = get_conn()
    c = conn.cursor()
    c.execute("""SELECT perspektif_bsc, SUM(budget_amount) total_budget,
        SUM(actual_amount) total_actual, COUNT(*) items FROM budget_items GROUP BY perspektif_bsc""")
    by_bsc = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT * FROM budget_items ORDER BY perspektif_bsc, id")
    items = [_norm(dict(r)) for r in c.fetchall()]
    c.execute("SELECT SUM(budget_amount) t, SUM(actual_amount) a FROM budget_items")
    totals = _norm(dict(c.fetchone()))
    conn.close()
    return templates.TemplateResponse("rev_budget.html", build_ctx(request, user, {
        "by_bsc": by_bsc, "items": items,
        "total_budget": totals['t'] or 0, "total_actual": totals['a'] or 0,
        "cur_year": cur_year,
    }))

@app.post("/revenue/budget/{bid}/update")
async def budget_update(bid: int, request: Request, user=Depends(login_required),
                        actual_amount: str = Form("0"), notes: str = Form("")):
    guard(user, "rev_budget")
    conn = get_conn()
    c = conn.cursor()
    c.execute("UPDATE budget_items SET actual_amount=%s, notes=%s WHERE id=%s",
                 (float(actual_amount or 0), notes, bid))
    conn.commit(); conn.close()
    return RedirectResponse("/revenue/budget", status_code=303)

# ── FIELD ACTIVITY ───────────────────────────────────────────────────────────

@app.get("/field-activity", response_class=HTMLResponse)
async def field_activity_page(request: Request, user=Depends(login_required),
                               tgl: str = None):
    guard(user, "field_activity")
    today_str = tgl or str(date.today())
    conn = get_conn()
    c = conn.cursor()

    # Ambil semua aktivitas hari ini (atau tanggal filter)
    c.execute("""
        SELECT fa.*, u.nama as sales_nama
        FROM field_activity_logs fa
        LEFT JOIN users u ON fa.user_id = u.id
        WHERE DATE(fa.checked_in_at) = %s
        ORDER BY fa.checked_in_at DESC
    """, (today_str,))
    activities = [_norm(dict(r)) for r in c.fetchall()]

    # Summary per user
    c.execute("""
        SELECT u.nama, COUNT(*) as visits,
               SUM(CASE WHEN fa.checked_out_at IS NOT NULL THEN fa.duration_minutes ELSE 0 END) as total_menit
        FROM field_activity_logs fa
        LEFT JOIN users u ON fa.user_id = u.id
        WHERE DATE(fa.checked_in_at) = %s
        GROUP BY u.nama ORDER BY visits DESC
    """, (today_str,))
    summary = [_norm(dict(r)) for r in c.fetchall()]
    conn.close()

    return templates.TemplateResponse("field_activity.html", build_ctx(request, user, {
        "activities": activities, "summary": summary,
        "today_str": today_str, "filter_tgl": today_str,
    }))


# ── DAILY REPORT ──────────────────────────────────────────────────────────────

@app.get("/daily-report", response_class=HTMLResponse)
async def daily_report_page(request: Request, user=Depends(login_required),
                             month: str = None, user_id: str = None, status: str = None):
    guard(user, "daily_report")
    cur_month = month or date.today().strftime("%Y-%m")
    conn = get_conn()
    c = conn.cursor()

    where  = ["TO_CHAR(dr.report_date, 'YYYY-MM') = %s"]
    params = [cur_month]

    if user_id:
        where.append("dr.user_id = %s")
        params.append(int(user_id))
    if status:
        where.append("dr.status = %s")
        params.append(status)

    where_str = " AND ".join(where)
    c.execute(f"""
        SELECT dr.*, u.nama as sales_nama
        FROM daily_reports dr
        LEFT JOIN users u ON dr.user_id = u.id
        WHERE {where_str}
        ORDER BY dr.report_date DESC, u.nama
    """, params)
    reports = [_norm(dict(r)) for r in c.fetchall()]

    # Daftar sales untuk filter
    c.execute("SELECT id, nama FROM users WHERE role_id IN (2,3) ORDER BY nama")
    sales_list = [dict(r) for r in c.fetchall()]

    # Summary bulan ini
    c.execute("""
        SELECT
            COUNT(*) as total,
            COUNT(CASE WHEN status='sent' THEN 1 END) as terkirim,
            COUNT(CASE WHEN status='draft' THEN 1 END) as draft,
            COUNT(DISTINCT user_id) as sales_aktif
        FROM daily_reports
        WHERE TO_CHAR(report_date, 'YYYY-MM') = %s
    """, (cur_month,))
    summary = _norm(dict(c.fetchone()))
    conn.close()

    return templates.TemplateResponse("daily_report.html", build_ctx(request, user, {
        "reports": reports, "summary": summary, "sales_list": sales_list,
        "cur_month": cur_month, "filter_user_id": user_id or "",
        "filter_status": status or "",
    }))


# ── IMPORT EXCEL ──────────────────────────────────────────────────────────────

@app.get("/import", response_class=HTMLResponse)
async def import_page(request: Request, user=Depends(login_required)):
    guard(user, "import")
    return templates.TemplateResponse("import.html", build_ctx(request, user, {}))


@app.post("/import")
async def import_excel(request: Request, user=Depends(login_required),
                       mode: str = Form("append")):
    from fastapi import UploadFile, File
    import pandas as pd
    import io, traceback

    guard(user, "import")

    form = await request.form()
    file: UploadFile = form.get("file")
    mode = form.get("mode", "append")  # "append" | "replace"

    if not file or not file.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse("import.html", build_ctx(request, user, {
            "error": "Harap upload file Excel (.xlsx / .xls)"
        }))

    content = await file.read()
    xl = pd.ExcelFile(io.BytesIO(content))
    sheets = xl.sheet_names

    results = {}
    errors = []

    def clean(v, default=None):
        if v is None: return default
        s = str(v).strip()
        return None if s in ('nan', 'NaT', 'None', '') else s

    def numval(v, default=0.0):
        try:
            f = float(v)
            return 0.0 if (f != f) else f
        except: return default

    def datestr(v):
        if v is None: return None
        try:
            ts = pd.Timestamp(v)
            return ts.strftime('%Y-%m-%d') if not pd.isna(ts) else None
        except: return None

    conn = get_conn()
    c = conn.cursor()

    try:
        # ── 1. PIPELINE ─────────────────────────────────────────────────────
        pipeline_sheet = next((s for s in sheets if 'pipeline' in s.lower()), None)
        if pipeline_sheet:
            try:
                df = pd.read_excel(xl, sheet_name=pipeline_sheet, header=None)
                # Cari baris header (ada 'Nama Company' atau 'ID')
                header_row = 0
                for i, row in df.iterrows():
                    vals = [str(v).strip() for v in row.values]
                    if any(k in vals for k in ['Nama Company', 'ID', 'Lead ID']):
                        header_row = i
                        break
                df.columns = [str(c).strip() for c in df.iloc[header_row]]
                df = df.iloc[header_row + 1:].reset_index(drop=True)

                if mode == "replace":
                    c.execute("DELETE FROM leads")

                imported = 0
                for _, row in df.iterrows():
                    lid = clean(row.get('ID') or row.get('Lead ID') or row.get('lead_id'))
                    if not lid or not (lid.startswith('LD-') or lid.startswith('ld-')):
                        continue
                    nama = clean(row.get('Nama Company'))
                    if not nama:
                        continue

                    pv   = numval(row.get('Propose Value', 0))
                    dv   = numval(row.get('Deal Value (Rp)', 0) or row.get('Deal Value', 0))
                    prob = numval(row.get('Probability (%)', 0) or row.get('Probability', 0))
                    wv   = dv * (prob / 100) if dv else pv * (prob / 100)

                    c.execute("""
                        INSERT INTO leads
                          (lead_id,nama_company,product,contact_person,segmen,sub_segmen,
                           source,stage,prioritas,tgl_masuk,propose_value,deal_value,
                           probability,exp_close_date,weighted_value,sales_owner,
                           next_fu_date,last_fu_date,last_fu_notes,fu_count,days_in_stage,remarks)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (lead_id) DO UPDATE SET
                          nama_company=EXCLUDED.nama_company, product=EXCLUDED.product,
                          contact_person=EXCLUDED.contact_person, segmen=EXCLUDED.segmen,
                          sub_segmen=EXCLUDED.sub_segmen, source=EXCLUDED.source,
                          stage=EXCLUDED.stage, prioritas=EXCLUDED.prioritas,
                          tgl_masuk=EXCLUDED.tgl_masuk, propose_value=EXCLUDED.propose_value,
                          deal_value=EXCLUDED.deal_value, probability=EXCLUDED.probability,
                          exp_close_date=EXCLUDED.exp_close_date, weighted_value=EXCLUDED.weighted_value,
                          sales_owner=EXCLUDED.sales_owner, next_fu_date=EXCLUDED.next_fu_date,
                          last_fu_date=EXCLUDED.last_fu_date, last_fu_notes=EXCLUDED.last_fu_notes,
                          fu_count=EXCLUDED.fu_count, days_in_stage=EXCLUDED.days_in_stage,
                          remarks=EXCLUDED.remarks
                    """, (
                        lid, nama,
                        clean(row.get('Product')), clean(row.get('Contact Person')),
                        clean(row.get('Segmen')), clean(row.get('Sub-segmen')),
                        clean(row.get('Source')), clean(row.get('Stage'), 'New'),
                        clean(row.get('Prioritas'), 'Warm'),
                        datestr(row.get('Tgl Masuk')), pv, dv, prob,
                        datestr(row.get('Exp. Close Date') or row.get('Exp Close Date')),
                        wv, clean(row.get('Sales Owner')),
                        datestr(row.get('Next FU Date')), datestr(row.get('Last FU Date')),
                        clean(row.get('Last FU Notes')),
                        int(numval(row.get('FU Count', 0))),
                        int(numval(row.get('Days in Stage', 0))),
                        clean(row.get('Remarks')),
                    ))
                    imported += 1
                results['pipeline'] = imported
            except Exception as e:
                errors.append(f"Pipeline: {e}")

        # ── 2. REVENUE TRACKER ───────────────────────────────────────────────
        rev_sheet = next((s for s in sheets if 'revenue tracker' in s.lower() or s.lower() == 'revenue'), None)
        if rev_sheet:
            try:
                df = pd.read_excel(xl, sheet_name=rev_sheet, header=None)
                header_row = 0
                for i, row in df.iterrows():
                    if 'Project ID' in str(row.values) or 'Client' in str(row.values):
                        header_row = i
                        break
                df.columns = [str(v).strip() for v in df.iloc[header_row]]
                df = df.iloc[header_row + 1:].reset_index(drop=True)

                if mode == "replace":
                    c.execute("DELETE FROM revenue_monthly")
                    c.execute("DELETE FROM revenue_projects")

                imported = 0
                for _, row in df.iterrows():
                    pid_raw = clean(row.get('Project ID'))
                    if not pid_raw: continue
                    try: pid_int = int(float(pid_raw))
                    except: continue
                    project_id = f"REV-{pid_int:04d}"

                    rev_target = numval(row.get('Revenue Target', 0))
                    rev_actual = numval(row.get('Actual Revenue', 0))
                    tahun_raw  = clean(row.get('Tahun') or row.get('Year'))
                    tahun      = int(float(tahun_raw)) if tahun_raw else date.today().year
                    type_raw   = clean(row.get('Type'))
                    tgt_month  = datestr(row.get('Target Month'))
                    status, risk = auto_status_risk(rev_target, rev_actual, type_raw, tahun, tgt_month)

                    c.execute("""
                        INSERT INTO revenue_projects
                          (project_id,lob,owner,product,client,kategori,type,target_month,
                           revenue_target,actual_revenue,status,invoice_date,payment_date,
                           notes,risk_level,action_required,pic,tahun)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (project_id) DO UPDATE SET
                          lob=EXCLUDED.lob, owner=EXCLUDED.owner, product=EXCLUDED.product,
                          client=EXCLUDED.client, kategori=EXCLUDED.kategori, type=EXCLUDED.type,
                          target_month=EXCLUDED.target_month, revenue_target=EXCLUDED.revenue_target,
                          actual_revenue=EXCLUDED.actual_revenue, status=EXCLUDED.status,
                          invoice_date=EXCLUDED.invoice_date, payment_date=EXCLUDED.payment_date,
                          notes=EXCLUDED.notes, risk_level=EXCLUDED.risk_level,
                          action_required=EXCLUDED.action_required, pic=EXCLUDED.pic, tahun=EXCLUDED.tahun
                    """, (
                        project_id,
                        clean(row.get('LOB'), 'DCSS'),
                        clean(row.get('Owner')), clean(row.get('Product')),
                        clean(row.get('Client')), clean(row.get('Kategori')),
                        type_raw, tgt_month,
                        rev_target, rev_actual, status,
                        datestr(row.get('Invoice Date')),
                        datestr(row.get('Payment Date')),
                        clean(row.get('Notes')), risk,
                        clean(row.get('Action Required')),
                        clean(row.get('PIC')), tahun,
                    ))
                    imported += 1
                results['revenue'] = imported
            except Exception as e:
                errors.append(f"Revenue: {e}")

        # ── 3. INVOICES ──────────────────────────────────────────────────────
        inv_sheet = next((s for s in sheets if 'invoice' in s.lower()), None)
        if inv_sheet:
            try:
                df = pd.read_excel(xl, sheet_name=inv_sheet, header=0)
                df.columns = [str(c).strip() for c in df.columns]

                if mode == "replace":
                    c.execute("DELETE FROM invoices")

                imported = 0
                for _, row in df.iterrows():
                    pid_raw = clean(row.get('Project ID'))
                    if not pid_raw: continue
                    try: pid_int = int(float(pid_raw))
                    except: continue
                    project_id = f"REV-{pid_int:04d}"

                    inv_nos = str(row.get('Invoice No', '') or '').replace('\n', '; ').strip()
                    if inv_nos in ('nan', 'None', ''): inv_nos = None

                    c.execute("""
                        INSERT INTO invoices
                          (project_id,lob,owner,product,client,invoice_no,invoice_date,
                           period,invoice_amount,paid_amount,paid_date,notes)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        project_id,
                        clean(row.get('LOB'), 'DCSS'),
                        clean(row.get('Owner')), clean(row.get('Product')),
                        clean(row.get('Client')), inv_nos,
                        datestr(row.get('Invoice Date')),
                        clean(row.get('Period')),
                        numval(row.get('Invoice Amount', 0)),
                        numval(row.get('Paid Amount', 0)),
                        datestr(row.get('Paid Date')),
                        clean(row.get('NOTES') or row.get('Notes')),
                    ))
                    imported += 1
                results['invoice'] = imported
            except Exception as e:
                errors.append(f"Invoice: {e}")

        # ── 4. KPI PROSPECTING ───────────────────────────────────────────────
        kpi_sheet = next((s for s in sheets if 'kpi' in s.lower()), None)
        if kpi_sheet:
            try:
                df = pd.read_excel(xl, sheet_name=kpi_sheet, header=None)
                header_row = None
                for i, row in df.iterrows():
                    if 'KPI Category' in str(row.values) or 'KPI Name' in str(row.values):
                        header_row = i
                        break
                if header_row is not None:
                    df.columns = [str(v).strip() for v in df.iloc[header_row]]
                    df = df.iloc[header_row + 1:].reset_index(drop=True)

                    if mode == "replace":
                        c.execute("DELETE FROM kpi_prospecting")

                    current_cat = None
                    sort_n = 0
                    imported = 0
                    for _, row in df.iterrows():
                        cat  = clean(row.get('KPI Category'))
                        name = clean(row.get('KPI Name'))
                        if cat and not name:
                            current_cat = cat
                            continue
                        if not name: continue
                        if cat: current_cat = cat
                        sort_n += 1
                        c.execute("""
                            INSERT INTO kpi_prospecting
                              (kpi_category,kpi_name,unit,target_annual,
                               q1_target,q1_actual,q2_target,q2_actual,
                               q3_target,q3_actual,q4_target,q4_actual,sort_order)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            ON CONFLICT DO NOTHING
                        """, (
                            current_cat, name,
                            clean(row.get('Unit'), 'Count'),
                            numval(row.get('Target 2026') or row.get('Target Annual', 0)),
                            numval(row.get('Q1 Target', 0)), numval(row.get('Q1 Actual', 0)),
                            numval(row.get('Q2 Target', 0)), numval(row.get('Q2 Actual', 0)),
                            numval(row.get('Q3 Target', 0)), numval(row.get('Q3 Actual', 0)),
                            numval(row.get('Q4 Target', 0)), numval(row.get('Q4 Actual', 0)),
                            sort_n,
                        ))
                        imported += 1
                    results['kpi'] = imported
            except Exception as e:
                errors.append(f"KPI: {e}")

        conn.commit()

    except Exception as e:
        conn.rollback()
        errors.append(f"Error umum: {traceback.format_exc()}")
    finally:
        conn.close()

    return templates.TemplateResponse("import.html", build_ctx(request, user, {
        "results": results,
        "errors": errors,
        "filename": file.filename,
        "mode": mode,
    }))


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)
